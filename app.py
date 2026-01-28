#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Flask web application for VR-TS Bonepile Statistics (multi-sheet)
"""

from flask import Flask, render_template, jsonify, request, redirect, url_for, flash
import pandas as pd
import numpy as np
import os
from datetime import datetime, timedelta, date
import pytz
from werkzeug.utils import secure_filename
import socket
import re
import glob
from collections import defaultdict
import pickle
import hashlib
import json

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['CACHE_FOLDER'] = 'cache'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
USER_MAPPING_PATH = os.path.join(app.config['CACHE_FOLDER'], 'user_mapping.json')

# Create upload and cache folders if not exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['CACHE_FOLDER'], exist_ok=True)


class SchemaError(Exception):
    """Raised when user-provided mapping can't be applied to the input file."""

    def __init__(self, message, details=None):
        super().__init__(message)
        self.details = details or {}


DEFAULT_USER_MAPPING = {
    "bonepile": {
        "file_name": "NV_IGS_VR144_Bonepile.xlsx",
        "sheet_name": "VR-TS1",
        # Excel header row number (1-based). Row 2 => pandas header=1.
        "header_row_excel": 2,
        # Backward-compat (older configs used 0-based pandas header). Keep for migration only.
        "header_row": 1,
        # Canonical field -> Excel column header (string). These must match df.columns values.
        "columns": {
            "sn": "SN",
            "bp_duration": "bp_duration",
            "nv_disposition": "NV Disposition",
            "result": "Status",
            "pic": "PIC",
            "igs_action": "IGS Action",
            "igs_status": "IGS Status",
            "fail_time": "fail_time",
        },
    },
    "fa_work_log": {
        "file_name": "FA_Work_Log.xlsx",
        "sheet_name": "Log",
        # Excel row number (1-based) to start reading data rows (skip header)
        "start_row": 2,
        # Excel column index (0-based within openpyxl row tuple): Column B=1, Column C=2
        "sn_col_index": 1,
        "wo_col_index": 2,
    },
}


def load_user_mapping():
    """Load persisted user mapping (single source of truth)."""
    try:
        if os.path.exists(USER_MAPPING_PATH):
            with open(USER_MAPPING_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, dict):
                    # Migrate bonepile header row to 1-based Excel row if missing
                    try:
                        bp = data.get('bonepile', {})
                        if isinstance(bp, dict) and 'header_row_excel' not in bp:
                            # Old meaning: header_row is 0-based pandas header index
                            old = bp.get('header_row', DEFAULT_USER_MAPPING['bonepile']['header_row'])
                            bp['header_row_excel'] = int(old) + 1
                            data['bonepile'] = bp
                            # Persist migration so UI shows the correct value next time
                            save_user_mapping(data)
                    except Exception:
                        pass
                    # Migrate bonepile single-sheet mapping to per-sheet mapping
                    try:
                        bp = data.get('bonepile', {})
                        if isinstance(bp, dict):
                            has_new = isinstance(bp.get('sheets'), dict) or isinstance(bp.get('selected_sheets'), list)
                            if not has_new:
                                legacy_sheet = (bp.get('sheet_name') or DEFAULT_USER_MAPPING['bonepile']['sheet_name'] or '').strip()
                                legacy_header_row_excel = get_bonepile_header_row_excel(bp)
                                legacy_cols = bp.get('columns', DEFAULT_USER_MAPPING['bonepile']['columns']) or {}
                                bp['selected_sheets'] = [legacy_sheet] if legacy_sheet else []
                                bp['sheets'] = {}
                                if legacy_sheet:
                                    bp['sheets'][legacy_sheet] = {
                                        "header_row_excel": int(legacy_header_row_excel),
                                        "columns": legacy_cols,
                                    }
                                data['bonepile'] = bp
                                save_user_mapping(data)
                    except Exception:
                        pass
                    return data
    except Exception:
        pass
    # Fallback to defaults if file missing or invalid
    return json.loads(json.dumps(DEFAULT_USER_MAPPING))


def normalize_bonepile_mapping(bonepile_mapping):
    """
    Normalize bonepile mapping to support multi-sheet + per-sheet mappings.

    Canonical shape:
    {
      "file_name": str,
      "selected_sheets": [str, ...],
      "sheets": {
        "<sheet>": { "header_row_excel": int, "columns": { ... } }
      }
    }

    Backward-compatible with legacy keys: sheet_name, header_row_excel/header_row, columns.
    """
    default_bp = DEFAULT_USER_MAPPING.get('bonepile', {}) or {}
    bp = bonepile_mapping if isinstance(bonepile_mapping, dict) else {}

    file_name = (bp.get('file_name') or default_bp.get('file_name') or '').strip()
    legacy_sheet = (bp.get('sheet_name') or default_bp.get('sheet_name') or '').strip()
    legacy_header_row_excel = get_bonepile_header_row_excel(bp)
    legacy_cols = bp.get('columns', default_bp.get('columns')) or {}

    selected = bp.get('selected_sheets')
    if isinstance(selected, str):
        selected = [selected]
    if not isinstance(selected, list):
        selected = []
    selected_sheets = [str(s).strip() for s in selected if str(s).strip()]

    sheets_cfg = bp.get('sheets')
    if not isinstance(sheets_cfg, dict):
        sheets_cfg = {}

    # Ensure legacy sheet exists in per-sheet configs (as a baseline)
    if legacy_sheet and legacy_sheet not in sheets_cfg:
        sheets_cfg[legacy_sheet] = {
            "header_row_excel": int(legacy_header_row_excel),
            "columns": json.loads(json.dumps(legacy_cols)),
        }

    # Default selection
    if not selected_sheets:
        selected_sheets = [legacy_sheet] if legacy_sheet else []

    # Ensure each selected sheet has a usable config
    for s in list(selected_sheets):
        if s not in sheets_cfg or not isinstance(sheets_cfg.get(s), dict):
            sheets_cfg[s] = {}
        sc = sheets_cfg.get(s) or {}
        if sc.get('header_row_excel') is None:
            sc['header_row_excel'] = int(legacy_header_row_excel)
        else:
            try:
                sc['header_row_excel'] = max(1, int(sc.get('header_row_excel')))
            except Exception:
                sc['header_row_excel'] = int(legacy_header_row_excel)
        cols = sc.get('columns')
        if not isinstance(cols, dict) or not cols:
            cols = json.loads(json.dumps(legacy_cols))
        sc['columns'] = cols
        sheets_cfg[s] = sc

    return {
        "file_name": file_name,
        "selected_sheets": selected_sheets,
        "sheets": sheets_cfg,
    }


def get_bonepile_header_row_excel(bonepile_mapping):
    """
    Return Excel row number (1-based) for the header row.
    Supports both new key (header_row_excel) and old key (header_row, 0-based).
    """
    if not isinstance(bonepile_mapping, dict):
        return int(DEFAULT_USER_MAPPING['bonepile']['header_row_excel'])
    if 'header_row_excel' in bonepile_mapping and bonepile_mapping.get('header_row_excel') is not None:
        try:
            return max(1, int(bonepile_mapping.get('header_row_excel')))
        except Exception:
            return int(DEFAULT_USER_MAPPING['bonepile']['header_row_excel'])
    # Backward compat
    try:
        return max(1, int(bonepile_mapping.get('header_row', DEFAULT_USER_MAPPING['bonepile']['header_row'])) + 1)
    except Exception:
        return int(DEFAULT_USER_MAPPING['bonepile']['header_row_excel'])


def bonepile_header_row_0_based(bonepile_mapping):
    """Convert Excel header row number (1-based) to pandas header index (0-based)."""
    return max(0, get_bonepile_header_row_excel(bonepile_mapping) - 1)


def save_user_mapping(mapping):
    """Persist user mapping to disk."""
    with open(USER_MAPPING_PATH, 'w', encoding='utf-8') as f:
        json.dump(mapping, f, indent=2, ensure_ascii=False)


def resolve_uploaded_or_local_path(file_name):
    """Prefer uploads/<file_name>, else project root <file_name>."""
    upload_path = os.path.join(app.config['UPLOAD_FOLDER'], file_name)
    if os.path.exists(upload_path):
        return upload_path
    if os.path.exists(file_name):
        return file_name
    return None


def list_excel_sheets(file_path):
    """Return list of sheet names from an Excel file."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    except Exception:
        return []


def get_pandas_columns(file_path, sheet_name, header_row_0_based):
    """
    Get columns via pandas (fallback when openpyxl header read fails).
    Uses nrows=0 so it only reads headers.
    """
    try:
        df0 = pd.read_excel(file_path, sheet_name=sheet_name, header=int(header_row_0_based), nrows=0)
        cols = []
        for idx, c in enumerate(list(df0.columns)):
            cols.append({"index": idx, "name": '' if c is None else str(c).strip()})
        return cols
    except Exception:
        return []


def get_excel_sheet_max_column(file_path, sheet_name):
    """Return max_column for a sheet (capped for UI)."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return 0
        ws = wb[sheet_name]
        max_col = int(ws.max_column or 0)
        wb.close()
        return max_col
    except Exception:
        return 0


def get_excel_header_values(file_path, sheet_name, header_row_0_based):
    """
    Read one row from Excel to build a list of column options with index.
    header_row_0_based: pandas-style header row index (0-based).
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        wb = load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[sheet_name]
        excel_row = int(header_row_0_based) + 1  # openpyxl is 1-based
        out = []
        for idx, cell in enumerate(ws[excel_row]):
            v = cell.value
            name = '' if v is None else str(v).strip()
            out.append({
                "value": f"__idx__{idx}",
                "name": name,
                "index": idx,
                "label": f"{get_column_letter(idx+1)} ({idx}) - {name if name else '(blank)'}",
            })
        wb.close()
        return out
    except Exception:
        return []


def resolve_df_column(df, configured_name):
    """Resolve a configured column name against df.columns (case/whitespace tolerant)."""
    if configured_name is None:
        return None
    if df is None:
        return None
    configured_str = str(configured_name).strip()

    # Preferred mapping: by index sentinel (__idx__N)
    if configured_str.startswith('__idx__'):
        try:
            idx = int(configured_str.replace('__idx__', '').strip())
            if 0 <= idx < len(df.columns):
                return df.columns[idx]
            return None
        except Exception:
            return None

    if configured_str in df.columns:
        return configured_str
    configured_norm = configured_str.lower().strip()
    for col in df.columns:
        if str(col).strip().lower() == configured_norm:
            return col
    return None

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_local_ip():
    """Get local IP address"""
    try:
        # Connect to a remote address to determine local IP
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return "127.0.0.1"

# Station order: FLA > FLB > AST > FTS > FCT > RIN > NVL
# Note: NVL is placed after RIN for TS2 flow.
STATION_ORDER = ['FLA', 'FLB', 'AST', 'FTS', 'FCT', 'RIN', 'NVL']

def sort_stations(stations):
    """
    Sort stations according to custom order: FLA > FLB > AST > FTS > FCT > RIN
    Stations not in the order list will be sorted alphabetically at the end
    """
    def get_station_order(station):
        try:
            return STATION_ORDER.index(station)
        except ValueError:
            # Station not in order list, put it at the end
            return len(STATION_ORDER) + ord(station[0]) if station else 999
    
    if isinstance(stations, (list, tuple)):
        return sorted(stations, key=get_station_order)
    elif isinstance(stations, set):
        return sorted(list(stations), key=get_station_order)
    else:
        return stations


def get_pass_station_for_part_number(part_number):
    """
    Determine which station counts as the final PASS station for a given part number.

    Rules:
    - If part number ends with/contains TS2 => final pass station is NVL
    - Otherwise (TS1 and future TS3+ unless overridden) => final pass station is FCT
    """
    try:
        pn = '' if part_number is None else str(part_number).strip().upper()
    except Exception:
        pn = ''
    # Examples: "675-24109-0000-TS2" or "...-TS2"
    if 'TS2' in pn:
        return 'NVL'
    return 'FCT'


def is_final_pass_test_entry(test_entry, cutoff_date=None):
    """
    Return True if this test entry should be considered the "final pass" for the SN.

    If part_number is present, use TS-based station rule.
    If part_number is missing/Unknown, fall back to date-based legacy rule:
    - Before 2026-01-01: PASS at RIN
    - After  2026-01-01: PASS at FCT
    """
    if not isinstance(test_entry, dict):
        return False
    if test_entry.get('status') != 'P':
        return False

    station = str(test_entry.get('station') or '').strip().upper()
    part_number = test_entry.get('part_number')
    pn = '' if part_number is None else str(part_number).strip()

    if pn and pn.lower() != 'unknown':
        return station == get_pass_station_for_part_number(pn)

    # Legacy fallback when part_number is not available
    if cutoff_date is None:
        cutoff_date = datetime(2026, 1, 1).date()
    test_date = test_entry.get('date')
    if isinstance(test_date, datetime):
        test_date = test_date.date()
    elif isinstance(test_date, pd.Timestamp):
        test_date = test_date.to_pydatetime().date()
    elif isinstance(test_date, str):
        try:
            test_date = datetime.strptime(test_date, '%Y-%m-%d').date()
        except Exception:
            test_date = None
    if not isinstance(test_date, date):
        test_date = cutoff_date

    if test_date >= cutoff_date:
        return station == 'FCT'
    return station == 'RIN'

# Function to check valid SN
def is_valid_sn(sn):
    if pd.isna(sn):
        return False
    if isinstance(sn, (int, float)):
        sn_str = str(int(sn))
    else:
        sn_str = str(sn).strip().replace('.0', '')
    return sn_str.startswith('183') and len(sn_str) == 13 and sn_str.isdigit()

# Function to check if status is in process
def is_in_process(status):
    if pd.isna(status):
        return False
    status_str = str(status).lower()
    in_process_keywords = ['waiting', 'testing', 'in process', 'in progress']
    return any(keyword in status_str for keyword in in_process_keywords)

# Function to check waiting for material
def is_waiting_for_material(text):
    if pd.isna(text):
        return False
    text_str = str(text).lower()
    # Check for material/component waits (expanded keywords)
    material_keywords = [
        'waiting for material',
        'waiting for cx9',
        'waiting for strata',
        'waiting for new material',
        'waiting for new strata',
        'waiting for bbay',
        'waiting for bf4',
    ]
    return any(keyword in text_str for keyword in material_keywords)

# Function to normalize SN format
def normalize_sn(sn):
    """Normalize SN format to ensure consistent matching"""
    if pd.isna(sn) or sn == '':
        return ''
    
    # Handle different input types
    if isinstance(sn, (int, float)):
        # If it's a number, convert to int first (removes .0), then to string
        # This preserves the full number without scientific notation
        sn_str = str(int(sn))
    else:
        # Convert to string and strip whitespace
        sn_str = str(sn).strip()
    
    # Remove .0 suffix if exists (from float conversion)
    if sn_str.endswith('.0'):
        sn_str = sn_str[:-2]
    
    # Extract only digits (remove any non-digit characters)
    sn_digits = ''.join(filter(str.isdigit, sn_str))
    
    # Validate SN format: should start with 18 and be 13 digits
    if sn_digits.startswith('18') and len(sn_digits) == 13:
        return sn_digits
    
    # If not valid format but has digits, try to pad or fix
    # Sometimes Excel might read SN as number and lose leading zeros
    if sn_digits.startswith('18') and len(sn_digits) < 13:
        # Pad with zeros if it starts with 18 but is shorter
        # This shouldn't happen for valid SNs, but handle it anyway
        sn_digits = sn_digits.ljust(13, '0')
        if len(sn_digits) == 13:
            return sn_digits
    
    # If still not valid, return the digit-only string anyway (might be a different format)
    return sn_digits if sn_digits else sn_str

# Function to normalize WO format
def normalize_wo(wo_str):
    """Normalize WO format: '000007016682-1' -> '7016682'"""
    if pd.isna(wo_str) or not wo_str:
        return ''
    wo_str = str(wo_str).strip()
    # Remove prefix "00000" if exists
    if wo_str.startswith('00000'):
        wo_str = wo_str[5:]
    # Remove suffix "-1" or "-X" if exists
    if '-' in wo_str:
        wo_str = wo_str.split('-')[0]
    return wo_str

# Function to parse test filename from daily test analysis
def extract_part_number_from_filename(filename):
    """
    Extract part number từ filename và normalize (chỉ lấy phần sau PB-xxxxx_)
    Pattern: IGSJ_PB-6306_675-24109-0000-TS1_1835225000016_F_RIN_20251230T161507Z
    Returns normalized part number (chỉ phần sau PB-xxxxx_) hoặc 'Unknown' nếu không tìm thấy
    """
    name = filename.replace('.zip', '')
    
    # Pattern 1: PB-XXXX_XXX-XXXXX-XXXX-TS<NUM> (full pattern with PB prefix)
    pattern1 = r'PB-\d+_(\d+-\d+-\d+-TS\d+)'
    match1 = re.search(pattern1, name)
    if match1:
        return match1.group(1)  # Chỉ lấy phần sau PB-xxxxx_
    
    # Pattern 2: PB-XXXX_XXX-XXXXX-XXXX (without TS1, with PB prefix)
    pattern2 = r'PB-\d+_(\d+-\d+-\d+)'
    match2 = re.search(pattern2, name)
    if match2:
        return match2.group(1)  # Chỉ lấy phần sau PB-xxxxx_
    
    # Pattern 3: XXX-XXXXX-XXXX-TS<NUM> (without PB prefix - giữ nguyên)
    pattern3 = r'(\d+-\d+-\d+-TS\d+)'
    match3 = re.search(pattern3, name)
    if match3:
        return match3.group(1)
    
    # Pattern 4: XXX-XXXXX-XXXX (without PB prefix and TS1 - giữ nguyên)
    pattern4 = r'(\d+-\d+-\d+)'
    match4 = re.search(pattern4, name)
    if match4:
        return match4.group(1)
    
    return 'Unknown'

def parse_test_filename(filename):
    """Parse filename để extract: SN, Status (F/P), Station, Part Number"""
    name = filename.replace('.zip', '')
    
    # Extract part number
    part_number = extract_part_number_from_filename(filename)
    
    # Pattern 1: _SN_Status_Station_ (ví dụ: _1835225000016_F_RIN_)
    pattern1 = r'_(\d{10,})_([FP])_([A-Z0-9]+)_'
    match1 = re.search(pattern1, name)
    if match1:
        sn = match1.group(1)
        status = match1.group(2)  # F hoặc P
        station = match1.group(3)  # RIN, FLA, etc.
        # Validate SN: phải bắt đầu bằng 18 và có 13 digits
        if sn.startswith('18') and len(sn) == 13:
            return (sn, status, station, part_number)
    
    # Pattern 2: Tìm SN 18xxxxxxxxxxx ở bất kỳ đâu, sau đó tìm _Status_Station_
    sn_match = re.search(r'(18\d{11})', name)
    if sn_match:
        sn = sn_match.group(1)
        after_sn = name[name.find(sn) + len(sn):]
        pattern2 = r'_([FP])_([A-Z0-9]+)_'
        match2 = re.search(pattern2, after_sn)
        if match2:
            status = match2.group(1)
            station = match2.group(2)
            return (sn, status, station, part_number)
    
    return None

# Function to get current date in California timezone
def get_current_ca_date():
    """Get current date in California timezone (PST/PDT)"""
    try:
        ca_tz = pytz.timezone('America/Los_Angeles')
        ca_now = datetime.now(ca_tz)
        return ca_now.date()
    except:
        # Fallback: assume UTC-8 (PST) if pytz not available
        utc_now = datetime.utcnow()
        ca_now = utc_now - timedelta(hours=8)  # PST = UTC-8
        return ca_now.date()

# Function to get cache file path for a date
def get_cache_file_path(date):
    """Get cache file path for a specific date"""
    date_str = date.strftime('%Y-%m-%d')
    cache_filename = f"daily_test_{date_str}.pkl"
    return os.path.join(app.config['CACHE_FOLDER'], cache_filename)

# Function to cleanup old cache files (older than 90 days)
def cleanup_old_cache():
    """Delete cache files older than 90 days"""
    try:
        cache_dir = app.config['CACHE_FOLDER']
        if not os.path.exists(cache_dir):
            return
        
        current_ca_date = get_current_ca_date()
        cutoff_date = current_ca_date - timedelta(days=90)
        
        deleted_count = 0
        for filename in os.listdir(cache_dir):
            if filename.startswith('daily_test_') and filename.endswith('.pkl'):
                # Extract date from filename: daily_test_YYYY-MM-DD.pkl
                try:
                    date_str = filename.replace('daily_test_', '').replace('.pkl', '')
                    file_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    
                    if file_date < cutoff_date:
                        file_path = os.path.join(cache_dir, filename)
                        os.remove(file_path)
                        deleted_count += 1
                except (ValueError, OSError):
                    # Skip files with invalid format or deletion errors
                    continue
        
        if deleted_count > 0:
            print(f"[CACHE] Cleaned up {deleted_count} old cache file(s) (older than 90 days)", flush=True)
    except Exception as e:
        print(f"[CACHE] Error cleaning up old cache: {e}", flush=True)

# Function to convert nested defaultdict and set to dict/list for pickle
def convert_to_dict(obj):
    """Recursively convert defaultdict and set to dict/list for pickle compatibility"""
    if isinstance(obj, defaultdict):
        return {k: convert_to_dict(v) for k, v in obj.items()}
    elif isinstance(obj, dict):
        return {k: convert_to_dict(v) for k, v in obj.items()}
    elif isinstance(obj, (list, tuple)):
        return type(obj)(convert_to_dict(item) for item in obj)
    elif isinstance(obj, set):
        return list(obj)  # Convert set to list for pickle
    else:
        return obj

# Function to load cached data for a date
def load_cached_data(date):
    """Load cached data for a specific date if exists (only for past dates, not today)"""
    # Don't cache current date - always load fresh data for today
    current_ca_date = get_current_ca_date()
    if isinstance(date, datetime):
        date_only = date.date()
    else:
        date_only = date
    
    if date_only >= current_ca_date:
        return None  # Don't use cache for today or future dates
    
    cache_file = get_cache_file_path(date)
    if os.path.exists(cache_file):
        try:
            # Check file size first
            file_size = os.path.getsize(cache_file)
            if file_size == 0:
                return None
            
            with open(cache_file, 'rb') as f:
                data = pickle.load(f)
                # Validate that cached data has required keys
                required_keys = ['all_sns', 'sn_test_info', 'sn_pass_rin', 'station_stats']
                if not isinstance(data, dict) or not all(key in data for key in required_keys):
                    return None
                return data
        except (EOFError, pickle.UnpicklingError, Exception):
            # If cache is corrupt, return None to force reload
            return None
    return None

# Function to save data to cache for a date
def save_to_cache(date, data):
    """Save data to cache for a specific date (only for past dates, not today)"""
    # Don't cache current date - always load fresh data for today
    current_ca_date = get_current_ca_date()
    if isinstance(date, datetime):
        date_only = date.date()
    else:
        date_only = date
    
    if date_only >= current_ca_date:
        return  # Don't save cache for today or future dates
    
    cache_file = get_cache_file_path(date)
    try:
        # Ensure cache directory exists
        cache_dir = os.path.dirname(cache_file)
        os.makedirs(cache_dir, exist_ok=True)
        
        # Use temporary file and rename for atomic write
        temp_file = cache_file + '.tmp'
        with open(temp_file, 'wb') as f:
            pickle.dump(data, f)
        
        # Atomic rename (works on Windows if target exists, may need to delete first)
        if os.path.exists(cache_file):
            os.remove(cache_file)
        os.rename(temp_file, cache_file)
    except Exception:
        # Clean up temp file if exists
        temp_file = cache_file + '.tmp'
        if os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except:
                pass
        pass

# Function to load daily test data from network path with caching
def load_daily_test_data(start_date, end_date):
    """
    Load test data từ network path cho date range với caching
    Returns: dict với test results grouped by WO, station, part number
    """
    base_path = r"\\10.16.137.111\Oberon\L10"
    
    all_sns = set()
    sn_test_info = defaultdict(list)  # {sn: [(date, status, station, filename, wo, part_number), ...]}
    sn_pass_rin = set()  # SNs đã PASS ở RIN
    station_stats = defaultdict(lambda: {'pass': 0, 'fail': 0})  # {station: {'pass': X, 'fail': Y}}
    wo_station_stats = defaultdict(lambda: defaultdict(lambda: {'pass': 0, 'fail': 0}))  # {wo: {station: {'pass': X, 'fail': Y}}}
    part_station_stats = defaultdict(lambda: defaultdict(lambda: {'pass': 0, 'fail': 0}))  # {part_number: {station: {'pass': X, 'fail': Y}}}
    part_stats = defaultdict(lambda: {'pass': 0, 'fail': 0})  # {part_number: {'pass': X, 'fail': Y}}
    
    # Load SN -> WO mapping
    sn_wo_mapping = load_fa_work_log()
    
    # Track part numbers per SN (một SN có thể có nhiều part numbers nếu test nhiều lần)
    sn_part_numbers = defaultdict(set)  # {sn: {part_number1, part_number2, ...}}
    
    # Iterate through date range with caching
    current_date = start_date
    while current_date <= end_date:
        # Try to load from cache first
        cached_data = load_cached_data(current_date)
        
        if cached_data:
            # Use cached data
            # Convert lists back to sets if needed (for compatibility with new cache format)
            cached_sns_list = cached_data.get('all_sns', [])
            cached_sns = set(cached_sns_list) if isinstance(cached_sns_list, list) else cached_sns_list
            cached_test_info = cached_data.get('sn_test_info', {})
            cached_pass_rin_list = cached_data.get('sn_pass_rin', [])
            cached_pass_rin = set(cached_pass_rin_list) if isinstance(cached_pass_rin_list, list) else cached_pass_rin_list
            cached_station_stats = cached_data.get('station_stats', {})
            cached_wo_station_stats = cached_data.get('wo_station_stats', {})
            cached_part_station_stats = cached_data.get('part_station_stats', {})
            cached_part_stats = cached_data.get('part_stats', {})
            cached_sn_part_numbers = cached_data.get('sn_part_numbers', {})
            
            # Merge cached data into main data structures
            all_sns.update(cached_sns)
            sn_pass_rin.update(cached_pass_rin)
            
            for sn, test_list in cached_test_info.items():
                if sn not in sn_test_info:
                    sn_test_info[sn] = []
                # Update WO for cached entries using current mapping
                for test_entry in test_list:
                    # Re-lookup WO from current mapping to ensure it's up-to-date
                    sn_normalized = normalize_sn(sn)
                    wo_from_mapping = sn_wo_mapping.get(sn_normalized, 'No WO')
                    wo_from_mapping = normalize_wo(wo_from_mapping) if wo_from_mapping != 'No WO' else wo_from_mapping
                    # Update WO in test_entry
                    test_entry['wo'] = wo_from_mapping
                sn_test_info[sn].extend(test_list)
            
            for station, stats in cached_station_stats.items():
                station_stats[station]['pass'] += stats.get('pass', 0)
                station_stats[station]['fail'] += stats.get('fail', 0)
            
            for wo, wo_stations in cached_wo_station_stats.items():
                for station, stats in wo_stations.items():
                    wo_station_stats[wo][station]['pass'] += stats.get('pass', 0)
                    wo_station_stats[wo][station]['fail'] += stats.get('fail', 0)
            
            for part, part_stations in cached_part_station_stats.items():
                for station, stats in part_stations.items():
                    part_station_stats[part][station]['pass'] += stats.get('pass', 0)
                    part_station_stats[part][station]['fail'] += stats.get('fail', 0)
            
            for part, stats in cached_part_stats.items():
                part_stats[part]['pass'] += stats.get('pass', 0)
                part_stats[part]['fail'] += stats.get('fail', 0)
            
            for sn, part_nums in cached_sn_part_numbers.items():
                if isinstance(part_nums, list):
                    sn_part_numbers[sn].update(part_nums)
                else:
                    sn_part_numbers[sn].update([part_nums])
        else:
            # Load from network path and cache it
            year = current_date.strftime("%Y")
            month = current_date.strftime("%m")
            day = current_date.strftime("%d")
            dir_path = os.path.join(base_path, year, month, day)
            
            # Data structures for this date only
            date_sns = set()
            date_test_info = defaultdict(list)
            date_pass_rin = set()
            date_station_stats = defaultdict(lambda: {'pass': 0, 'fail': 0})
            date_wo_station_stats = defaultdict(lambda: defaultdict(lambda: {'pass': 0, 'fail': 0}))
            date_part_station_stats = defaultdict(lambda: defaultdict(lambda: {'pass': 0, 'fail': 0}))
            date_part_stats = defaultdict(lambda: {'pass': 0, 'fail': 0})
            date_sn_part_numbers = defaultdict(set)
            
            # Check if network path is accessible
            try:
                if os.path.isdir(dir_path):
                    zip_files = glob.glob(os.path.join(dir_path, "**", "*.zip"), recursive=True)
                    
                    for file_path in zip_files:
                        filename = os.path.basename(file_path)
                        parsed = parse_test_filename(filename)
                        
                        if parsed:
                            sn, status, station, part_number = parsed
                            # Normalize SN to ensure matching with mapping
                            sn_normalized = normalize_sn(sn)
                            date_sns.add(sn_normalized)
                            all_sns.add(sn_normalized)
                            wo = sn_wo_mapping.get(sn_normalized, 'No WO')
                            
                            wo = normalize_wo(wo) if wo != 'No WO' else wo
                            
                            # Track part number for this SN (use normalized SN for consistency)
                            date_sn_part_numbers[sn_normalized].add(part_number)
                            sn_part_numbers[sn_normalized].add(part_number)
                            
                            test_entry = {
                                'date': current_date,
                                'status': status,
                                'station': station,
                                'filename': filename,
                                'wo': wo,
                                'part_number': part_number
                            }
                            
                            date_test_info[sn_normalized].append(test_entry)
                            if sn_normalized not in sn_test_info:
                                sn_test_info[sn_normalized] = []
                            sn_test_info[sn_normalized].append(test_entry)
                            
                            if status == 'F':  # Fail
                                station_stats[station]['fail'] += 1
                                wo_station_stats[wo][station]['fail'] += 1
                                part_station_stats[part_number][station]['fail'] += 1
                                part_stats[part_number]['fail'] += 1
                                
                                date_station_stats[station]['fail'] += 1
                                date_wo_station_stats[wo][station]['fail'] += 1
                                date_part_station_stats[part_number][station]['fail'] += 1
                                date_part_stats[part_number]['fail'] += 1
                            elif status == 'P':  # Pass
                                station_stats[station]['pass'] += 1
                                wo_station_stats[wo][station]['pass'] += 1
                                part_station_stats[part_number][station]['pass'] += 1
                                part_stats[part_number]['pass'] += 1
                                
                                date_station_stats[station]['pass'] += 1
                                date_wo_station_stats[wo][station]['pass'] += 1
                                date_part_station_stats[part_number][station]['pass'] += 1
                                date_part_stats[part_number]['pass'] += 1
                                
                                # Determine PASS based on part number (TS1/TS2) and station.
                                # Keep legacy fallback when part_number is missing.
                                cutoff_date = datetime(2026, 1, 1).date()
                                if is_final_pass_test_entry(test_entry, cutoff_date=cutoff_date):
                                    date_pass_rin.add(sn_normalized)
                                    sn_pass_rin.add(sn_normalized)
            except (OSError, PermissionError):
                # Network path not accessible, skip this date
                pass
            
            # Cache the data for this date (only if we processed files)
            if date_sns or date_test_info:
                # Convert nested defaultdict and set to dict/list for pickle compatibility
                cache_data = {
                    'all_sns': convert_to_dict(date_sns),  # Convert set to list
                    'sn_test_info': convert_to_dict(date_test_info),  # Convert nested defaultdict to dict
                    'sn_pass_rin': convert_to_dict(date_pass_rin),  # Convert set to list
                    'station_stats': convert_to_dict(date_station_stats),
                    'wo_station_stats': convert_to_dict(date_wo_station_stats),  # Convert nested defaultdict
                    'part_station_stats': convert_to_dict(date_part_station_stats),  # Convert nested defaultdict
                    'part_stats': convert_to_dict(date_part_stats),
                    'sn_part_numbers': convert_to_dict(date_sn_part_numbers),
                    'cached_date': current_date
                }
                save_to_cache(current_date, cache_data)
        
        current_date += timedelta(days=1)
    
    # Calculate totals by part number (count unique SNs per part number)
    part_tray_stats = defaultdict(lambda: {'pass': 0, 'fail': 0, 'total': 0})
    for sn in all_sns:
        is_pass = sn in sn_pass_rin
        # Một SN có thể có nhiều part numbers, đếm cho mỗi part number
        for part_num in sn_part_numbers[sn]:
            part_tray_stats[part_num]['total'] += 1
            if is_pass:
                part_tray_stats[part_num]['pass'] += 1
            else:
                part_tray_stats[part_num]['fail'] += 1
    
    # Calculate totals
    total_trays = len(all_sns)
    total_pass = len(sn_pass_rin)
    total_fail = total_trays - total_pass
    
    return {
        'all_sns': all_sns,
        'sn_test_info': dict(sn_test_info),
        'sn_pass_rin': sn_pass_rin,
        'station_stats': dict(station_stats),
        'wo_station_stats': dict(wo_station_stats),
        'part_station_stats': dict(part_station_stats),
        'part_stats': dict(part_stats),
        'part_tray_stats': dict(part_tray_stats),
        'sn_part_numbers': {k: list(v) for k, v in sn_part_numbers.items()},
        'total_trays': total_trays,
        'total_pass': total_pass,
        'total_fail': total_fail,
        'sn_wo_mapping': sn_wo_mapping
    }

# Function to parse date from text (format: MM/DD or M/D)
def parse_date_from_text(text):
    """Parse date from text like '12/24', '12/29', '1/5' etc."""
    if pd.isna(text):
        return None
    text_str = str(text).strip()
    # Pattern: MM/DD or M/D (with optional year)
    pattern = r'(\d{1,2})/(\d{1,2})(?:\s|$|:)'
    match = re.search(pattern, text_str)
    if match:
        month = int(match.group(1))
        day = int(match.group(2))
        # Assume current year or previous year if month > current month
        current_year = datetime.now().year
        try:
            date = datetime(current_year, month, day)
            if date > datetime.now():
                date = datetime(current_year - 1, month, day)
            return date
        except ValueError:
            return None
    return None

# Function to parse all dispositions from text
def parse_dispositions_from_text(text):
    """Parse all dispositions from text, return list of (date, description) tuples"""
    if pd.isna(text):
        return []
    text_str = str(text)
    dispositions = []
    # Improved pattern to match various formats:
    # - "12/24: description"
    # - "12/24 description" (without colon)
    # - "12/26 9pm: description"
    # - "12/29: [Name] description"
    # Pattern matches: MM/DD (optional time) (optional colon) description
    pattern = r'(\d{1,2})/(\d{1,2})(?:\s+\d+[ap]m)?\s*:?\s*([^\n\r]+?)(?=\s*\d{1,2}/|\Z)'
    matches = re.finditer(pattern, text_str, re.MULTILINE | re.IGNORECASE | re.DOTALL)
    for match in matches:
        month = int(match.group(1))
        day = int(match.group(2))
        description = match.group(3).strip()
        # Remove trailing colons, extra spaces, and newlines
        description = description.rstrip(':').strip().replace('\n', ' ').replace('\r', ' ')
        # Clean up multiple spaces
        description = ' '.join(description.split())
        if not description:
            continue
        current_year = datetime.now().year
        try:
            # Cutoff date: December 1, 2025
            cutoff_date = datetime(2025, 12, 1)
            
            # Try current year first
            date = datetime(current_year, month, day)
            # If date is in the future, use previous year
            if date > datetime.now():
                date = datetime(current_year - 1, month, day)
            
            # Only add if date >= December 1, 2025
            if date >= cutoff_date:
                dispositions.append((date, description))
        except ValueError:
            continue
    return dispositions


def parse_nv_dispositions_mmdd_colon(text):
    """
    Parse NV Disposition text where each disposition is in format:
      MM/DD : blabla
    One cell can contain multiple dispositions.

    Returns list of (date, description) tuples.
    """
    if pd.isna(text):
        return []
    text_str = str(text)
    dispositions = []

    # Strictly require ":" after the date, per new definition
    # Supports optional whitespace around ":".
    pattern = r'(\d{1,2})/(\d{1,2})\s*:\s*([^\n\r]+?)(?=\s*\d{1,2}/\d{1,2}\s*:|\Z)'
    matches = re.finditer(pattern, text_str, re.MULTILINE | re.IGNORECASE | re.DOTALL)

    current_year = datetime.now().year
    for match in matches:
        try:
            month = int(match.group(1))
            day = int(match.group(2))
        except Exception:
            continue

        description = (match.group(3) or '').strip()
        description = description.rstrip(':').strip().replace('\n', ' ').replace('\r', ' ')
        description = ' '.join(description.split())

        # Keep behavior consistent with previous parsing: if parsed date is in the future,
        # treat it as previous year.
        try:
            disp_date = datetime(current_year, month, day)
            if disp_date > datetime.now():
                disp_date = datetime(current_year - 1, month, day)
        except ValueError:
            continue

        dispositions.append((disp_date, description))

    return dispositions


def get_latest_entry_from_date_desc_list(entries):
    """
    Given a list of (datetime, description) tuples, return the latest (date, description).
    If multiple entries have the same date, prefer the later entry in the original list.
    Returns (None, '') when empty.
    """
    if not entries:
        return None, ''
    try:
        _, (d, desc) = max(
            enumerate(entries),
            key=lambda t: ((t[1][0] or datetime.min), t[0]),
        )
        return d, desc
    except Exception:
        return None, ''

# Function to get latest date from dispositions
def get_latest_date_from_dispositions(dispositions):
    """Get the latest date from a list of (date, description) tuples"""
    if not dispositions:
        return None
    dates = [d[0] for d in dispositions if d[0] is not None]
    if not dates:
        return None
    return max(dates)

# Helper function to get column name with fallback to index
def get_column_name(df, column_name, fallback_index=None):
    """
    Get column name from dataframe with fallback to index.
    Tries exact match first, then partial match, then fallback index.
    
    Args:
        df: pandas DataFrame
        column_name: Name of the column to search for (case-insensitive, trimmed)
        fallback_index: Column index (0-based) to use if column name not found
    
    Returns:
        Column name (string) if found, None if not found and no valid fallback
    """
    if df is None or df.empty:
        return None
    
    # Try to find column by exact name match (case-insensitive, trimmed)
    column_name_lower = str(column_name).strip().lower()
    for col in df.columns:
        if str(col).strip().lower() == column_name_lower:
            return col
    
    # Try partial match for special cases like "IGS Status" or "IGS Action"
    # This handles cases where column might be named slightly differently
    if 'igs' in column_name_lower:
        # For IGS columns, try to find any column containing both keywords
        keywords = column_name_lower.split()
        if len(keywords) >= 2:
            for col in df.columns:
                col_lower = str(col).strip().lower()
                # Check if column contains all keywords
                if all(keyword in col_lower for keyword in keywords):
                    return col
    
    # Fallback to index if provided
    if fallback_index is not None and fallback_index < len(df.columns):
        fallback_col = df.columns[fallback_index]
        # Verify the fallback column exists
        if fallback_col in df.columns:
            return fallback_col
    
    # Return None if no match found and no valid fallback
    return None

# Helper function to safely get value from row
def safe_get_row_value(row, column_name, default=''):
    """
    Safely get value from pandas Series (row) with fallback to default.
    
    Args:
        row: pandas Series (row from iterrows())
        column_name: Column name or index to access
        default: Default value if column doesn't exist or is NaN
    
    Returns:
        Value from row or default
    """
    if column_name is None:
        return default
    
    try:
        if column_name in row.index:
            value = row[column_name]
            return value if pd.notna(value) else default
        else:
            return default
    except (KeyError, IndexError, TypeError):
        return default

# Function to load FA Work Log and create SN -> WO mapping
def load_fa_work_log():
    """Load FA_Work_Log.xlsx and create SN -> WO mapping"""
    sn_wo_mapping = {}

    mapping = load_user_mapping().get('fa_work_log', {})
    file_name = mapping.get('file_name', DEFAULT_USER_MAPPING['fa_work_log']['file_name'])
    sheet_name = mapping.get('sheet_name', DEFAULT_USER_MAPPING['fa_work_log']['sheet_name'])
    start_row = int(mapping.get('start_row', DEFAULT_USER_MAPPING['fa_work_log']['start_row']))
    sn_col_index = int(mapping.get('sn_col_index', DEFAULT_USER_MAPPING['fa_work_log']['sn_col_index']))
    wo_col_index = int(mapping.get('wo_col_index', DEFAULT_USER_MAPPING['fa_work_log']['wo_col_index']))

    fa_work_log_path = resolve_uploaded_or_local_path(file_name)
    if not fa_work_log_path:
        return sn_wo_mapping
    
    try:
        # Read Excel file directly using openpyxl to preserve SN precision
        # This is critical for large numbers like 1830126000016 which can lose precision if read as float
        from openpyxl import load_workbook
        
        # Load workbook - read as string to preserve exact format
        wb = load_workbook(fa_work_log_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise SchemaError(
                f"FA Work Log mapping error: sheet '{sheet_name}' not found. Please update mapping in Upload.",
                details={"file": file_name, "available_sheets": list(wb.sheetnames)},
            )
        ws = wb[sheet_name]
        
        loaded_count = 0
        skipped_count = 0
        
        # Read rows starting from configured row (1-based)
        for row in ws.iter_rows(min_row=max(1, start_row), values_only=True):
            if not row:
                continue

            if sn_col_index >= len(row) or wo_col_index >= len(row):
                continue

            sn_value = row[sn_col_index]
            wo_value = row[wo_col_index]
                
            if sn_value is not None and wo_value is not None:
                # Convert SN to string - handle both numeric and string types
                # For large integers stored as float in Excel, preserve precision
                if isinstance(sn_value, float):
                    # Check if it's a whole number (integer stored as float)
                    if sn_value.is_integer():
                        # Convert to int first to avoid scientific notation, then to string
                        sn_original = str(int(sn_value))
                    else:
                        # Should not happen for SNs, but handle it
                        sn_original = f"{sn_value:.0f}"
                elif isinstance(sn_value, int):
                    sn_original = str(sn_value)
                else:
                    sn_original = str(sn_value).strip() if sn_value else ''
                
                wo_original = str(wo_value).strip() if wo_value else ''
                
                if sn_original and sn_original.lower() != 'nan' and wo_original and wo_original.lower() != 'nan':
                    sn_str = normalize_sn(sn_original)  # Normalize SN format
                    wo_str = normalize_wo(wo_original)  # Normalize WO format
                    
                    if sn_str and len(sn_str) == 13 and sn_str.startswith('18'):
                        sn_wo_mapping[sn_str] = wo_str
                        loaded_count += 1
                    else:
                        skipped_count += 1
        
        wb.close()
    except SchemaError:
        raise
    except Exception as e:
        # Log error for debugging
        import traceback
        print(f"Error loading FA Work Log: {e}")
        traceback.print_exc()
        pass
    
    return sn_wo_mapping

# Function to load Bonepile list with fail_time mapping
def load_bonepile_list():
    """Load Bonepile Excel and return dict mapping SN -> fail_time (datetime or None)."""
    bonepile_fail_time = {}  # {sn: fail_time (datetime|None)}

    raw_mapping = load_user_mapping().get('bonepile', {})
    bp = normalize_bonepile_mapping(raw_mapping)
    file_name = bp.get('file_name', DEFAULT_USER_MAPPING['bonepile']['file_name'])
    selected_sheets = bp.get('selected_sheets') or []
    sheets_cfg = bp.get('sheets') or {}

    bonepile_file = resolve_uploaded_or_local_path(file_name)
    if not bonepile_file:
        return bonepile_fail_time
    
    if not selected_sheets:
        raise SchemaError(
            "Bonepile mapping error: no sheets selected. Please update mapping in Upload.",
            details={"file": file_name, "available_sheets": list_excel_sheets(bonepile_file)},
        )

    try:
        for sheet_name in selected_sheets:
            sc = sheets_cfg.get(sheet_name, {}) if isinstance(sheets_cfg, dict) else {}
            header_row_excel = int(sc.get('header_row_excel') or DEFAULT_USER_MAPPING['bonepile']['header_row_excel'])
            header_row = max(0, header_row_excel - 1)
            columns = sc.get('columns', {}) if isinstance(sc.get('columns', {}), dict) else {}
            try:
                df = pd.read_excel(bonepile_file, sheet_name=sheet_name, header=header_row)
            except Exception as e:
                raise SchemaError(
                    f"Bonepile mapping error: failed to read sheet '{sheet_name}' (header row={header_row_excel}). Please update mapping in Upload. ({e})",
                    details={"file": file_name, "sheet": sheet_name, "header_row_excel": header_row_excel, "header_row": header_row},
                )

            sn_col = resolve_df_column(df, columns.get('sn'))
            fail_time_col = resolve_df_column(df, columns.get('fail_time'))

            if not sn_col:
                raise SchemaError(
                    f"Bonepile mapping error: SN column not found for sheet '{sheet_name}'. Please update mapping in Upload.",
                    details={
                        "file": file_name,
                        "sheet": sheet_name,
                        "header_row_excel": header_row_excel,
                        "header_row": header_row,
                        "expected": {"sn": columns.get('sn')},
                        "available_columns": [str(c) for c in list(df.columns)],
                    },
                )
            
            # Remove duplicate header
            if len(df) > 0 and sn_col in df.columns:
                try:
                    if str(df.iloc[0][sn_col]).strip().lower() == 'sn':
                        df = df.iloc[1:].reset_index(drop=True)
                except Exception:
                    pass
            
            # Filter valid SNs
            valid_sn_records = df[df[sn_col].apply(is_valid_sn)].copy()
            
            # Get SN -> fail_time mapping
            for idx, row in valid_sn_records.iterrows():
                sn = row[sn_col]
                if isinstance(sn, (int, float)):
                    sn_str = str(int(sn))
                else:
                    sn_str = str(sn).strip().replace('.0', '')
                
                if not sn_str:
                    continue
                
                # Get fail_time (optional)
                fail_time = row.get(fail_time_col) if fail_time_col else None
                
                # Parse fail_time to datetime
                fail_time_dt = None
                if pd.notna(fail_time):
                    try:
                        # Try to parse as datetime
                        if isinstance(fail_time, datetime):
                            fail_time_dt = fail_time
                        elif isinstance(fail_time, str):
                            # Try various date formats
                            for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S']:
                                try:
                                    fail_time_dt = datetime.strptime(fail_time.strip(), fmt)
                                    break
                                except ValueError:
                                    continue
                        elif isinstance(fail_time, (int, float)):
                            # Excel serial date number
                            try:
                                fail_time_dt = pd.to_datetime(fail_time, origin='1899-12-30', unit='D')
                                if isinstance(fail_time_dt, pd.Timestamp):
                                    fail_time_dt = fail_time_dt.to_pydatetime()
                            except Exception:
                                pass
                    except Exception:
                        pass
                
                # Merge strategy: prefer a non-null fail_time; if both present, keep the earliest
                if sn_str in bonepile_fail_time:
                    existing = bonepile_fail_time.get(sn_str)
                    if existing is None and fail_time_dt is not None:
                        bonepile_fail_time[sn_str] = fail_time_dt
                    elif existing is not None and fail_time_dt is not None:
                        bonepile_fail_time[sn_str] = min(existing, fail_time_dt)
                else:
                    bonepile_fail_time[sn_str] = fail_time_dt
    except SchemaError:
        raise
    except Exception as e:
        print(f"[ERROR] load_bonepile_list: {e}", flush=True)
        import traceback
        print(traceback.format_exc(), flush=True)
        pass
    
    return bonepile_fail_time

# Function to parse timestamp from filename
def parse_timestamp_from_filename(filename):
    """
    Parse timestamp từ filename: 20251230T161507Z
    Format: YYYYMMDDTHHMMSSZ
    Note: Timestamp trong filename là local time (PST/PDT), không phải UTC
    Returns: datetime object (localized to CA timezone) hoặc None nếu không parse được
    """
    try:
        # Pattern: YYYYMMDDTHHMMSSZ
        pattern = r'(\d{8})T(\d{6})Z'
        match = re.search(pattern, filename)
        if match:
            date_str = match.group(1)  # 20260112
            time_str = match.group(2)    # 232647
            dt_str = f"{date_str}T{time_str}"
            # Parse as naive datetime first
            dt_naive = datetime.strptime(dt_str, "%Y%m%dT%H%M%S")
            # Localize to California timezone (treat as local time, not UTC)
            ca_tz = pytz.timezone('America/Los_Angeles')
            dt_ca = ca_tz.localize(dt_naive)
            return dt_ca
    except Exception as e:
        pass
    return None

# Function to convert datetime to California timezone
def convert_to_ca_time(dt):
    """
    Convert datetime to California timezone
    If already timezone-aware, return as is (assume it's already in CA timezone)
    If naive, assume it's already in CA timezone and localize it
    """
    if dt is None:
        return None
    
    # If already timezone-aware, return as is (assume it's already CA time)
    if hasattr(dt, 'tzinfo') and dt.tzinfo:
        return dt
    
    # If naive, assume it's already in CA timezone and localize it
    try:
        ca_tz = pytz.timezone('America/Los_Angeles')
        return ca_tz.localize(dt)
    except:
        return dt

# Function to filter test entries by time range
def filter_by_datetime_range(test_entries, start_datetime, end_datetime):
    """
    Filter test entries by datetime range
    start_datetime: datetime object (timezone-aware or naive)
    end_datetime: datetime object (timezone-aware or naive)
    """
    filtered = []
    
    # Ensure both datetimes are timezone-aware (CA timezone)
    try:
        ca_tz = pytz.timezone('America/Los_Angeles')
        if isinstance(start_datetime, str):
            # Parse string format: "YYYY-MM-DD HH:MM"
            start_datetime = datetime.strptime(start_datetime, '%Y-%m-%d %H:%M')
        if isinstance(end_datetime, str):
            end_datetime = datetime.strptime(end_datetime, '%Y-%m-%d %H:%M')
        
        # Localize if naive
        if not hasattr(start_datetime, 'tzinfo') or not start_datetime.tzinfo:
            start_time = ca_tz.localize(start_datetime)
        else:
            start_time = start_datetime.astimezone(ca_tz) if start_datetime.tzinfo else ca_tz.localize(start_datetime)
            
        if not hasattr(end_datetime, 'tzinfo') or not end_datetime.tzinfo:
            end_time = ca_tz.localize(end_datetime)
        else:
            end_time = end_datetime.astimezone(ca_tz) if end_datetime.tzinfo else ca_tz.localize(end_datetime)
    except Exception as e:
        # If parsing fails, return empty list
        return []
    
    # Get date range for fallback (when test_time_ca is not available)
    start_date = start_time.date()
    end_date = end_time.date()
    
    for entry in test_entries:
        test_time = entry.get('test_time_ca')
        if test_time:
            try:
                # Ensure test_time is timezone-aware
                if not hasattr(test_time, 'tzinfo') or not test_time.tzinfo:
                    # Make it timezone-aware using CA timezone
                    try:
                        ca_tz = pytz.timezone('America/Los_Angeles')
                        test_time = ca_tz.localize(test_time)
                    except:
                        # If localization fails, skip this entry
                        continue
                
                # Compare times (both should be timezone-aware now)
                if start_time <= test_time <= end_time:
                    filtered.append(entry)
            except Exception as e:
                # Skip entries with invalid time
                continue
        else:
            # Entry không có test_time_ca - fallback: dùng date từ entry
            # Include if entry date is within the date range
            entry_date = entry.get('date')
            if entry_date:
                # Convert entry_date to date if needed
                if isinstance(entry_date, datetime):
                    entry_date = entry_date.date()
                elif isinstance(entry_date, str):
                    entry_date = datetime.strptime(entry_date, '%Y-%m-%d').date()
                elif isinstance(entry_date, pd.Timestamp):
                    entry_date = entry_date.to_pydatetime().date()
                
                # Include if entry date is within range (inclusive)
                if start_date <= entry_date <= end_date:
                    filtered.append(entry)
    
    return filtered

# Function to load hourly report data
def load_hourly_report_data(start_datetime, end_datetime):
    """
    Load hourly report data for selected datetime range
    start_datetime: datetime object or string "YYYY-MM-DD HH:MM"
    end_datetime: datetime object or string "YYYY-MM-DD HH:MM"
    """
    base_path = r"\\10.16.137.111\Oberon\L10"
    
    # Parse datetime strings if needed
    if isinstance(start_datetime, str):
        start_datetime = datetime.strptime(start_datetime, '%Y-%m-%d %H:%M')
    if isinstance(end_datetime, str):
        end_datetime = datetime.strptime(end_datetime, '%Y-%m-%d %H:%M')
    
    # Get date range
    start_date = start_datetime.date() if isinstance(start_datetime, datetime) else start_datetime
    end_date = end_datetime.date() if isinstance(end_datetime, datetime) else end_datetime
    
    # Search dates: search from 1 day before start_date to 1 day after end_date
    # to ensure we don't miss any data (data might be in different folders)
    dates_to_search = []
    current_date = start_date - timedelta(days=1)
    while current_date <= end_date + timedelta(days=1):
        dates_to_search.append(current_date)
        current_date += timedelta(days=1)
    
    # Load bonepile list
    bonepile_fail_time = load_bonepile_list()
    bonepile_sns = set(bonepile_fail_time.keys())
    
    # Load SN -> WO mapping
    sn_wo_mapping = load_fa_work_log()
    
    # Collect all test entries
    all_test_entries = []
    total_files = 0
    parsed_files = 0
    files_with_timestamp = 0
    
    for search_date in dates_to_search:
        year = search_date.strftime("%Y")
        month = search_date.strftime("%m")
        day = search_date.strftime("%d")
        dir_path = os.path.join(base_path, year, month, day)
        
        try:
            if not os.path.isdir(dir_path):
                continue
            
            # Find all zip files
            zip_files = glob.glob(os.path.join(dir_path, "**", "*.zip"), recursive=True)
            total_files += len(zip_files)
            
            for file_path in zip_files:
                try:
                    filename = os.path.basename(file_path)
                    parsed = parse_test_filename(filename)
                    
                    if parsed:
                        parsed_files += 1
                        sn, status, station, part_number = parsed
                        sn_normalized = normalize_sn(sn)
                        
                        # Parse timestamp from filename
                        utc_dt = parse_timestamp_from_filename(filename)
                        test_time_ca = None
                        if utc_dt:
                            test_time_ca = convert_to_ca_time(utc_dt)
                            files_with_timestamp += 1
                        
                        wo = sn_wo_mapping.get(sn_normalized, 'No WO')
                        wo = normalize_wo(wo) if wo != 'No WO' else wo
                        
                        test_entry = {
                            'sn': sn_normalized,
                            'status': status,
                            'station': station,
                            'part_number': part_number,
                            'filename': filename,
                            'wo': wo,
                            'test_time_ca': test_time_ca,
                            'date': search_date
                        }
                        
                        all_test_entries.append(test_entry)
                except Exception as e:
                    # Skip files that can't be parsed
                    continue
        except Exception as e:
            # Skip directories that can't be accessed
            continue
    
    # Filter by datetime range
    filtered_entries = filter_by_datetime_range(all_test_entries, start_datetime, end_datetime)
    
    # Group by SN
    sn_data = defaultdict(lambda: {
        'tests': [],
        'part_numbers': set(),
        'stations': set(),
        'bonepile': False,
        'pass_fail': 'FAIL'
    })
    
    for entry in filtered_entries:
        sn = entry['sn']
        sn_data[sn]['tests'].append(entry)
        sn_data[sn]['part_numbers'].add(entry['part_number'])
        sn_data[sn]['stations'].add(entry['station'])
        sn_data[sn]['bonepile'] = (sn in bonepile_sns)
    
    # Determine pass/fail for each SN and track last pass time
    cutoff_date = datetime(2026, 1, 1).date()
    for sn, details in sn_data.items():
        tests = details['tests']
        if not tests:
            details['pass_fail'] = 'FAIL'
            details['last_pass_time'] = None
            continue
        
        # Check pass/fail based on rule:
        # - TS1: final pass at FCT
        # - TS2: final pass at NVL
        # - Fallback (no part_number): legacy date-based rule
        is_pass = False
        last_pass_time = None
        
        for test in tests:
            if is_final_pass_test_entry(test, cutoff_date=cutoff_date):
                is_pass = True
                # Track the latest pass time
                test_time = test.get('test_time_ca')
                if test_time:
                    if last_pass_time is None or test_time > last_pass_time:
                        last_pass_time = test_time
        
        details['pass_fail'] = 'PASS' if is_pass else 'FAIL'
        details['last_pass_time'] = last_pass_time
    
    # Convert sets to lists for JSON serialization
    for sn, details in sn_data.items():
        if isinstance(details.get('part_numbers'), set):
            details['part_numbers'] = sorted(list(details['part_numbers']))
        if isinstance(details.get('stations'), set):
            details['stations'] = sort_stations(list(details['stations']))
        elif isinstance(details.get('stations'), list):
            details['stations'] = sort_stations(details['stations'])
        # Convert datetime to string for JSON serialization
        if details.get('last_pass_time') and isinstance(details['last_pass_time'], datetime):
            details['last_pass_time'] = details['last_pass_time'].isoformat()
    
    # Calculate statistics
    all_sns = set(sn_data.keys())
    bonepile_sns_in_data = {sn for sn in all_sns if sn_data[sn]['bonepile']}
    igs_sns_in_data = all_sns - bonepile_sns_in_data
    
    all_pass_sns = {sn for sn in all_sns if sn_data[sn]['pass_fail'] == 'PASS'}
    all_fail_sns = all_sns - all_pass_sns
    
    bonepile_pass_sns = {sn for sn in bonepile_sns_in_data if sn_data[sn]['pass_fail'] == 'PASS'}
    bonepile_fail_sns = bonepile_sns_in_data - bonepile_pass_sns
    
    igs_pass_sns = {sn for sn in igs_sns_in_data if sn_data[sn]['pass_fail'] == 'PASS'}
    igs_fail_sns = igs_sns_in_data - igs_pass_sns
    
    stats = {
        'all': {
            'total_sns': len(all_sns),
            'pass_count': len(all_pass_sns),
            'fail_count': len(all_fail_sns),
            'bonepile': len(bonepile_sns_in_data),
            'fresh': len(igs_sns_in_data),
            'pass_rate': (len(all_pass_sns) / len(all_sns) * 100) if all_sns else 0
        },
        'bonepile': {
            'total_sns': len(bonepile_sns_in_data),
            'pass_count': len(bonepile_pass_sns),
            'fail_count': len(bonepile_fail_sns),
            'pass_rate': (len(bonepile_pass_sns) / len(bonepile_sns_in_data) * 100) if bonepile_sns_in_data else 0
        },
        'igs': {
            'total_sns': len(igs_sns_in_data),
            'pass_count': len(igs_pass_sns),
            'fail_count': len(igs_fail_sns),
            'pass_rate': (len(igs_pass_sns) / len(igs_sns_in_data) * 100) if igs_sns_in_data else 0
        }
    }
    
    # Return with 'statistics' key for consistency
    return {
        'statistics': stats,
        'sn_details': sn_data
    }

# Load data
def load_data(filename=None):
    user_mapping = load_user_mapping()
    bonepile_mapping = user_mapping.get('bonepile', {})
    default_bonepile_mapping = DEFAULT_USER_MAPPING['bonepile']

    if filename is None:
        # Prefer the configured Bonepile file
        configured_file_name = bonepile_mapping.get('file_name', default_bonepile_mapping['file_name'])
        excel_file = resolve_uploaded_or_local_path(configured_file_name)

        # Fallback: use most recent uploaded Excel if configured file isn't found
        if not excel_file:
            upload_folder = app.config['UPLOAD_FOLDER']
            excel_files = [f for f in os.listdir(upload_folder) if f.endswith(('.xlsx', '.xls'))]
            if excel_files:
                excel_file = os.path.join(upload_folder, sorted(excel_files)[-1])
            else:
                return None
    else:
        excel_file = filename
    
    if not os.path.exists(excel_file):
        return None
    
    bp = normalize_bonepile_mapping(bonepile_mapping)
    selected_sheets = bp.get('selected_sheets') or []
    sheets_cfg = bp.get('sheets') or {}

    if not selected_sheets:
        raise SchemaError(
            "Bonepile mapping error: no sheets selected. Please update mapping in Upload.",
            details={"file": excel_file, "available_sheets": list_excel_sheets(excel_file)},
        )

    normalized_frames = []
    for sheet_name in selected_sheets:
        sc = sheets_cfg.get(sheet_name, {}) if isinstance(sheets_cfg, dict) else {}
        header_row_excel = int(sc.get('header_row_excel') or default_bonepile_mapping.get('header_row_excel', 2))
        header_row = max(0, header_row_excel - 1)
        configured_cols = sc.get('columns', {}) if isinstance(sc.get('columns', {}), dict) else {}

        try:
            df_sheet = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row)
        except Exception as e:
            raise SchemaError(
                f"Bonepile mapping error: failed to read sheet '{sheet_name}' (header row={header_row_excel}). Please update mapping in Upload. ({e})",
                details={"file": excel_file, "sheet": sheet_name, "header_row_excel": header_row_excel, "header_row": header_row},
            )

        # Resolve required columns from saved mapping (exact, but case/whitespace tolerant)
        sn_col = resolve_df_column(df_sheet, configured_cols.get('sn'))
        pic_col = resolve_df_column(df_sheet, configured_cols.get('pic'))
        result_col = resolve_df_column(df_sheet, configured_cols.get('result'))
        igs_action_col = resolve_df_column(df_sheet, configured_cols.get('igs_action'))
        igs_status_col = resolve_df_column(df_sheet, configured_cols.get('igs_status'))
        bp_duration_col = resolve_df_column(df_sheet, configured_cols.get('bp_duration'))
        nv_disposition_col = resolve_df_column(df_sheet, configured_cols.get('nv_disposition'))
        fail_time_col = resolve_df_column(df_sheet, configured_cols.get('fail_time'))

        required = {
            "sn": sn_col,
            "result": result_col,
            "pic": pic_col,
            "nv_disposition": nv_disposition_col,
            "igs_action": igs_action_col,
            "igs_status": igs_status_col,
            "bp_duration": bp_duration_col,
        }
        missing = [k for k, v in required.items() if not v]
        if missing:
            raise SchemaError(
                f"Bonepile mapping error: missing required columns {missing} for sheet '{sheet_name}'. Please update mapping in Upload.",
                details={
                    "file": excel_file,
                    "sheet": sheet_name,
                    "header_row_excel": header_row_excel,
                    "header_row": header_row,
                    "expected": {k: configured_cols.get(k) for k in missing},
                    "configured_columns": configured_cols,
                    "available_columns": [str(c) for c in list(df_sheet.columns)],
                },
            )

        # Remove duplicate header
        if len(df_sheet) > 0 and sn_col in df_sheet.columns:
            try:
                if str(df_sheet.iloc[0][sn_col]).strip().lower() == 'sn':
                    df_sheet = df_sheet.iloc[1:].reset_index(drop=True)
            except Exception:
                pass

        # Normalize to canonical columns so downstream logic is sheet-agnostic
        df_norm = pd.DataFrame({
            "sn": df_sheet[sn_col],
            "result": df_sheet[result_col],
            "pic": df_sheet[pic_col],
            "nv_disposition": df_sheet[nv_disposition_col],
            "igs_action": df_sheet[igs_action_col],
            "igs_status": df_sheet[igs_status_col],
            "bp_duration": df_sheet[bp_duration_col],
            "fail_time": df_sheet[fail_time_col] if fail_time_col else None,
        })
        df_norm["source_sheet"] = sheet_name
        normalized_frames.append(df_norm)

    df = pd.concat(normalized_frames, ignore_index=True) if normalized_frames else pd.DataFrame()

    # Canonical column names (used throughout)
    sn_col = "sn"
    pic_col = "pic"
    result_col = "result"
    igs_action_col = "igs_action"
    igs_status_col = "igs_status"
    bp_duration_col = "bp_duration"
    nv_disposition_col = "nv_disposition"

    # Status helpers (Status column is mapped into canonical "result")
    def is_status_fail(v):
        s = str(v).upper().strip()
        return s.startswith('FAIL')  # accept "FAIL", "Fail", "FAIL - ...", etc.

    def is_status_all_pass(v):
        s = str(v).upper().strip()
        return s.startswith('ALL PASS')
    
    # Filter valid SN
    valid_sn_records = df[df[sn_col].apply(is_valid_sn)].copy()
    
    # Load FA Work Log mapping
    sn_wo_mapping = load_fa_work_log()
    
    # Get unique SNs (total trays in BP)
    unique_sns = valid_sn_records[sn_col].unique()
    
    # Calculate dispositions based on NV Disposition column only.
    # Each "MM/DD : blabla" segment counts as 1 disposition.
    all_dispositions = []  # List of all dispositions with details
    disposition_by_row = {}  # Map row index to disposition info

    for idx, row in valid_sn_records.iterrows():
        sn = safe_get_row_value(row, sn_col, '')
        sn_str = str(int(sn)) if isinstance(sn, (int, float)) else str(sn).strip().replace('.0', '')
        nv_disp_text = safe_get_row_value(row, nv_disposition_col, '')
        igs_action_text = safe_get_row_value(row, igs_action_col, '')
        igs_status_text = safe_get_row_value(row, igs_status_col, '')
        result_text = safe_get_row_value(row, result_col, '')
        
        # Parse dispositions from NV Disposition column (strict mm/dd: format)
        nv_dispositions = parse_nv_dispositions_mmdd_colon(str(nv_disp_text))

        # Pending definition (row-level): Status=Fail + PIC=IGS
        result_upper = str(result_text).upper().strip()
        pic_upper = str(safe_get_row_value(row, pic_col, '')).upper().strip()
        is_fail_igs_row = is_status_fail(result_text) and (pic_upper == 'IGS')
        
        # Store dispositions for this row
        row_dispositions = []
        for disp_date, desc in nv_dispositions:
            wo = sn_wo_mapping.get(sn_str, '')
            # Normalize WO if it exists
            if wo:
                wo = normalize_wo(wo)
            row_dispositions.append({
                'date': disp_date,
                'description': desc,
                'sn': sn_str,
                'wo': wo,
                'row_idx': idx
            })
            all_dispositions.append({
                'date': disp_date,
                'description': desc,
                'sn': sn_str,
                'wo': wo,
                'row_idx': idx,
                # Pending/Completed are defined simply from row status for now.
                # Completed will be computed as Total - Pending at metric level.
                'is_pending': is_fail_igs_row,
                'is_completed': (not is_fail_igs_row),
                'result': result_upper,
                'pic': pic_upper,
            })
        
        
        disposition_by_row[idx] = {
            'dispositions': row_dispositions,
            'is_pending': is_fail_igs_row,
            'sn': sn_str
        }

    # Sort dispositions by date (then SN for stability)
    try:
        all_dispositions.sort(key=lambda d: (d.get('date') or datetime.min, d.get('sn') or '', d.get('row_idx') or 0))
    except Exception:
        pass
    
    # Fail records: COUNTIF(Status="Fail") - đếm số dòng có status = fail (per mapping)
    fail_records = valid_sn_records[
        valid_sn_records[result_col].apply(is_status_fail)
    ].copy()
    
    # Pass records: COUNTIF(Status="ALL PASS") - đếm số dòng có status = all pass (per mapping)
    pass_records = valid_sn_records[
        valid_sn_records[result_col].apply(is_status_all_pass)
    ].copy()
    
    # Get unique SNs for fail and pass
    unique_fail_sns = set(fail_records[sn_col].unique())
    unique_pass_sns = set(pass_records[sn_col].unique())
    
    # Fail records with PIC = IGS (for IGS Action/Status analysis)
    fail_igs_records = fail_records[
        fail_records[pic_col].astype(str).str.upper().str.strip() == 'IGS'
    ].copy()
    
    # Fail with empty IGS Action (only for IGS records)
    fail_with_empty_action = fail_igs_records[
        (fail_igs_records[igs_action_col].isna()) | 
        (fail_igs_records[igs_action_col].astype(str).str.strip() == '') |
        (fail_igs_records[igs_action_col].astype(str).str.strip() == 'nan')
    ]
    
    # In process dispositions (only for IGS records)
    # NOTE: Avoid DataFrame-as-mask bug when fail_with_empty_action is empty.
    if len(fail_with_empty_action) > 0:
        mask_has_action = ~fail_igs_records[sn_col].isin(set(fail_with_empty_action[sn_col].unique()))
        fail_with_action = fail_igs_records[mask_has_action]
    else:
        fail_with_action = fail_igs_records.copy()
    
    if igs_status_col and igs_status_col in fail_with_action.columns and len(fail_with_action) > 0:
        in_process_records = fail_with_action[
            fail_with_action[igs_status_col].apply(is_in_process)
        ]
    else:
        in_process_records = pd.DataFrame()
    
    # Waiting for material
    if len(in_process_records) > 0:
        if igs_status_col and igs_status_col in in_process_records.columns:
            status_filter = in_process_records[igs_status_col].apply(is_waiting_for_material)
        else:
            status_filter = pd.Series([False] * len(in_process_records), index=in_process_records.index)
        
        if igs_action_col and igs_action_col in in_process_records.columns:
            action_filter = in_process_records[igs_action_col].apply(is_waiting_for_material)
        else:
            action_filter = pd.Series([False] * len(in_process_records), index=in_process_records.index)
        
        waiting_material_records = in_process_records[status_filter | action_filter]
    else:
        waiting_material_records = pd.DataFrame()
    
    # Current dispositions (Status = FAIL, PIC = IGS)
    current_dispositions_completed = []
    current_dispositions_waiting = []
    current_dispositions_testing = []
    current_dispositions_waiting_material = []
    
    # IMPORTANT: Use the mapped Status column (result_col) rather than any hardcoded column name.
    # Status values may be "Fail"/"FAIL", so compare case-insensitively.
    current_mask_fail = valid_sn_records[result_col].astype(str).str.upper().str.strip().str.startswith('FAIL')
    current_mask_pic_igs = valid_sn_records[pic_col].astype(str).str.upper().str.strip() == 'IGS'
    current_igs_fail_records = valid_sn_records[current_mask_fail & current_mask_pic_igs].copy()

    for idx, row in current_igs_fail_records.iterrows():
        sn = safe_get_row_value(row, sn_col, '')
        sn_str = str(int(sn)) if isinstance(sn, (int, float)) else str(sn).strip().replace('.0', '')
        igs_status = safe_get_row_value(row, igs_status_col, '')
        igs_status_lower = str(igs_status).lower()
        igs_action_text = safe_get_row_value(row, igs_action_col, '')
        nv_disp_text = safe_get_row_value(row, nv_disposition_col, '')

        # Parse latest NV Disposition date (mm/dd:) and latest IGS Action date (mm/dd:)
        nv_entries = parse_nv_dispositions_mmdd_colon(str(nv_disp_text))
        igs_entries = parse_nv_dispositions_mmdd_colon(str(igs_action_text))
        nv_latest_date, _ = get_latest_entry_from_date_desc_list(nv_entries)
        igs_latest_date, igs_latest_desc = get_latest_entry_from_date_desc_list(igs_entries)
        igs_latest_desc_lower = str(igs_latest_desc).lower()
        
        wo = sn_wo_mapping.get(sn_str, '')
        # Normalize WO if it exists
        if wo:
            wo = normalize_wo(wo)
        
        disposition_info = {
            'sn': sn_str,
            'wo': wo,
            'igs_status': str(igs_status),
            'igs_action_latest_date': igs_latest_date.strftime('%Y-%m-%d') if isinstance(igs_latest_date, datetime) else '',
            'nv_dispo_latest_date': nv_latest_date.strftime('%Y-%m-%d') if isinstance(nv_latest_date, datetime) else '',
            'igs_action_latest_text': str(igs_latest_desc),
            'row_idx': idx
        }

        # Classification rules (priority order):
        # 1) If IGS Status contains "waiting for NV ..." => treat as Completed (override)
        if ('waiting for nv dispo' in igs_status_lower or
            'waiting for nv disposition' in igs_status_lower or
            'waiting for nv' in igs_status_lower):
            current_dispositions_completed.append(disposition_info)
            continue

        # 2) If NV Disposition is empty => treat as Completed
        if not nv_entries:
            current_dispositions_completed.append(disposition_info)
            continue

        # 3) Waiting for material: IGS Status or latest IGS Action contains material keywords
        material_keywords = [
            'waiting for material',
            'waiting for strata',
            'waiting for cx9',
            'waiting for bbay',
            'waiting for bf4',
            'waiting for new material',
            'waiting for new strata',
        ]
        if any(k in igs_status_lower for k in material_keywords) or any(k in igs_latest_desc_lower for k in material_keywords):
            current_dispositions_waiting_material.append(disposition_info)
            continue

        # 4) Testing: IGS Status or latest IGS Action contains "testing"
        if ('testing' in igs_status_lower) or ('testing' in igs_latest_desc_lower):
            current_dispositions_testing.append(disposition_info)
            continue

        # 5) Waiting IGS action: IGS Action empty OR latest IGS Action date < latest NV Disposition date
        if (not igs_entries) or (isinstance(igs_latest_date, datetime) and isinstance(nv_latest_date, datetime) and igs_latest_date < nv_latest_date):
            current_dispositions_waiting.append(disposition_info)
            continue
        if (not isinstance(igs_latest_date, datetime)) and isinstance(nv_latest_date, datetime):
            current_dispositions_waiting.append(disposition_info)
            continue

        # 6) Otherwise Completed (igs_latest_date >= nv_latest_date)
        current_dispositions_completed.append(disposition_info)
    
    return {
        'df': valid_sn_records,
        'selected_sheets': selected_sheets,
        'unique_sns': unique_sns,
        'unique_fail_sns': unique_fail_sns,
        'unique_pass_sns': unique_pass_sns,
        'fail_records': fail_records,
        'pass_records': pass_records,
        'fail_igs_records': fail_igs_records,
        'fail_with_empty_action': fail_with_empty_action,
        'in_process_records': in_process_records,
        'waiting_material_records': waiting_material_records,
        'all_dispositions': all_dispositions,
        'disposition_by_row': disposition_by_row,  # Disposition theo row
        'sn_wo_mapping': sn_wo_mapping,  # Mapping SN -> WO
        'current_dispositions_completed': current_dispositions_completed,
        'current_dispositions_waiting': current_dispositions_waiting,
        'current_dispositions_testing': current_dispositions_testing,
        'current_dispositions_waiting_material': current_dispositions_waiting_material,
        'cols': {
            'sn': sn_col,
            'nv_disposition': nv_disposition_col,
            'igs_action': igs_action_col,
            'igs_status': igs_status_col,
            'bp_duration': bp_duration_col,
            'result': result_col,
            'pic': pic_col
        }
    }

@app.route('/')
def index():
    try:
        data = load_data()
        if data is None:
            return render_template('index.html', stats=None, error="No Excel file found. Please upload a file.", ip=get_local_ip())
        
        # Calculate statistics
        fail_empty_sns = set(data['fail_with_empty_action'][data['cols']['sn']].unique()) if len(data['fail_with_empty_action']) > 0 else set()
        
        # Calculate disposition statistics (NV Disposition only)
        total_dispositions = len(data['all_dispositions']) if 'all_dispositions' in data else 0
        pending_dispositions = sum(1 for d in data['all_dispositions'] if d.get('is_pending')) if 'all_dispositions' in data else 0
        completed_dispositions = max(0, total_dispositions - pending_dispositions)
        
        stats = {
            'total_trays': len(data['unique_sns']),
            'total_fail': len(data['fail_records']),
            'total_pass': len(data['pass_records']),
            'total_fail_unique': len(data['unique_fail_sns']),
            'total_pass_unique': len(data['unique_pass_sns']),
            'total_dispositions': total_dispositions,
            'completed_dispositions': completed_dispositions,
            'fail_empty_action': len(fail_empty_sns),
            'in_process': len(data['in_process_records']),
            'in_process_unique': len(set(data['in_process_records'][data['cols']['sn']].unique())) if len(data['in_process_records']) > 0 else 0,
            'waiting_material': len(data['waiting_material_records']),
            'waiting_material_unique': len(set(data['waiting_material_records'][data['cols']['sn']].unique())) if len(data['waiting_material_records']) > 0 else 0
        }
    
        # Calculate average duration
        fail_empty_action_sns = fail_empty_sns
        completed_records = data['df'][
            ~(data['df'][data['cols']['sn']].isin(fail_empty_action_sns) & 
              (data['df'][data['cols']['pic']].astype(str).str.upper().str.strip() == 'IGS') & 
              (data['df'][data['cols']['result']].astype(str).str.upper().str.strip() == 'FAIL'))
        ]
        
        bp_durations = []
        if data['cols']['bp_duration'] in completed_records.columns:
            for d in completed_records[data['cols']['bp_duration']].dropna():
                try:
                    dur = float(d)
                    if dur >= 0:
                        bp_durations.append(dur)
                except:
                    pass
        
        stats['avg_duration'] = float(np.mean(bp_durations)) if bp_durations else 0.0
        stats['median_duration'] = float(np.median(bp_durations)) if bp_durations else 0.0
        stats['min_duration'] = float(np.min(bp_durations)) if bp_durations else 0.0
        stats['max_duration'] = float(np.max(bp_durations)) if bp_durations else 0.0
        stats['std_duration'] = float(np.std(bp_durations)) if bp_durations else 0.0
        stats['duration_count'] = len(bp_durations)
        
        return render_template(
            'index.html',
            stats=stats,
            error=None,
            ip=get_local_ip(),
            selected_sheets=data.get('selected_sheets') if isinstance(data, dict) else None,
        )
    except SchemaError as e:
        msg = f"{str(e)}"
        return render_template(
            'index.html',
            stats=None,
            error=msg,
            error_details=getattr(e, 'details', None),
            ip=get_local_ip(),
        )
    except Exception as e:
        return render_template('index.html', stats=None, error=f"Error loading data: {str(e)}", ip=get_local_ip())

@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    # Shared mapping context for the Upload page (Upload + Settings in one place)
    def build_mapping_context(current_mapping):
        # Ensure keys exist
        if 'bonepile' not in current_mapping:
            current_mapping['bonepile'] = json.loads(json.dumps(DEFAULT_USER_MAPPING['bonepile']))
        if 'fa_work_log' not in current_mapping:
            current_mapping['fa_work_log'] = json.loads(json.dumps(DEFAULT_USER_MAPPING['fa_work_log']))

        bonepile_path = resolve_uploaded_or_local_path(current_mapping['bonepile'].get('file_name', DEFAULT_USER_MAPPING['bonepile']['file_name']))
        bonepile_sheets = list_excel_sheets(bonepile_path) if bonepile_path else []

        # Normalize to multi-sheet mapping (without persisting on GET)
        bp_norm = normalize_bonepile_mapping(current_mapping.get('bonepile', {}))
        selected_sheets = bp_norm.get('selected_sheets') or []
        sheets_cfg = bp_norm.get('sheets') or {}

        # Filter invalid selections if workbook is available
        if bonepile_sheets:
            selected_sheets = [s for s in selected_sheets if s in bonepile_sheets]

        # Default selection: prefer VR-TS1 if present
        if bonepile_sheets and not selected_sheets:
            selected_sheets = ['VR-TS1'] if 'VR-TS1' in bonepile_sheets else [bonepile_sheets[0]]

        # Editing sheet (which mapping panel shows)
        edit_sheet = (request.args.get('edit_sheet') or '').strip()
        if not edit_sheet:
            edit_sheet = selected_sheets[0] if selected_sheets else (bonepile_sheets[0] if bonepile_sheets else (bp_norm.get('selected_sheets') or [''])[0])
        if bonepile_sheets and edit_sheet not in bonepile_sheets:
            edit_sheet = bonepile_sheets[0]

        # Ensure edit sheet has a config
        if edit_sheet and edit_sheet not in sheets_cfg:
            # Seed from legacy defaults
            legacy = current_mapping.get('bonepile', {}) or {}
            sheets_cfg[edit_sheet] = {
                "header_row_excel": get_bonepile_header_row_excel(legacy),
                "columns": json.loads(json.dumps(legacy.get('columns', DEFAULT_USER_MAPPING['bonepile']['columns']) or {})),
            }
        edit_cfg = sheets_cfg.get(edit_sheet, {}) if isinstance(sheets_cfg, dict) else {}
        edit_header_row_excel = int(edit_cfg.get('header_row_excel') or DEFAULT_USER_MAPPING['bonepile']['header_row_excel'])
        edit_header_row = max(0, edit_header_row_excel - 1)

        bonepile_header_errors = []
        bonepile_headers = []
        if bonepile_path and edit_sheet:
            # 1) openpyxl header row -> includes index labels
            try:
                bonepile_headers = get_excel_header_values(bonepile_path, edit_sheet, edit_header_row)
            except Exception as e:
                bonepile_header_errors.append(f"openpyxl header read failed: {e}")

            # 2) pandas fallback (nrows=0)
            if not bonepile_headers:
                try:
                    from openpyxl.utils import get_column_letter
                    cols = get_pandas_columns(bonepile_path, edit_sheet, edit_header_row)
                    bonepile_headers = [{
                        "value": f"__idx__{c['index']}",
                        "name": c.get("name", ""),
                        "index": c["index"],
                        "label": f"{get_column_letter(c['index']+1)} ({c['index']}) - {c.get('name') or '(blank)'}",
                    } for c in cols]
                except Exception as e:
                    bonepile_header_errors.append(f"pandas header read failed: {e}")

            # 3) last resort: show pure index options (no header text)
            if not bonepile_headers:
                try:
                    from openpyxl.utils import get_column_letter
                    max_col = get_excel_sheet_max_column(bonepile_path, edit_sheet)
                    max_col = min(int(max_col or 0), 200)
                    bonepile_headers = [{
                        "value": f"__idx__{i}",
                        "name": "",
                        "index": i,
                        "label": f"{get_column_letter(i+1)} ({i}) - (unknown header)",
                    } for i in range(max_col)]
                except Exception as e:
                    bonepile_header_errors.append(f"index fallback failed: {e}")

        # Normalize configured column selections (name -> __idx__N) for the editing sheet only
        try:
            edit_cols = (edit_cfg.get('columns') or {}) if isinstance(edit_cfg, dict) else {}
            first_idx_by_name = {}
            for h in bonepile_headers:
                n = str(h.get('name', '')).strip()
                if not n:
                    continue
                key = n.lower()
                if key not in first_idx_by_name:
                    first_idx_by_name[key] = h.get('value')
            for k, v in list(edit_cols.items()):
                if not v:
                    continue
                v_str = str(v).strip()
                if v_str.startswith('__idx__'):
                    continue
                resolved_idx = first_idx_by_name.get(v_str.lower())
                if resolved_idx:
                    edit_cols[k] = resolved_idx
            edit_cfg['columns'] = edit_cols
            sheets_cfg[edit_sheet] = edit_cfg
        except Exception:
            pass

        # Expose normalized mapping to the template (without persisting)
        current_mapping['bonepile']['selected_sheets'] = selected_sheets
        current_mapping['bonepile']['sheets'] = sheets_cfg
        # Keep legacy keys as a convenience for older code paths
        if edit_sheet:
            current_mapping['bonepile']['sheet_name'] = edit_sheet
            current_mapping['bonepile']['header_row_excel'] = edit_header_row_excel
            current_mapping['bonepile']['header_row'] = max(0, edit_header_row_excel - 1)
            current_mapping['bonepile']['columns'] = edit_cfg.get('columns', {}) if isinstance(edit_cfg, dict) else {}

        worklog_path = resolve_uploaded_or_local_path(current_mapping['fa_work_log'].get('file_name', DEFAULT_USER_MAPPING['fa_work_log']['file_name']))
        worklog_sheets = list_excel_sheets(worklog_path) if worklog_path else []
        if worklog_sheets and current_mapping['fa_work_log'].get('sheet_name') not in worklog_sheets:
            current_mapping['fa_work_log']['sheet_name'] = worklog_sheets[0]
        worklog_max_col = (
            get_excel_sheet_max_column(worklog_path, current_mapping['fa_work_log'].get('sheet_name', DEFAULT_USER_MAPPING['fa_work_log']['sheet_name']))
            if worklog_path
            else 0
        )
        worklog_max_col = min(int(worklog_max_col or 0), 60)

        def col_label(idx):
            try:
                from openpyxl.utils import get_column_letter
                return f"{get_column_letter(idx + 1)} ({idx})"
            except Exception:
                return str(idx)

        worklog_col_options = [{"value": i, "label": col_label(i)} for i in range(max(0, worklog_max_col))]

        return {
            "mapping": current_mapping,
            "bonepile_path": bonepile_path,
            "bonepile_sheets": bonepile_sheets,
            "bonepile_selected_sheets": selected_sheets,
            "bonepile_edit_sheet": edit_sheet,
            "bonepile_edit_header_row_excel": edit_header_row_excel,
            "bonepile_edit_columns": (edit_cfg.get('columns') or {}) if isinstance(edit_cfg, dict) else {},
            "bonepile_headers": bonepile_headers,
            "bonepile_header_errors": bonepile_header_errors,
            "bonepile_header_row_excel": edit_header_row_excel,
            "worklog_path": worklog_path,
            "worklog_sheets": worklog_sheets,
            "worklog_col_options": worklog_col_options,
        }

    if request.method == 'POST':
        form_type = (request.form.get('form_type') or '').strip().lower()

        # ---- Save mapping settings ----
        if form_type == 'settings':
            mapping = load_user_mapping()
            try:
                bp = mapping.get('bonepile') or json.loads(json.dumps(DEFAULT_USER_MAPPING['bonepile']))
                wl = mapping.get('fa_work_log') or json.loads(json.dumps(DEFAULT_USER_MAPPING['fa_work_log']))

                # ---- Bonepile (multi-sheet + per-sheet mapping) ----
                bp_norm = normalize_bonepile_mapping(bp)
                bp['file_name'] = (request.form.get('bonepile_file_name') or bp.get('file_name') or DEFAULT_USER_MAPPING['bonepile']['file_name']).strip()

                selected = request.form.getlist('bonepile_selected_sheets')
                selected = [s.strip() for s in selected if s and str(s).strip()]

                edit_sheet = (request.form.get('bonepile_edit_sheet') or '').strip()
                if not edit_sheet and selected:
                    edit_sheet = selected[0]

                # If user unchecks everything, keep at least the sheet being edited (or previous selection)
                if not selected:
                    if edit_sheet:
                        selected = [edit_sheet]
                    else:
                        prev = bp_norm.get('selected_sheets') or []
                        selected = prev[:1] if prev else [DEFAULT_USER_MAPPING['bonepile']['sheet_name']]

                bp_norm['selected_sheets'] = selected
                if not isinstance(bp_norm.get('sheets'), dict):
                    bp_norm['sheets'] = {}

                # Update mapping for the sheet currently being edited
                if edit_sheet:
                    sc = bp_norm['sheets'].get(edit_sheet) or {}
                    sc['header_row_excel'] = int(request.form.get(
                        'bonepile_header_row',
                        sc.get('header_row_excel', DEFAULT_USER_MAPPING['bonepile']['header_row_excel'])
                    ))
                    sc_cols = sc.get('columns', {}) if isinstance(sc.get('columns', {}), dict) else {}
                    for field in ['sn', 'result', 'pic', 'nv_disposition', 'igs_action', 'igs_status', 'bp_duration', 'fail_time']:
                        sc_cols[field] = (request.form.get(f'bonepile_col_{field}') or '').strip()
                    sc['columns'] = sc_cols
                    bp_norm['sheets'][edit_sheet] = sc

                # Persist normalized multi-sheet mapping
                bp['selected_sheets'] = bp_norm.get('selected_sheets') or []
                bp['sheets'] = bp_norm.get('sheets') or {}

                # Keep legacy single-sheet keys in sync (use the first selected sheet)
                primary_sheet = (bp['selected_sheets'][0] if bp.get('selected_sheets') else edit_sheet) or DEFAULT_USER_MAPPING['bonepile']['sheet_name']
                bp['sheet_name'] = primary_sheet
                primary_cfg = (bp.get('sheets') or {}).get(primary_sheet, {}) if isinstance(bp.get('sheets'), dict) else {}
                primary_header_row_excel = int(primary_cfg.get('header_row_excel') or DEFAULT_USER_MAPPING['bonepile']['header_row_excel'])
                bp['header_row_excel'] = primary_header_row_excel
                bp['header_row'] = max(0, primary_header_row_excel - 1)
                primary_cols = primary_cfg.get('columns', {}) if isinstance(primary_cfg.get('columns', {}), dict) else {}
                bp['columns'] = primary_cols

                wl['file_name'] = (request.form.get('worklog_file_name') or wl.get('file_name') or DEFAULT_USER_MAPPING['fa_work_log']['file_name']).strip()
                wl['sheet_name'] = (request.form.get('worklog_sheet_name') or wl.get('sheet_name') or DEFAULT_USER_MAPPING['fa_work_log']['sheet_name']).strip()
                wl['start_row'] = int(request.form.get('worklog_start_row', wl.get('start_row', DEFAULT_USER_MAPPING['fa_work_log']['start_row'])))
                wl['sn_col_index'] = int(request.form.get('worklog_sn_col_index', wl.get('sn_col_index', DEFAULT_USER_MAPPING['fa_work_log']['sn_col_index'])))
                wl['wo_col_index'] = int(request.form.get('worklog_wo_col_index', wl.get('wo_col_index', DEFAULT_USER_MAPPING['fa_work_log']['wo_col_index'])))

                mapping['bonepile'] = bp
                mapping['fa_work_log'] = wl
                save_user_mapping(mapping)
                flash('Settings saved successfully.')
            except Exception as e:
                flash(f'Error saving settings: {e}')

            return redirect(url_for('upload_file') + '#settings')

        # ---- Upload files (default) ----
        uploaded_files = []

        # Handle NV_IGS_VR144_Bonepile.xlsx
        if 'file_bonepile' in request.files:
            file = request.files['file_bonepile']
            if file and file.filename != '' and allowed_file(file.filename):
                filename = 'NV_IGS_VR144_Bonepile.xlsx'
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                uploaded_files.append('NV_IGS_VR144_Bonepile.xlsx')

        # Handle FA_Work_Log.xlsx
        if 'file_fa_work_log' in request.files:
            file = request.files['file_fa_work_log']
            if file and file.filename != '' and allowed_file(file.filename):
                filename = 'FA_Work_Log.xlsx'
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                uploaded_files.append('FA_Work_Log.xlsx')

        if uploaded_files:
            flash(f'Files uploaded successfully: {", ".join(uploaded_files)}')
        else:
            flash('No valid files uploaded. Please upload .xlsx or .xls files.')

        # Upload only (do not auto-run dashboard). Redirect back to mapping/settings.
        return redirect(url_for('upload_file') + '#settings')

    mapping = load_user_mapping()
    ctx = build_mapping_context(mapping)
    return render_template('upload.html', ip=get_local_ip(), **ctx)


@app.route('/settings', methods=['GET', 'POST'])
def settings():
    # Backwards-compatible endpoint: settings now lives inside /upload
    return redirect(url_for('upload_file') + '#settings')

@app.route('/api/sn-list/<category>')
def get_sn_list(category):
    try:
        data = load_data()
        if data is None:
            return jsonify({'error': 'No data available'}), 404
        
        if category == 'total':
            sns = sorted([str(int(sn)) if isinstance(sn, (int, float)) else str(sn) 
                         for sn in data['unique_sns']])
        elif category == 'fail':
            sns = sorted([str(int(sn)) if isinstance(sn, (int, float)) else str(sn) 
                         for sn in data['unique_fail_sns']])
        elif category == 'pass':
            sns = sorted([str(int(sn)) if isinstance(sn, (int, float)) else str(sn) 
                         for sn in data['unique_pass_sns']])
        else:
            sns = []
        
        return jsonify({'sns': sns, 'count': len(sns)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/fail-empty-action')
def get_fail_empty_action():
    try:
        data = load_data()
        if data is None:
            return jsonify({'error': 'No data available'}), 404
        
        result = []
        for idx, row in data['fail_with_empty_action'].iterrows():
            sn = row[data['cols']['sn']]
            sn_str = str(int(sn)) if isinstance(sn, (int, float)) else str(sn)
            nv_disp = row[data['cols']['nv_disposition']] if pd.notna(row[data['cols']['nv_disposition']]) else ''
            
            result.append({
                'sn': sn_str,
                'nv_disposition': str(nv_disp)
            })
        
        return jsonify({'data': result, 'count': len(result)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/in-process')
def get_in_process():
    try:
        data = load_data()
        if data is None:
            return jsonify({'error': 'No data available'}), 404
        
        result = []
        for idx, row in data['in_process_records'].iterrows():
            sn = row[data['cols']['sn']]
            sn_str = str(int(sn)) if isinstance(sn, (int, float)) else str(sn)
            nv_disp = row[data['cols']['nv_disposition']] if pd.notna(row[data['cols']['nv_disposition']]) else ''
            igs_action = row[data['cols']['igs_action']] if pd.notna(row[data['cols']['igs_action']]) else ''
            
            result.append({
                'sn': sn_str,
                'nv_disposition': str(nv_disp),
                'igs_action': str(igs_action)
            })
        
        return jsonify({'data': result, 'count': len(result)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/waiting-material')
def get_waiting_material():
    try:
        data = load_data()
        if data is None:
            return jsonify({'error': 'No data available'}), 404
        
        result = []
        for idx, row in data['waiting_material_records'].iterrows():
            sn = safe_get_row_value(row, data['cols']['sn'], '')
            sn_str = str(int(sn)) if isinstance(sn, (int, float)) else str(sn)
            nv_disp = safe_get_row_value(row, data['cols']['nv_disposition'], '')
            igs_action = safe_get_row_value(row, data['cols']['igs_action'], '')
            igs_status = safe_get_row_value(row, data['cols']['igs_status'], '')
            
            result.append({
                'sn': sn_str,
                'nv_disposition': str(nv_disp),
                'igs_action': str(igs_action),
                'igs_status': str(igs_status)
            })
        
        return jsonify({'data': result, 'count': len(result)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/disposition-stats')
def get_disposition_stats():
    """Get disposition statistics with filters for date range and WO"""
    try:
        data = load_data()
        if data is None:
            return jsonify({'error': 'No data available'}), 404
        
        # Get filter parameters
        start_date_str = request.args.get('start_date', '')
        end_date_str = request.args.get('end_date', '')
        wo_filter = request.args.get('wo', '')
        
        # Parse dates
        start_date = None
        end_date = None
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            except:
                pass
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
                # Include the entire end date
                end_date = end_date + timedelta(days=1)
            except:
                pass
        
        # Filter dispositions
        filtered_dispositions = []
        for disp in data['all_dispositions']:
            # Filter by date
            if start_date and disp['date'] < start_date:
                continue
            if end_date and disp['date'] >= end_date:
                continue
            # Filter by WO
            if wo_filter and wo_filter != 'ALL' and disp['wo'] != wo_filter:
                continue
            filtered_dispositions.append(disp)
        
        # Calculate statistics
        total_dispositions = len(filtered_dispositions)
        pending_dispositions = sum(1 for d in filtered_dispositions if d.get('is_pending'))
        completed_dispositions = max(0, total_dispositions - pending_dispositions)
        
        # Calculate average per day/week
        if filtered_dispositions:
            # If date filters are provided, use them; otherwise use min/max from dispositions
            if start_date and end_date:
                # Use the filter date range (end_date was already +1 day for filtering, so subtract 1)
                actual_end_date = end_date - timedelta(days=1)
                days_diff = (actual_end_date - start_date).days + 1
                weeks_diff = days_diff / 7.0
            elif start_date:
                # Only start date provided, use max date from dispositions
                dates = [d['date'] for d in filtered_dispositions]
                max_date = max(dates)
                days_diff = (max_date - start_date).days + 1
                weeks_diff = days_diff / 7.0
            elif end_date:
                # Only end date provided, use min date from dispositions
                dates = [d['date'] for d in filtered_dispositions]
                min_date = min(dates)
                actual_end_date = end_date - timedelta(days=1)
                days_diff = (actual_end_date - min_date).days + 1
                weeks_diff = days_diff / 7.0
            else:
                # No date filter, use min/max from dispositions
                dates = [d['date'] for d in filtered_dispositions]
                min_date = min(dates)
                max_date = max(dates)
                days_diff = (max_date - min_date).days + 1
                weeks_diff = days_diff / 7.0
            
            avg_per_day = total_dispositions / days_diff if days_diff > 0 else 0
            avg_per_week = total_dispositions / weeks_diff if weeks_diff > 0 else 0
        else:
            avg_per_day = 0
            avg_per_week = 0
        
        # Get unique WOs and dates for dropdowns
        unique_wos = sorted(set([d['wo'] for d in data['all_dispositions'] if d['wo']]))
        unique_dates = sorted(set([d['date'].strftime('%Y-%m-%d') for d in data['all_dispositions']]))
        
        return jsonify({
            'total_dispositions': total_dispositions,
            'completed_dispositions': completed_dispositions,
            'pending_dispositions': pending_dispositions,
            'avg_per_day': round(avg_per_day, 2),
            'avg_per_week': round(avg_per_week, 2),
            'unique_wos': unique_wos,
            'unique_dates': unique_dates
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/current-dispositions')
def get_current_dispositions():
    """Get current dispositions (Status = Fail, PIC = IGS) with status"""
    try:
        data = load_data()
        if data is None:
            return jsonify({'error': 'No data available'}), 404
        
        return jsonify({
            'completed': data['current_dispositions_completed'],
            'waiting': data['current_dispositions_waiting'],
            'testing': data.get('current_dispositions_testing', []),
            'waiting_material': data['current_dispositions_waiting_material'],
            'total_completed': len(data['current_dispositions_completed']),
            'total_waiting': len(data['current_dispositions_waiting']),
            'total_testing': len(data.get('current_dispositions_testing', [])),
            'total_waiting_material': len(data['current_dispositions_waiting_material'])
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/all-dispositions')
def get_all_dispositions():
    """Get all dispositions (supports date/WO filters)"""
    try:
        data = load_data()
        if data is None:
            return jsonify({'error': 'No data available'}), 404

        # Optional filters (same semantics as /api/disposition-stats)
        start_date_str = request.args.get('start_date', '')
        end_date_str = request.args.get('end_date', '')
        wo_filter = request.args.get('wo', '')

        start_date = None
        end_date = None
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            except Exception:
                start_date = None
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1)
            except Exception:
                end_date = None
        
        # Format dispositions for display
        dispositions_list = []
        for disp in data.get('all_dispositions', []):
            # Filter by date
            if start_date and disp.get('date') and disp['date'] < start_date:
                continue
            if end_date and disp.get('date') and disp['date'] >= end_date:
                continue
            # Filter by WO
            if wo_filter and wo_filter != 'ALL' and disp.get('wo', '') != wo_filter:
                continue

            dispositions_list.append({
                'date': disp['date'].strftime('%Y-%m-%d') if disp['date'] else '',
                'description': disp['description'],
                'sn': disp['sn'],
                'wo': disp['wo'],
                'is_completed': disp.get('is_completed', False),
                'is_pending': disp.get('is_pending', False),
                'row_idx': disp['row_idx']
            })
        
        # Sort by date
        dispositions_list.sort(key=lambda x: x['date'] if x['date'] else '')
        
        return jsonify({
            'data': dispositions_list,
            'total': len(dispositions_list),
            'pending': sum(1 for d in dispositions_list if d.get('is_pending')),
            'completed': max(0, len(dispositions_list) - sum(1 for d in dispositions_list if d.get('is_pending'))),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/wo-statistics')
def get_wo_statistics():
    """Get statistics grouped by WO with filters"""
    try:
        data = load_data()
        if data is None:
            return jsonify({'error': 'No data available'}), 404
        
        # Get filter parameters
        start_date_str = request.args.get('start_date', '')
        end_date_str = request.args.get('end_date', '')
        wo_filter = request.args.get('wo', '')
        
        # Parse dates
        start_date = None
        end_date = None
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            except:
                pass
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
                end_date = end_date + timedelta(days=1)
            except:
                pass
        
        # Filter dispositions by date and WO
        filtered_dispositions = []
        for disp in data.get('all_dispositions', []):
            # Filter by date
            if start_date and disp['date'] < start_date:
                continue
            if end_date and disp['date'] >= end_date:
                continue
            # Filter by WO
            if wo_filter and wo_filter != 'ALL' and disp.get('wo', '') != wo_filter:
                continue
            filtered_dispositions.append(disp)
        
        # Get unique SNs from filtered dispositions
        filtered_sns = set([d['sn'] for d in filtered_dispositions])
        
        # Group by WO
        wo_stats = {}
        sn_wo_mapping = data.get('sn_wo_mapping', {})
        
        # Count trays (unique SN) by WO from filtered data
        for sn in data['unique_sns']:
            sn_str = str(int(sn)) if isinstance(sn, (int, float)) else str(sn).strip().replace('.0', '')
            # Only count if SN is in filtered dispositions or no filter applied
            if filtered_dispositions and sn_str not in filtered_sns:
                continue
                
            wo = sn_wo_mapping.get(sn_str, '')
            if wo:
                wo = normalize_wo(wo)
            if not wo:
                wo = 'No WO'
            
            if wo not in wo_stats:
                wo_stats[wo] = {
                    'wo': wo,
                    'tray_pass': 0,
                    'tray_fail': 0,
                    'dispositions': 0,
                    'dispositions_completed': 0
                }
        
        # Count pass/fail trays by WO
        for sn in data['unique_pass_sns']:
            sn_str = str(int(sn)) if isinstance(sn, (int, float)) else str(sn).strip().replace('.0', '')
            if filtered_dispositions and sn_str not in filtered_sns:
                continue
            wo = sn_wo_mapping.get(sn_str, '')
            if wo:
                wo = normalize_wo(wo)
            if not wo:
                wo = 'No WO'
            if wo in wo_stats:
                wo_stats[wo]['tray_pass'] += 1
        
        for sn in data['unique_fail_sns']:
            sn_str = str(int(sn)) if isinstance(sn, (int, float)) else str(sn).strip().replace('.0', '')
            if filtered_dispositions and sn_str not in filtered_sns:
                continue
            wo = sn_wo_mapping.get(sn_str, '')
            if wo:
                wo = normalize_wo(wo)
            if not wo:
                wo = 'No WO'
            if wo in wo_stats:
                wo_stats[wo]['tray_fail'] += 1
        
        # Count dispositions by WO from filtered dispositions
        for disp in filtered_dispositions:
            wo = disp.get('wo', '')
            if not wo:
                wo = 'No WO'
            if wo not in wo_stats:
                wo_stats[wo] = {
                    'wo': wo,
                    'tray_pass': 0,
                    'tray_fail': 0,
                    'dispositions': 0,
                    'dispositions_completed': 0
                }
            wo_stats[wo]['dispositions'] += 1
            if disp.get('is_completed', False):
                wo_stats[wo]['dispositions_completed'] += 1
        
        # Convert to list and sort
        wo_list = list(wo_stats.values())
        wo_list.sort(key=lambda x: x['wo'])
        
        return jsonify({
            'data': wo_list,
            'total_wo': len(wo_list)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/daily-test-analysis')
def daily_test_analysis_page():
    """Render Daily Test Analysis page"""
    return render_template('daily_test_analysis.html', ip=get_local_ip())

@app.route('/api/daily-test-analysis')
def get_daily_test_analysis():
    """Get daily test analysis data with date range filters"""
    try:
        start_date_str = request.args.get('start_date', '')
        end_date_str = request.args.get('end_date', '')
        
        if not start_date_str or not end_date_str:
            return jsonify({'error': 'Start date and end date are required'}), 400
        
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        if start_date > end_date:
            return jsonify({'error': 'Start date must be before end date'}), 400
        
        # Load test data
        test_data = load_daily_test_data(start_date, end_date)
        
        # Prepare WO statistics
        wo_stats = defaultdict(lambda: {
            'wo': '',
            'tray_pass': 0,
            'tray_fail': 0,
            'stations': defaultdict(lambda: {'pass': 0, 'fail': 0})
        })
        
        # Prepare Part Number statistics
        part_stats = defaultdict(lambda: {
            'part_number': '',
            'tray_pass': 0,
            'tray_fail': 0,
            'tray_total': 0,
            'stations': defaultdict(lambda: {'pass': 0, 'fail': 0})
        })
        
        # Track 1 time pass statistics
        one_time_pass_wo = defaultdict(lambda: {'total': 0, 'one_time_pass': 0})
        one_time_pass_part = defaultdict(lambda: {'total': 0, 'one_time_pass': 0})
        
        # Process SN test info
        for sn, test_list in test_data['sn_test_info'].items():
            wo = test_list[0]['wo'] if test_list else 'No WO'
            
            # Determine if SN passed
            # After 2026-01-01: pass FCT = pass (all WOs)
            # Before 2026-01-01: only pass RIN = pass
            is_pass = False
            cutoff_date = datetime(2026, 1, 1).date()
            
            for test in test_list:
                test_date = test.get('date')
                if not test_date:
                    # Try to get date from filename or use current date
                    test_date = datetime.now().date()
                # Ensure test_date is a date object
                elif isinstance(test_date, datetime):
                    test_date = test_date.date()
                elif isinstance(test_date, str):
                    test_date = datetime.strptime(test_date, '%Y-%m-%d').date()
                elif isinstance(test_date, pd.Timestamp):
                    test_date = test_date.to_pydatetime().date()
                elif not isinstance(test_date, date):
                    test_date = datetime.now().date()
                
                if test_date >= cutoff_date:
                    # New rule: pass FCT = pass
                    if test['station'] == 'FCT' and test['status'] == 'P':
                        is_pass = True
                        break
                else:
                    # Old rule: only pass if pass RIN
                    if test['station'] == 'RIN' and test['status'] == 'P':
                        is_pass = True
                        break
            
            # Check if SN is 1 time pass:
            # 1. All tests must be pass (no fail at any station)
            # 2. Must have pass at RIN station
            # 3. Only 1 test entry per station (no retest)
            is_one_time_pass = False
            if is_pass and test_list:
                # Check if all tests are pass
                all_pass = all(test['status'] == 'P' for test in test_list)
                # Check if has RIN pass
                has_rin_pass = any(test['station'] == 'RIN' and test['status'] == 'P' for test in test_list)
                # Check if only 1 test per station (no retest)
                stations_count = {}
                for test in test_list:
                    st = test['station']
                    stations_count[st] = stations_count.get(st, 0) + 1
                no_retest = all(count == 1 for count in stations_count.values())
                
                is_one_time_pass = all_pass and has_rin_pass and no_retest
            
            # Update WO statistics
            if wo not in wo_stats:
                wo_stats[wo]['wo'] = wo
            if is_pass:
                wo_stats[wo]['tray_pass'] += 1
            else:
                wo_stats[wo]['tray_fail'] += 1
            
            # Update 1 time pass statistics for WO
            one_time_pass_wo[wo]['total'] += 1
            if is_one_time_pass:
                one_time_pass_wo[wo]['one_time_pass'] += 1
            
            # Update Part Number statistics (count unique SN per part number)
            part_numbers_for_sn = test_data['sn_part_numbers'].get(sn, [])
            for part_num in part_numbers_for_sn:
                if part_num not in part_stats:
                    part_stats[part_num]['part_number'] = part_num
                part_stats[part_num]['tray_total'] += 1
                if is_pass:
                    part_stats[part_num]['tray_pass'] += 1
                else:
                    part_stats[part_num]['tray_fail'] += 1
                
                # Update 1 time pass statistics for Part Number
                one_time_pass_part[part_num]['total'] += 1
                if is_one_time_pass:
                    one_time_pass_part[part_num]['one_time_pass'] += 1
            
            # Count by station for both WO and Part Number
            for test in test_list:
                station = test['station']
                status = test['status']
                part_num = test.get('part_number', 'Unknown')
                
                if status == 'P':
                    wo_stats[wo]['stations'][station]['pass'] += 1
                    part_stats[part_num]['stations'][station]['pass'] += 1
                elif status == 'F':
                    wo_stats[wo]['stations'][station]['fail'] += 1
                    part_stats[part_num]['stations'][station]['fail'] += 1
        
        # Helper function to calculate station pass percentages
        def calculate_station_pass_percentages(stations_dict):
            """Calculate pass percentages for each station, sorted by station order"""
            station_pcts = {}
            # Sort stations according to custom order: FLA > FLB > AST > FTS > FCT > RIN
            sorted_stations = sort_stations(list(stations_dict.keys()))
            for station in sorted_stations:
                st_stats = stations_dict[station]
                st_total = st_stats['pass'] + st_stats['fail']
                st_pass_pct = (st_stats['pass'] / st_total * 100) if st_total > 0 else 0
                station_pcts[station] = round(st_pass_pct, 2)
            return station_pcts
        
        # Get WO -> Part Number mapping for sorting
        wo_part_mapping = {}
        for sn, test_list in test_data['sn_test_info'].items():
            wo = test_list[0]['wo'] if test_list else 'No WO'
            part_numbers = test_data['sn_part_numbers'].get(sn, [])
            if wo not in wo_part_mapping and part_numbers:
                wo_part_mapping[wo] = part_numbers[0]  # Use first part number
        
        # Convert to list for JSON with percentages
        wo_list = []
        for wo, stats in wo_stats.items():
            total_trays = stats['tray_pass'] + stats['tray_fail']
            pass_pct = (stats['tray_pass'] / total_trays * 100) if total_trays > 0 else 0
            fail_pct = (stats['tray_fail'] / total_trays * 100) if total_trays > 0 else 0
            
            # Calculate station percentages using helper function
            station_pcts = calculate_station_pass_percentages(stats['stations'])
            
            wo_list.append({
                'wo': wo,
                'tray_pass': stats['tray_pass'],
                'tray_fail': stats['tray_fail'],
                'tray_total': total_trays,
                'pass_percentage': round(pass_pct, 2),
                'fail_percentage': round(fail_pct, 2),
                'stations': {k: dict(v) for k, v in stats['stations'].items()},
                'station_pass_percentages': station_pcts,
                'part_number': wo_part_mapping.get(wo, 'Unknown')
            })
        # Sort by WO (ascending)
        wo_list.sort(key=lambda x: x['wo'])
        
        # Convert part number statistics to list with percentages
        part_list = []
        for part_num, stats in part_stats.items():
            pass_pct = (stats['tray_pass'] / stats['tray_total'] * 100) if stats['tray_total'] > 0 else 0
            fail_pct = (stats['tray_fail'] / stats['tray_total'] * 100) if stats['tray_total'] > 0 else 0
            
            # Calculate station percentages using helper function
            station_pcts = calculate_station_pass_percentages(stats['stations'])
            
            part_list.append({
                'part_number': part_num,
                'tray_pass': stats['tray_pass'],
                'tray_fail': stats['tray_fail'],
                'tray_total': stats['tray_total'],
                'pass_percentage': round(pass_pct, 2),
                'fail_percentage': round(fail_pct, 2),
                'stations': {k: dict(v) for k, v in stats['stations'].items()},
                'station_pass_percentages': station_pcts
            })
        part_list.sort(key=lambda x: x['part_number'])
        
        # Prepare station statistics with percentages
        station_list = []
        for station, stats in test_data['station_stats'].items():
            total = stats['pass'] + stats['fail']
            pass_pct = (stats['pass'] / total * 100) if total > 0 else 0
            fail_pct = (stats['fail'] / total * 100) if total > 0 else 0
            station_list.append({
                'station': station,
                'pass': stats['pass'],
                'fail': stats['fail'],
                'total': total,
                'pass_percentage': round(pass_pct, 2),
                'fail_percentage': round(fail_pct, 2)
            })
        # Sort stations according to custom order: FLA > FLB > AST > FTS > FCT > RIN
        def get_station_sort_key(station_item):
            station = station_item['station']
            try:
                return STATION_ORDER.index(station)
            except ValueError:
                return len(STATION_ORDER) + 1
        station_list.sort(key=get_station_sort_key)
        
        # Helper function to create heatmap rows
        def create_heatmap_rows(data_list, station_key='station_pass_percentages', 
                                id_key='wo', extra_keys=None):
            """Create heatmap rows from data list"""
            # Collect all stations
            all_stations = set()
            for item in data_list:
                all_stations.update(item[station_key].keys())
            all_stations = sort_stations(all_stations)
            
            # Create rows
            rows = []
            for item in data_list:
                row = {id_key: item[id_key]}
                if extra_keys:
                    for key in extra_keys:
                        row[key] = item.get(key, '')
                for station in all_stations:
                    row[station] = item[station_key].get(station, 0)
                rows.append(row)
            
            return rows, all_stations
        
        # Prepare heatmap data: WO x Station (Pass %)
        wo_station_heatmap, all_stations = create_heatmap_rows(
            wo_list, 
            station_key='station_pass_percentages',
            id_key='wo',
            extra_keys=['part_number']
        )
        
        # Prepare heatmap data: Part Number x Station (Pass %)
        part_station_heatmap, _ = create_heatmap_rows(
            part_list,
            station_key='station_pass_percentages',
            id_key='part_number',
            extra_keys=None
        )
        
        # Prepare 1 time pass statistics: WO grouped by Part Number
        one_time_pass_list = []
        for wo_item in wo_list:
            wo = wo_item['wo']
            part_num = wo_item['part_number']
            stats = one_time_pass_wo[wo]
            one_time_pass_pct = (stats['one_time_pass'] / stats['total'] * 100) if stats['total'] > 0 else 0
            
            one_time_pass_list.append({
                'part_number': part_num,
                'wo': wo,
                'total': stats['total'],
                'one_time_pass': stats['one_time_pass'],
                'one_time_pass_percentage': round(one_time_pass_pct, 2)
            })
        # Sort by part number first, then by WO (both ascending)
        one_time_pass_list.sort(key=lambda x: (x['part_number'], x['wo']))
        
        return jsonify({
            'total_trays': test_data['total_trays'],
            'total_pass': test_data['total_pass'],
            'total_fail': test_data['total_fail'],
            'wo_statistics': wo_list,
            'part_statistics': part_list,
            'station_statistics': station_list,
            'wo_station_heatmap': wo_station_heatmap,
            'part_station_heatmap': part_station_heatmap,
            'all_stations': all_stations,
            'one_time_pass_statistics': one_time_pass_list
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/daily-test-sn-details')
def get_sn_details():
    """Get detailed SN information for a specific WO, station, or part number"""
    try:
        start_date_str = request.args.get('start_date', '')
        end_date_str = request.args.get('end_date', '')
        wo = request.args.get('wo', '')
        station = request.args.get('station', '')
        part_number = request.args.get('part_number', '')
        status_filter = request.args.get('status', '')  # 'pass' or 'fail'
        
        if not start_date_str or not end_date_str:
            return jsonify({'error': 'Start date and end date are required'}), 400
        
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        test_data = load_daily_test_data(start_date, end_date)
        
        result = []
        for sn, test_list in test_data['sn_test_info'].items():
            wo_sn = test_list[0]['wo'] if test_list else 'No WO'
            # Determine if SN passed
            # After 2026-01-01: pass FCT = pass (all WOs)
            # Before 2026-01-01: only pass RIN = pass
            is_pass = False
            cutoff_date = datetime(2026, 1, 1).date()
            
            for test in test_list:
                test_date = test.get('date')
                if not test_date:
                    # Try to get date from filename or use current date
                    test_date = datetime.now().date()
                # Ensure test_date is a date object
                elif isinstance(test_date, datetime):
                    test_date = test_date.date()
                elif isinstance(test_date, str):
                    test_date = datetime.strptime(test_date, '%Y-%m-%d').date()
                elif isinstance(test_date, pd.Timestamp):
                    test_date = test_date.to_pydatetime().date()
                elif hasattr(test_date, 'date'):
                    test_date = test_date.date()
                elif not isinstance(test_date, date):
                    test_date = datetime.now().date()
                
                if test_date >= cutoff_date:
                    # New rule: pass FCT = pass
                    if test['station'] == 'FCT' and test['status'] == 'P':
                        is_pass = True
                        break
                else:
                    # Old rule: only pass if pass RIN
                    if test['station'] == 'RIN' and test['status'] == 'P':
                        is_pass = True
                        break
            part_numbers_for_sn = test_data['sn_part_numbers'].get(sn, [])
            
            # Apply filters
            if wo and wo != 'ALL' and wo_sn != wo:
                continue
            if part_number and part_number != 'ALL' and part_number not in part_numbers_for_sn:
                continue
            if status_filter:
                if status_filter.lower() == 'pass' and not is_pass:
                    continue
                if status_filter.lower() == 'fail' and is_pass:
                    continue
            
            # Get stations for this SN (sorted by station order)
            stations = {}
            for test in test_list:
                st = test['station']
                if st not in stations:
                    stations[st] = {'pass': 0, 'fail': 0}
                if test['status'] == 'P':
                    stations[st]['pass'] += 1
                else:
                    stations[st]['fail'] += 1
            
            # Sort stations dict keys according to custom order: FLA > FLB > AST > FTS > FCT > RIN
            sorted_station_keys = sort_stations(list(stations.keys()))
            stations_sorted = {k: stations[k] for k in sorted_station_keys}
            
            # Filter by station if specified
            if station and station != 'ALL':
                if station not in stations_sorted:
                    continue
            
            result.append({
                'sn': sn,
                'wo': wo_sn,
                'part_numbers': sorted(part_numbers_for_sn) if part_numbers_for_sn else [],  # Sort part numbers
                'status': 'PASS' if is_pass else 'FAIL',
                'stations': stations_sorted,
                'test_count': len(test_list)
            })
        
        # Sort by SN (ascending)
        result.sort(key=lambda x: x['sn'])
        
        return jsonify({
            'data': result,
            'count': len(result)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/debug-comparison')
def debug_comparison():
    """Render debug comparison page"""
    return render_template('debug_comparison.html')

@app.route('/api/debug-comparison')
def get_debug_comparison():
    """Get debug comparison data (IGS vs NV)"""
    try:
        start_date_str = request.args.get('start_date', '')
        end_date_str = request.args.get('end_date', '')
        
        if not start_date_str or not end_date_str:
            return jsonify({'error': 'Start date and end date are required'}), 400
        
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        if start_date > end_date:
            return jsonify({'error': 'Start date must be before end date'}), 400
        
        # Load Bonepile list with fail_time mapping
        bonepile_fail_time = load_bonepile_list()  # {sn: fail_time (datetime)}
        
        # Load daily test data
        test_data = load_daily_test_data(start_date, end_date)
        
        # Categorize SNs: NV debug vs IGS debug based on test date vs fail_time
        nv_debug_sns = set()  # SNs tested after fail_time
        igs_debug_sns = set()  # SNs tested before fail_time or not in Bonepile
        
        # Statistics by date
        daily_stats = defaultdict(lambda: {
            'nv_debug': {'pass': 0, 'fail': 0, 'total': 0, 'pass_rin': 0},
            'igs_debug': {'pass': 0, 'fail': 0, 'total': 0, 'pass_rin': 0}
        })
        
        # Statistics by station
        station_stats = defaultdict(lambda: {
            'nv_debug': {'pass': 0, 'fail': 0, 'total': 0},
            'igs_debug': {'pass': 0, 'fail': 0, 'total': 0}
        })
        
        # Overall statistics
        overall_stats = {
            'nv_debug': {'pass': 0, 'fail': 0, 'total': 0, 'pass_rin': 0},
            'igs_debug': {'pass': 0, 'fail': 0, 'total': 0, 'pass_rin': 0}
        }
        
        # Process each SN and each test entry
        cutoff_date = datetime(2026, 1, 1).date()
        for sn, test_list in test_data['sn_test_info'].items():
            # Get fail_time for this SN from Bonepile
            fail_time = bonepile_fail_time.get(sn)
            
            # Determine if SN passed
            # After 2026-01-01: pass FCT = pass (all WOs)
            # Before 2026-01-01: only pass RIN = pass
            is_pass = False
            
            # Check all tests for this SN
            for test_entry in test_list:
                test_date = test_entry.get('date', start_date)
                if isinstance(test_date, str):
                    test_date = datetime.strptime(test_date, '%Y-%m-%d').date()
                elif isinstance(test_date, pd.Timestamp):
                    test_date = test_date.to_pydatetime().date()
                elif hasattr(test_date, 'date'):
                    test_date = test_date.date()
                
                if test_date >= cutoff_date:
                    # New rule: pass FCT = pass
                    if test_entry['station'] == 'FCT' and test_entry['status'] == 'P':
                        is_pass = True
                        break
                else:
                    # Old rule: only pass if pass RIN
                    if test_entry['station'] == 'RIN' and test_entry['status'] == 'P':
                        is_pass = True
                        break
            
            # Process each test entry individually
            for test_entry in test_list:
                test_date = test_entry.get('date', start_date)
                if isinstance(test_date, str):
                    test_date = datetime.strptime(test_date, '%Y-%m-%d')
                elif isinstance(test_date, pd.Timestamp):
                    test_date = test_date.to_pydatetime()
                
                # Normalize test_date to date only
                if isinstance(test_date, datetime):
                    test_date = test_date.date()
                elif isinstance(test_date, str):
                    test_date = datetime.strptime(test_date, '%Y-%m-%d').date()
                elif isinstance(test_date, pd.Timestamp):
                    test_date = test_date.to_pydatetime().date()
                elif hasattr(test_date, 'date'):
                    test_date = test_date.date()
                
                # Determine debug type based on test_date vs fail_time
                # - If SN has fail_time in Bonepile:
                #   - test_date < fail_time → IGS Debug
                #   - test_date >= fail_time → NV Debug
                # - If SN does NOT have fail_time in Bonepile:
                #   - Only count as IGS Debug if has at least 1 fail
                if fail_time:
                    # Normalize fail_time to date only
                    if isinstance(fail_time, datetime):
                        fail_time_date = fail_time.date()
                    elif hasattr(fail_time, 'date'):
                        fail_time_date = fail_time.date()
                    else:
                        fail_time_date = fail_time
                    
                    if test_date < fail_time_date:
                        # Test before fail_time → IGS Debug
                        debug_type = 'igs_debug'
                        igs_debug_sns.add(sn)
                    else:
                        # Test after or on fail_time → NV Debug
                        debug_type = 'nv_debug'
                        nv_debug_sns.add(sn)
                else:
                    # SN not in Bonepile or no fail_time
                    # Count as IGS Debug if:
                    # 1. Test is a fail, OR
                    # 2. Test is pass at RIN station (before 2026), OR
                    # 3. Test is pass at FCT station (after 2026)
                    if test_entry['status'] == 'F':
                        debug_type = 'igs_debug'
                        igs_debug_sns.add(sn)
                    elif test_entry['status'] == 'P':
                        # After 2026: pass FCT = pass
                        # Before 2026: only pass RIN = pass
                        if test_date >= cutoff_date:
                            if test_entry['station'] == 'FCT':
                                debug_type = 'igs_debug'
                                igs_debug_sns.add(sn)
                            else:
                                # Skip if pass at other stations (not FCT) after 2026
                                continue
                        else:
                            if test_entry['station'] == 'RIN':
                                # Pass at RIN must be counted (Pass RIN is subset of Pass)
                                debug_type = 'igs_debug'
                                igs_debug_sns.add(sn)
                            else:
                                # Skip if pass at other stations (not RIN) before 2026
                                continue
                    else:
                        continue
                
                # Update daily stats
                date_str = test_date.strftime('%Y-%m-%d')
                daily_stats[date_str][debug_type]['total'] += 1
                if test_entry['status'] == 'P':
                    daily_stats[date_str][debug_type]['pass'] += 1
                    # Count as Pass RIN if:
                    # - Pass at RIN station (before 2026), OR
                    # - Pass at FCT station (after 2026)
                    if test_date >= cutoff_date:
                        if test_entry['station'] == 'FCT':
                            daily_stats[date_str][debug_type]['pass_rin'] += 1
                    else:
                        if test_entry['station'] == 'RIN':
                            daily_stats[date_str][debug_type]['pass_rin'] += 1
                else:
                    daily_stats[date_str][debug_type]['fail'] += 1
                
                # Update station stats
                station = test_entry['station']
                if test_entry['status'] == 'P':
                    station_stats[station][debug_type]['pass'] += 1
                else:
                    station_stats[station][debug_type]['fail'] += 1
                station_stats[station][debug_type]['total'] += 1
            
            # Track unique SNs for overall stats (count unique SNs, not test entries)
            # A SN can belong to both IGS and NV debug if it has tests before and after fail_time
            sn_has_igs_tests = False
            sn_has_nv_tests = False
            sn_has_igs_pass = False
            sn_has_nv_pass = False
            sn_has_igs_pass_rin = False
            sn_has_nv_pass_rin = False
            
            # Check all test entries to determine SN's debug types
            for test_entry in test_list:
                test_date = test_entry.get('date', start_date)
                # Normalize test_date to date only
                if isinstance(test_date, str):
                    test_date = datetime.strptime(test_date, '%Y-%m-%d').date()
                elif isinstance(test_date, pd.Timestamp):
                    test_date = test_date.to_pydatetime().date()
                elif isinstance(test_date, datetime):
                    test_date = test_date.date()
                elif hasattr(test_date, 'date'):
                    test_date = test_date.date()
                
                # Determine debug type for this test entry
                if fail_time:
                    # Normalize fail_time to date only
                    if isinstance(fail_time, datetime):
                        fail_time_date = fail_time.date()
                    elif hasattr(fail_time, 'date'):
                        fail_time_date = fail_time.date()
                    else:
                        fail_time_date = fail_time
                    if test_date < fail_time_date:
                        sn_has_igs_tests = True
                        if test_entry['status'] == 'P':
                            sn_has_igs_pass = True
                            # After 2026: pass FCT = pass_rin, Before 2026: only pass RIN = pass_rin
                            if test_date >= cutoff_date:
                                if test_entry['station'] == 'FCT':
                                    sn_has_igs_pass_rin = True
                            else:
                                if test_entry['station'] == 'RIN':
                                    sn_has_igs_pass_rin = True
                    else:
                        sn_has_nv_tests = True
                        if test_entry['status'] == 'P':
                            sn_has_nv_pass = True
                            # After 2026: pass FCT = pass_rin, Before 2026: only pass RIN = pass_rin
                            if test_date >= cutoff_date:
                                if test_entry['station'] == 'FCT':
                                    sn_has_nv_pass_rin = True
                            else:
                                if test_entry['station'] == 'RIN':
                                    sn_has_nv_pass_rin = True
                else:
                    # Not in Bonepile - only IGS if fail or pass at RIN (before 2026) or FCT (after 2026)
                    test_date_for_check = test_entry.get('date', start_date)
                    if isinstance(test_date_for_check, str):
                        test_date_for_check = datetime.strptime(test_date_for_check, '%Y-%m-%d').date()
                    elif isinstance(test_date_for_check, pd.Timestamp):
                        test_date_for_check = test_date_for_check.to_pydatetime().date()
                    elif hasattr(test_date_for_check, 'date'):
                        test_date_for_check = test_date_for_check.date()
                    
                    is_pass_station = False
                    if test_date_for_check >= cutoff_date:
                        is_pass_station = test_entry['status'] == 'P' and test_entry['station'] == 'FCT'
                    else:
                        is_pass_station = test_entry['status'] == 'P' and test_entry['station'] == 'RIN'
                    
                    if test_entry['status'] == 'F' or is_pass_station:
                        sn_has_igs_tests = True
                        if is_pass_station:
                            sn_has_igs_pass = True
                            sn_has_igs_pass_rin = True
            
            # Update overall stats (count unique SNs)
            # For summary boxes: Pass = Pass RIN (only count as pass if passed at RIN)
            if sn_has_igs_tests:
                overall_stats['igs_debug']['total'] += 1
                if sn_has_igs_pass_rin:  # Only count as pass if passed RIN
                    overall_stats['igs_debug']['pass'] += 1
                    overall_stats['igs_debug']['pass_rin'] += 1
                else:
                    overall_stats['igs_debug']['fail'] += 1
            
            if sn_has_nv_tests:
                overall_stats['nv_debug']['total'] += 1
                if sn_has_nv_pass_rin:  # Only count as pass if passed RIN
                    overall_stats['nv_debug']['pass'] += 1
                    overall_stats['nv_debug']['pass_rin'] += 1
                else:
                    overall_stats['nv_debug']['fail'] += 1
        
        # Validate daily stats: Ensure Pass RIN <= Pass for consistency
        # This ensures Pass RIN is always a subset of Pass
        for date_str in daily_stats:
            for debug_type in ['nv_debug', 'igs_debug']:
                if daily_stats[date_str][debug_type]['pass_rin'] > daily_stats[date_str][debug_type]['pass']:
                    # Cap Pass RIN at Pass value to ensure consistency
                    daily_stats[date_str][debug_type]['pass_rin'] = daily_stats[date_str][debug_type]['pass']
        
        # Convert daily_stats to list sorted by date
        daily_list = []
        current_date = start_date
        while current_date <= end_date:
            date_str = current_date.strftime('%Y-%m-%d')
            if date_str in daily_stats:
                stats = daily_stats[date_str]
                # Ensure Pass RIN <= Pass (double check for safety)
                nv_pass_rin = min(stats['nv_debug']['pass_rin'], stats['nv_debug']['pass'])
                igs_pass_rin = min(stats['igs_debug']['pass_rin'], stats['igs_debug']['pass'])
                
                daily_list.append({
                    'date': date_str,
                    'nv_debug': {
                        'pass': stats['nv_debug']['pass'],
                        'fail': stats['nv_debug']['fail'],
                        'total': stats['nv_debug']['total'],
                        'pass_rin': nv_pass_rin,
                        'pass_pct': round((stats['nv_debug']['pass'] / stats['nv_debug']['total'] * 100) if stats['nv_debug']['total'] > 0 else 0, 2),
                        'pass_rin_pct': round((nv_pass_rin / stats['nv_debug']['total'] * 100) if stats['nv_debug']['total'] > 0 else 0, 2)
                    },
                    'igs_debug': {
                        'pass': stats['igs_debug']['pass'],
                        'fail': stats['igs_debug']['fail'],
                        'total': stats['igs_debug']['total'],
                        'pass_rin': igs_pass_rin,
                        'pass_pct': round((stats['igs_debug']['pass'] / stats['igs_debug']['total'] * 100) if stats['igs_debug']['total'] > 0 else 0, 2),
                        'pass_rin_pct': round((igs_pass_rin / stats['igs_debug']['total'] * 100) if stats['igs_debug']['total'] > 0 else 0, 2)
                    }
                })
            else:
                daily_list.append({
                    'date': date_str,
                    'nv_debug': {'pass': 0, 'fail': 0, 'total': 0, 'pass_rin': 0, 'pass_pct': 0, 'pass_rin_pct': 0},
                    'igs_debug': {'pass': 0, 'fail': 0, 'total': 0, 'pass_rin': 0, 'pass_pct': 0, 'pass_rin_pct': 0}
                })
            current_date += timedelta(days=1)
        
        # Convert station_stats to list sorted by station order
        station_list = []
        for station in sort_stations(station_stats.keys()):
            stats = station_stats[station]
            nv_total = stats['nv_debug']['total']
            igs_total = stats['igs_debug']['total']
            station_list.append({
                'station': station,
                'nv_debug': {
                    'pass': stats['nv_debug']['pass'],
                    'fail': stats['nv_debug']['fail'],
                    'total': nv_total,
                    'pass_pct': round((stats['nv_debug']['pass'] / nv_total * 100) if nv_total > 0 else 0, 2),
                    'fail_pct': round((stats['nv_debug']['fail'] / nv_total * 100) if nv_total > 0 else 0, 2)
                },
                'igs_debug': {
                    'pass': stats['igs_debug']['pass'],
                    'fail': stats['igs_debug']['fail'],
                    'total': igs_total,
                    'pass_pct': round((stats['igs_debug']['pass'] / igs_total * 100) if igs_total > 0 else 0, 2),
                    'fail_pct': round((stats['igs_debug']['fail'] / igs_total * 100) if igs_total > 0 else 0, 2)
                }
            })
        
        # Validate daily stats: Ensure Pass RIN <= Pass for consistency
        # This ensures Pass RIN is always a subset of Pass
        for date_str in daily_stats:
            for debug_type in ['nv_debug', 'igs_debug']:
                if daily_stats[date_str][debug_type]['pass_rin'] > daily_stats[date_str][debug_type]['pass']:
                    # Cap Pass RIN at Pass value to ensure consistency
                    daily_stats[date_str][debug_type]['pass_rin'] = daily_stats[date_str][debug_type]['pass']
        
        # Pass = Pass RIN for summary boxes (they are the same now)
        # Calculate overall percentages
        nv_total = overall_stats['nv_debug']['total']
        igs_total = overall_stats['igs_debug']['total']
        
        # Set pass = pass_rin for summary boxes consistency
        overall_stats['nv_debug']['pass'] = overall_stats['nv_debug']['pass_rin']
        overall_stats['igs_debug']['pass'] = overall_stats['igs_debug']['pass_rin']
        
        overall_stats['nv_debug']['pass_pct'] = round((overall_stats['nv_debug']['pass_rin'] / nv_total * 100) if nv_total > 0 else 0, 2)
        overall_stats['nv_debug']['fail_pct'] = round((overall_stats['nv_debug']['fail'] / nv_total * 100) if nv_total > 0 else 0, 2)
        overall_stats['nv_debug']['pass_rin_pct'] = round((overall_stats['nv_debug']['pass_rin'] / nv_total * 100) if nv_total > 0 else 0, 2)
        
        overall_stats['igs_debug']['pass_pct'] = round((overall_stats['igs_debug']['pass_rin'] / igs_total * 100) if igs_total > 0 else 0, 2)
        overall_stats['igs_debug']['fail_pct'] = round((overall_stats['igs_debug']['fail'] / igs_total * 100) if igs_total > 0 else 0, 2)
        overall_stats['igs_debug']['pass_rin_pct'] = round((overall_stats['igs_debug']['pass_rin'] / igs_total * 100) if igs_total > 0 else 0, 2)
        
        return jsonify({
            'overall_stats': overall_stats,
            'daily_stats': daily_list,
            'station_stats': station_list,
            'nv_debug_count': len(nv_debug_sns),
            'igs_debug_count': len(igs_debug_sns),
            'all_stations': sort_stations(list(station_stats.keys()))
        })
    except Exception as e:
        import traceback
        print(f"[ERROR] Debug comparison: {e}", flush=True)
        print(traceback.format_exc(), flush=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/debug-comparison-sn-list')
def get_debug_comparison_sn_list():
    """Get SN list for debug comparison by debug type and status"""
    try:
        start_date_str = request.args.get('start_date', '')
        end_date_str = request.args.get('end_date', '')
        debug_type = request.args.get('debug_type', '')  # 'nv_debug' or 'igs_debug'
        status = request.args.get('status', '')  # 'pass', 'fail', or 'all'
        
        if not start_date_str or not end_date_str or not debug_type:
            return jsonify({'error': 'Start date, end date, and debug_type are required'}), 400
        
        if debug_type not in ['nv_debug', 'igs_debug']:
            return jsonify({'error': 'debug_type must be nv_debug or igs_debug'}), 400
        
        if status not in ['pass', 'fail', 'all']:
            return jsonify({'error': 'status must be pass, fail, or all'}), 400
        
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        if start_date > end_date:
            return jsonify({'error': 'Start date must be before end date'}), 400
        
        # Load Bonepile list with fail_time mapping
        bonepile_fail_time = load_bonepile_list()
        
        # Load daily test data
        test_data = load_daily_test_data(start_date, end_date)
        
        # Get SNs that passed RIN
        sn_pass_rin = test_data.get('sn_pass_rin', set())
        
        result_sns = []
        
        # Process each SN
        for sn, test_list in test_data['sn_test_info'].items():
            fail_time = bonepile_fail_time.get(sn)
            
            # Determine if this SN belongs to the requested debug type
            sn_belongs_to_debug_type = False
            sn_has_pass_rin = sn in sn_pass_rin
            
            # Check all test entries to determine SN's debug type
            for test_entry in test_list:
                test_date = test_entry.get('date', start_date)
                if isinstance(test_date, str):
                    test_date = datetime.strptime(test_date, '%Y-%m-%d')
                elif isinstance(test_date, pd.Timestamp):
                    test_date = test_date.to_pydatetime()
                test_date = test_date.replace(hour=0, minute=0, second=0, microsecond=0)
                
                if fail_time:
                    fail_time_date = fail_time.replace(hour=0, minute=0, second=0, microsecond=0)
                    if test_date < fail_time_date:
                        if debug_type == 'igs_debug':
                            sn_belongs_to_debug_type = True
                            break
                    else:
                        if debug_type == 'nv_debug':
                            sn_belongs_to_debug_type = True
                            break
                else:
                    # Not in Bonepile - only IGS if fail or pass at RIN
                    if debug_type == 'igs_debug':
                        if test_entry['status'] == 'F' or (test_entry['status'] == 'P' and test_entry['station'] == 'RIN'):
                            sn_belongs_to_debug_type = True
                            break
            
            if not sn_belongs_to_debug_type:
                continue
            
            # Filter by status
            if status == 'pass' and not sn_has_pass_rin:
                continue
            if status == 'fail' and sn_has_pass_rin:
                continue
            
            result_sns.append(sn)
        
        # Sort SNs
        result_sns.sort()
        
        return jsonify({
            'count': len(result_sns),
            'sns': result_sns
        })
    except Exception as e:
        import traceback
        print(f"[ERROR] Debug comparison SN list: {e}", flush=True)
        print(traceback.format_exc(), flush=True)
        return jsonify({'error': str(e)}), 500

# Context processor to inject show_daily_test_button=False for all templates
@app.context_processor
def inject_daily_test_button():
    """Inject show_daily_test_button=False into all templates (default for main app)"""
    return dict(show_daily_test_button=False)

@app.route('/hourly-report')
def hourly_report():
    """Main page for hourly report"""
    return render_template('hourly_report.html', ip=get_local_ip())

@app.route('/api/hourly-report-data', methods=['POST'])
def get_hourly_report_data():
    """
    Get hourly report data for selected datetime range
    Request: {'start_datetime': '2026-01-10 09:00', 'end_datetime': '2026-01-10 17:00', 'include_sns': true/false}
    """
    try:
        data = request.json
        start_datetime_str = data.get('start_datetime')
        end_datetime_str = data.get('end_datetime')
        include_sns = data.get('include_sns', True)
        
        if not start_datetime_str or not end_datetime_str:
            return jsonify({'success': False, 'error': 'Start datetime and end datetime are required'}), 400
        
        # Parse datetime strings
        start_datetime = datetime.strptime(start_datetime_str, '%Y-%m-%d %H:%M')
        end_datetime = datetime.strptime(end_datetime_str, '%Y-%m-%d %H:%M')
        
        if end_datetime <= start_datetime:
            return jsonify({'success': False, 'error': 'End datetime must be after start datetime'}), 400
        
        try:
            processed_data = load_hourly_report_data(start_datetime_str, end_datetime_str)
            
            # Ensure data has 'statistics' key for consistency
            if 'statistics' not in processed_data:
                processed_data = {'statistics': processed_data}
            
            # Don't save to cache - hourly report should always be fresh
            
            return jsonify({
                'success': True,
                'data': processed_data,
                'cached': False
            })
        except Exception as load_error:
            import traceback
            print(f"Error loading hourly report data: {load_error}", flush=True)
            print(traceback.format_exc(), flush=True)
            # Return empty data structure instead of error
            empty_data = {
                'statistics': {
                    'all': {'total_sns': 0, 'pass_count': 0, 'fail_count': 0, 'bonepile': 0, 'fresh': 0, 'pass_rate': 0},
                    'bonepile': {'total_sns': 0, 'pass_count': 0, 'fail_count': 0, 'pass_rate': 0},
                    'igs': {'total_sns': 0, 'pass_count': 0, 'fail_count': 0, 'pass_rate': 0}
                },
                'sn_details': {}
            }
            return jsonify({
                'success': True,
                'data': empty_data,
                'cached': False,
                'warning': f'No data found or error loading: {str(load_error)}'
            })
    except Exception as e:
        import traceback
        print(f"Error in get_hourly_report_data: {e}", flush=True)
        print(traceback.format_exc(), flush=True)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/hourly-report-sn-list', methods=['POST'])
def get_hourly_report_sn_list():
    """
    Get SN list for specific category
    Request: {'start_datetime': '2026-01-10 09:00', 'end_datetime': '2026-01-10 17:00', 'category': 'all', 'type': 'total'}
    category: 'all', 'bonepile', 'igs'
    type: 'total', 'pass', 'fail', 'pass_rate'
    """
    try:
        data = request.json
        start_datetime_str = data.get('start_datetime')
        end_datetime_str = data.get('end_datetime')
        category = data.get('category', 'all')
        type_filter = data.get('type', 'total')
        
        if not start_datetime_str or not end_datetime_str:
            return jsonify({'success': False, 'error': 'Start datetime and end datetime are required'}), 400
        
        # Load hourly report data
        try:
            report_data = load_hourly_report_data(start_datetime_str, end_datetime_str)
            sn_details = report_data.get('sn_details', {})
        except Exception as e:
            import traceback
            print(f"Error loading hourly report data in SN list: {e}", flush=True)
            print(traceback.format_exc(), flush=True)
            return jsonify({'success': False, 'error': str(e)}), 500
        
        # Filter SNs based on category and type
        result_sns = []
        
        for sn, details in sn_details.items():
            # Filter by category
            is_bonepile = details.get('bonepile', False)
            if category == 'bonepile' and not is_bonepile:
                continue
            if category == 'igs' and is_bonepile:
                continue
            
            # Filter by type
            pass_fail = details.get('pass_fail', 'FAIL')
            if type_filter == 'pass' and pass_fail != 'PASS':
                continue
            if type_filter == 'fail' and pass_fail != 'FAIL':
                continue
            # type_filter == 'total' or 'pass_rate' includes all
            
            # Get part numbers and stations
            part_numbers = details.get('part_numbers', set())
            stations = details.get('stations', set())
            
            result_sns.append({
                'sn': sn,
                'bonepile': is_bonepile,
                'pass_fail': pass_fail,
                'part_numbers': sorted(list(part_numbers)) if isinstance(part_numbers, set) else (part_numbers if isinstance(part_numbers, list) else []),
                'stations': sorted(list(stations)) if isinstance(stations, set) else (stations if isinstance(stations, list) else [])
            })
        
        # Sort by SN
        result_sns.sort(key=lambda x: x['sn'])
        
        return jsonify({
            'success': True,
            'sns': result_sns,
            'count': len(result_sns)
        })
    except Exception as e:
        import traceback
        print(f"Error in get_hourly_report_sn_list: {e}", flush=True)
        print(traceback.format_exc(), flush=True)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/hourly-report-download')
def download_hourly_report_csv():
    """
    Download hourly report as CSV
    """
    try:
        start_datetime_str = request.args.get('start_datetime')
        end_datetime_str = request.args.get('end_datetime')
        
        if not start_datetime_str or not end_datetime_str:
            return jsonify({'error': 'Start datetime and end datetime are required'}), 400
        
        # Load hourly report data
        report_data = load_hourly_report_data(start_datetime_str, end_datetime_str)
        sn_details = report_data.get('sn_details', {})
        
        # Create CSV content
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(['SN', 'Bonepile', 'Pass/Fail', 'Part Numbers', 'Stations', 'Test Count'])
        
        # Data rows
        for sn, details in sorted(sn_details.items()):
            # Get stations and sort them according to STATION_ORDER
            stations_list = details.get('stations', [])
            if isinstance(stations_list, set):
                stations_list = list(stations_list)
            sorted_stations = sort_stations(stations_list)
            
            writer.writerow([
                sn,
                'Yes' if details.get('bonepile', False) else 'No',
                details.get('pass_fail', 'FAIL'),
                ', '.join(sorted(details.get('part_numbers', []))),
                ', '.join(sorted_stations),
                len(details.get('tests', []))
            ])
        
        # Create response
        from flask import Response
        # Create filename from datetime range
        filename_start = start_datetime_str.replace(' ', '_').replace(':', '')
        filename_end = end_datetime_str.replace(' ', '_').replace(':', '')
        response = Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=hourly_report_{filename_start}_to_{filename_end}.csv'}
        )
        
        return response
    except Exception as e:
        import traceback
        print(f"Error in download_hourly_report_csv: {e}", flush=True)
        print(traceback.format_exc(), flush=True)
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    cleanup_old_cache()
    
    local_ip = get_local_ip()
    port = 5001
    print("=" * 80)
    print("VR-TS Bonepile Statistics Dashboard")
    print("=" * 80)
    print(f"Starting server...")
    print(f"Local access: http://localhost:{port}")
    print(f"Network access: http://{local_ip}:{port}")
    print("=" * 80)
    app.run(debug=True, host='0.0.0.0', port=port)

