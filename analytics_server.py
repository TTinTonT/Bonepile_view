#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Standalone Analytics Dashboard server (port 5555).

- Uses UI/UX from analytics_templates/analytics_dashboard.html (copied from Analytics_Dashboard.html)
- Scans Oberon L10 zip logs into SQLite cache (raw table)
- Computes fast aggregates from SQLite for UI (daily/weekly/monthly + summary matrix + SKU table)

This server is intentionally separate from app.py / daily_test_app.py.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import sqlite3
import threading
import time
import csv
import io
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pytz
from flask import Flask, Response, jsonify, render_template, request

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None


# -----------------------------
# Config
# -----------------------------

BASE_PATH = r"\\10.16.137.111\Oberon\L10"

APP_DIR = os.path.dirname(os.path.abspath(__file__))
ANALYTICS_CACHE_DIR = os.path.join(APP_DIR, "analytics_cache")
DB_PATH = os.path.join(ANALYTICS_CACHE_DIR, "analytics.db")
STATE_PATH = os.path.join(ANALYTICS_CACHE_DIR, "raw_state.json")

# Uploaded NV/IGS bonepile workbook (single file; replaced on each upload)
BONEPILE_UPLOAD_PATH = os.path.join(ANALYTICS_CACHE_DIR, "bonepile_upload.xlsx")
BONEPILE_ALLOWED_SHEETS = ["VR-TS1", "TS2-SKU002", "TS2-SKU010"]
BONEPILE_REQUIRED_FIELDS = ["sn", "nv_disposition", "status", "pic", "igs_action", "igs_status"]

# Excel export templates (formatting preserved in exported XLSX)
TEMPLATES_DIR = os.path.join(APP_DIR, "templates")
SKU_SUMMARY_TEMPLATE_PATH = os.path.join(TEMPLATES_DIR, "SKU_Summary.xlsx")
TRAY_SUMMARY_TEMPLATE_PATH = os.path.join(TEMPLATES_DIR, "Tray_Summary_Template.xlsx")

CA_TZ = pytz.timezone("America/Los_Angeles")
TW_TZ = pytz.timezone("Asia/Taipei")

AUTO_SCAN_EVERY_SECONDS = 60  # Auto-scan every 1 minute
# Each auto-scan: refresh last N minutes (delete cache in that window, then rescan) so data is always fresh.
REFRESH_WINDOW_MINUTES = 180  # 3 hours
RETENTION_DAYS = 90

# IMPORTANT:
# Oberon zip filenames end with "...YYYYMMDDTHHMMSSZ" but in this environment the timestamp
# should be treated as *California local time* (PST/PDT), matching the hourly report logic.
# If you change this mode, cached rows must be rebuilt because computed CA fields will change.
TIMESTAMP_MODE = "ca_local_suffix_v3"

# Final pass rules
PASS_AT_FCT_PART_NUMBERS = {
    "675-24109-0010-TS2",
}


# -----------------------------
# Flask app
# -----------------------------

app = Flask(
    __name__,
    template_folder=os.path.join(APP_DIR, "analytics_templates"),
)


# -----------------------------
# Auto scan / retention status (shared with UI)
# -----------------------------


auto_status_lock = threading.Lock()
next_auto_scan_ms: Optional[int] = None
last_retention_cleanup_ms: Optional[int] = None


# -----------------------------
# Utilities
# -----------------------------


def ensure_dirs() -> None:
    os.makedirs(ANALYTICS_CACHE_DIR, exist_ok=True)


def utc_ms(dt: datetime) -> int:
    """Convert aware datetime to epoch milliseconds."""
    if dt.tzinfo is None:
        raise ValueError("utc_ms expects tz-aware datetime")
    return int(dt.timestamp() * 1000)


def parse_timestamp_from_filename(filename: str) -> Optional[datetime]:
    """
    Parse timestamp from suffix like: 20260107T160130Z.

    NOTE: Even though the suffix ends with "Z", for this project we treat it as
    California local time (PST/PDT), matching the hourly report implementation.

    Returns tz-aware CA datetime.
    """
    try:
        m = re.search(r"(\d{8})T(\d{6})Z", filename)
        if not m:
            return None
        dt_naive = datetime.strptime(f"{m.group(1)}T{m.group(2)}", "%Y%m%dT%H%M%S")
        return CA_TZ.localize(dt_naive)
    except Exception:
        return None


def _parse_ca_input_datetime(s: str, *, is_end: bool) -> Optional[datetime]:
    """
    Parse user-provided datetime string in CA timezone.
    Accepts:
      - YYYY-MM-DD HH:MM
      - YYYY-MM-DD HH:MM:SS

    For end times with only minute precision, treat as inclusive of that minute
    by setting seconds to 59.
    """
    if not s:
        return None
    s = str(s).strip()
    try:
        if re.search(r"\d:\d\d:\d\d$", s):
            dt = datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
            return CA_TZ.localize(dt)
        dt = datetime.strptime(s, "%Y-%m-%d %H:%M")
        dt_ca = CA_TZ.localize(dt)
        if is_end:
            dt_ca = dt_ca + timedelta(seconds=59)
        return dt_ca
    except Exception:
        return None


def convert_to_ca_time(dt: datetime) -> datetime:
    """
    Convert a datetime to CA tz-aware datetime.

    - If dt is naive, assume it is already CA local time and localize it.
    - If dt is tz-aware, convert it to CA timezone.
    """
    if dt.tzinfo is None:
        return CA_TZ.localize(dt)
    return dt.astimezone(CA_TZ)


def ca_fields_from_utc(utc_dt: datetime) -> Tuple[int, str, int, str, str]:
    """
    Return:
    - ca_ms (epoch ms of the instant)
    - ca_date (YYYY-MM-DD in CA)
    - ca_hour (0-23 in CA)
    - ca_week (Sunday-start week in CA, formatted "YYYY-MM-DD~YYYY-MM-DD")
    - ca_month (YYYY-MM in CA)
    """
    # Despite the name, accept any tz-aware datetime and normalize to CA.
    ca_dt = convert_to_ca_time(utc_dt)
    ca_ms = utc_ms(ca_dt)
    ca_date = ca_dt.strftime("%Y-%m-%d")
    ca_hour = int(ca_dt.strftime("%H"))
    # Week starts on Sunday and ends on Saturday.
    # Python weekday(): Monday=0..Sunday=6, so days_since_sunday maps Sunday->0, Monday->1, ..., Saturday->6.
    days_since_sunday = (ca_dt.weekday() + 1) % 7
    week_start = (ca_dt - timedelta(days=days_since_sunday)).date()
    week_end = (week_start + timedelta(days=6))
    ca_week = f"{week_start.strftime('%Y-%m-%d')}~{week_end.strftime('%Y-%m-%d')}"
    ca_month = ca_dt.strftime("%Y-%m")
    return ca_ms, ca_date, ca_hour, ca_week, ca_month


def parse_source_token(filename: str) -> Tuple[Optional[int], Optional[str]]:
    """
    Parse Bonepile marker token right after "IGSJ_".
    Examples:
    - IGSJ_NA_... => is_bonepile=0
    - IGSJ_PB-71108_... => is_bonepile=1, pb_id="PB-71108"
    """
    try:
        m = re.match(r"^IGSJ_([^_]+)_", filename)
        if not m:
            return None, None
        token = (m.group(1) or "").strip()
        token_upper = token.upper()
        if token_upper == "NA":
            return 0, None
        if token_upper.startswith("PB-"):
            return 1, token
        return None, None
    except Exception:
        return None, None


def extract_part_number_from_filename(filename: str) -> str:
    """
    Extract part number from filename (normalized). Examples:
    IGSJ_PB-71108_675-24109-0002-TS1_1830126000087_P_FLA_20260107T163248Z.zip
      => 675-24109-0002-TS1
    """
    name = filename.replace(".zip", "")
    # PB-XXXX_XXX-XXXXX-XXXX-TS<NUM>
    m = re.search(r"PB-\d+_(\d+-\d+-\d+-TS\d+)", name)
    if m:
        return m.group(1)
    # PB-XXXX_XXX-XXXXX-XXXX
    m = re.search(r"PB-\d+_(\d+-\d+-\d+)", name)
    if m:
        return m.group(1)
    # XXX-XXXXX-XXXX-TS<NUM>
    m = re.search(r"(\d+-\d+-\d+-TS\d+)", name)
    if m:
        return m.group(1)
    # XXX-XXXXX-XXXX
    m = re.search(r"(\d+-\d+-\d+)", name)
    if m:
        return m.group(1)
    return "Unknown"


def parse_test_filename(filename: str) -> Optional[Tuple[str, str, str, str]]:
    """
    Parse filename to extract:
      SN (string of digits)
      status: 'F' or 'P'
      station: e.g. RIN, FLA...
      part_number: extracted from file name
    """
    name = filename.replace(".zip", "")
    part_number = extract_part_number_from_filename(filename)

    # Pattern 1: _SN_Status_Station_
    m = re.search(r"_(\d{10,})_([FP])_([A-Z0-9]+)_", name)
    if m:
        sn, status, station = m.group(1), m.group(2), m.group(3)
        if sn.startswith("18") and len(sn) == 13:
            return sn, status, station, part_number

    # Pattern 2: find SN anywhere then find _Status_Station_
    sn_match = re.search(r"(18\d{11})", name)
    if sn_match:
        sn = sn_match.group(1)
        after_sn = name[name.find(sn) + len(sn) :]
        m2 = re.search(r"_([FP])_([A-Z0-9]+)_", after_sn)
        if m2:
            status, station = m2.group(1), m2.group(2)
            return sn, status, station, part_number

    return None


def get_pass_station_for_part_number(part_number: str) -> str:
    pn = "" if part_number is None else str(part_number).strip().upper()
    if pn in PASS_AT_FCT_PART_NUMBERS:
        return "FCT"
    if "TS2" in pn:
        return "NVL"
    return "FCT"


def is_final_pass(sn_status: str, station: str, part_number: str) -> bool:
    """Final pass for hourly-style tray logic (no legacy fallback here)."""
    if sn_status != "P":
        return False
    st = str(station or "").strip().upper()
    pn = "" if part_number is None else str(part_number).strip()
    if not pn or pn.lower() == "unknown":
        # If unknown, treat as not a final pass (keep it strict for analytics app)
        return False
    return st == get_pass_station_for_part_number(pn)


def ca_range_to_tw_dates(start_ca: datetime, end_ca: datetime) -> List[datetime.date]:
    """
    Map a CA datetime range to a list of Taiwan dates to scan, with a small safety margin.
    """
    if start_ca.tzinfo is None:
        start_ca = CA_TZ.localize(start_ca)
    if end_ca.tzinfo is None:
        end_ca = CA_TZ.localize(end_ca)
    start_tw = start_ca.astimezone(TW_TZ)
    end_tw = end_ca.astimezone(TW_TZ)
    d0 = start_tw.date()
    d1 = end_tw.date()
    # safety: scan one day before/after to avoid boundary issues
    d0 = d0 - timedelta(days=1)
    d1 = d1 + timedelta(days=1)
    days: List[datetime.date] = []
    cur = d0
    while cur <= d1:
        days.append(cur)
        cur += timedelta(days=1)
    return days


# -----------------------------
# SQLite cache
# -----------------------------


db_init_lock = threading.Lock()
db_initialized = False


def ensure_db_ready(force: bool = False) -> None:
    """
    Ensure analytics cache directory + SQLite schema exist.

    Why:
    - Users may delete analytics.db/raw_state.json to "reset".
    - SQLite will recreate an empty file, but schema will be missing -> runtime errors.
    This function makes the server self-healing.
    """
    global db_initialized
    with db_init_lock:
        if db_initialized and not force:
            return
        ensure_dirs()
        init_db()
        db_initialized = True


def connect_db() -> sqlite3.Connection:
    # IMPORTANT:
    # connect_db must NOT call ensure_db_ready() to avoid recursive init:
    # ensure_db_ready -> init_db -> connect_db -> ensure_db_ready ...
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    # IMPORTANT:
    # init_db must open SQLite directly (not via connect_db) so ensure_db_ready() can call init_db safely.
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    try:
        # DB metadata (used for cache compatibility checks)
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS meta (
              key TEXT PRIMARY KEY,
              value TEXT NOT NULL
            );
            """
        )
        conn.commit()

        # If timestamp interpretation changes, cached computed CA fields are invalid.
        row = conn.execute("SELECT value FROM meta WHERE key = 'timestamp_mode';").fetchone()
        old_mode = row["value"] if row else None

        # Detect legacy caches (no meta row yet) that were built with the old behavior.
        raw_table_exists = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='raw_entries';"
        ).fetchone()
        raw_has_rows = False
        if raw_table_exists:
            try:
                raw_has_rows = bool(
                    conn.execute("SELECT 1 FROM raw_entries LIMIT 1;").fetchone()
                )
            except Exception:
                raw_has_rows = False

        needs_reset = (old_mode is None and raw_has_rows) or (old_mode is not None and old_mode != TIMESTAMP_MODE)
        if needs_reset:
            # Hard reset raw cache tables + scan state (safe: cache only).
            conn.execute("DROP TABLE IF EXISTS raw_entries;")
            conn.execute(
                "INSERT OR REPLACE INTO meta (key, value) VALUES (?, ?);",
                ("timestamp_mode", TIMESTAMP_MODE),
            )
            conn.commit()
            try:
                RawState().save()
            except Exception:
                pass
        elif old_mode is None:
            # Fresh DB (or empty cache): record the current mode.
            conn.execute(
                "INSERT OR REPLACE INTO meta (key, value) VALUES (?, ?);",
                ("timestamp_mode", TIMESTAMP_MODE),
            )
            conn.commit()

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS raw_entries (
              utc_ms INTEGER NOT NULL,
              ca_ms INTEGER NOT NULL,
              ca_date TEXT NOT NULL,
              ca_hour INTEGER NOT NULL,
              ca_week TEXT NOT NULL,
              ca_month TEXT NOT NULL,
              filename TEXT NOT NULL,
              folder_path TEXT NOT NULL,
              sn TEXT NOT NULL,
              status TEXT NOT NULL,
              station TEXT NOT NULL,
              part_number TEXT NOT NULL,
              is_bonepile INTEGER,
              pb_id TEXT,
              PRIMARY KEY (utc_ms, filename)
            );
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_raw_ca_ms ON raw_entries (ca_ms);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_raw_sn_ca ON raw_entries (sn, ca_ms);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_raw_ca_date ON raw_entries (ca_date);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_raw_ca_week ON raw_entries (ca_week);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_raw_ca_month ON raw_entries (ca_month);")

        # NV/IGS workbook parsed rows (per sheet)
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS bonepile_entries (
              sheet TEXT NOT NULL,
              excel_row INTEGER NOT NULL,
              sn TEXT NOT NULL,
              nvpn TEXT,
              status TEXT,
              pic TEXT,
              igs_status TEXT,
              nv_disposition TEXT,
              igs_action TEXT,
              nv_dispo_count INTEGER,
              igs_action_count INTEGER,
              updated_at_ca_ms INTEGER,
              PRIMARY KEY (sheet, excel_row)
            );
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bp_sn ON bonepile_entries (sn);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bp_sheet_sn ON bonepile_entries (sheet, sn);")
        conn.commit()
    finally:
        conn.close()


def get_db_data_range_ca_ms() -> Tuple[Optional[int], Optional[int]]:
    """
    Return (min_ca_ms, max_ca_ms) based on actual rows in SQLite.
    This reflects the newest/oldest data we truly have (not the requested scan range).
    """
    try:
        conn = connect_db()
        try:
            row = conn.execute(
                "SELECT MIN(ca_ms) AS min_ca_ms, MAX(ca_ms) AS max_ca_ms FROM raw_entries;"
            ).fetchone()
            if not row:
                return None, None
            return row["min_ca_ms"], row["max_ca_ms"]
        finally:
            conn.close()
    except sqlite3.OperationalError:
        # Likely schema missing (db was deleted). Re-init once and retry.
        ensure_db_ready(force=True)
        conn = connect_db()
        try:
            row = conn.execute(
                "SELECT MIN(ca_ms) AS min_ca_ms, MAX(ca_ms) AS max_ca_ms FROM raw_entries;"
            ).fetchone()
            if not row:
                return None, None
            return row["min_ca_ms"], row["max_ca_ms"]
        finally:
            conn.close()


# -----------------------------
# Raw state
# -----------------------------


@dataclass
class RawState:
    min_ca_ms: Optional[int] = None
    max_ca_ms: Optional[int] = None
    min_key: Optional[Tuple[int, str]] = None  # (utc_ms, filename)
    max_key: Optional[Tuple[int, str]] = None  # (utc_ms, filename)
    min_path: Optional[str] = None
    max_path: Optional[str] = None
    last_scan_ca_ms: Optional[int] = None
    # Record full-day rescan runs by hour -> YYYY-MM-DD (CA) to avoid repeating after restarts.
    full_day_runs: Optional[Dict[str, str]] = None
    # NV/IGS workbook upload + mapping + parse status
    bonepile_file: Optional[Dict[str, Any]] = None
    bonepile_mapping: Optional[Dict[str, Any]] = None  # per-sheet mapping config
    bonepile_sheet_status: Optional[Dict[str, Any]] = None  # per-sheet parse status/result

    @staticmethod
    def load() -> "RawState":
        if not os.path.exists(STATE_PATH):
            return RawState()
        try:
            with open(STATE_PATH, "r", encoding="utf-8") as f:
                data = json.load(f) or {}
            return RawState(
                min_ca_ms=data.get("min_ca_ms"),
                max_ca_ms=data.get("max_ca_ms"),
                min_key=tuple(data["min_key"]) if isinstance(data.get("min_key"), list) else None,
                max_key=tuple(data["max_key"]) if isinstance(data.get("max_key"), list) else None,
                min_path=data.get("min_path"),
                max_path=data.get("max_path"),
                last_scan_ca_ms=data.get("last_scan_ca_ms"),
                full_day_runs=data.get("full_day_runs") if isinstance(data.get("full_day_runs"), dict) else None,
                bonepile_file=data.get("bonepile_file") if isinstance(data.get("bonepile_file"), dict) else None,
                bonepile_mapping=data.get("bonepile_mapping") if isinstance(data.get("bonepile_mapping"), dict) else None,
                bonepile_sheet_status=data.get("bonepile_sheet_status")
                if isinstance(data.get("bonepile_sheet_status"), dict)
                else None,
            )
        except Exception:
            return RawState()

    def save(self) -> None:
        tmp = STATE_PATH + ".tmp"
        data = {
            "min_ca_ms": self.min_ca_ms,
            "max_ca_ms": self.max_ca_ms,
            "min_key": list(self.min_key) if self.min_key else None,
            "max_key": list(self.max_key) if self.max_key else None,
            "min_path": self.min_path,
            "max_path": self.max_path,
            "last_scan_ca_ms": self.last_scan_ca_ms,
            "full_day_runs": self.full_day_runs or None,
            "bonepile_file": self.bonepile_file or None,
            "bonepile_mapping": self.bonepile_mapping or None,
            "bonepile_sheet_status": self.bonepile_sheet_status or None,
        }
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        if os.path.exists(STATE_PATH):
            os.remove(STATE_PATH)
        os.replace(tmp, STATE_PATH)


# -----------------------------
# Scanning
# -----------------------------


scan_lock = threading.Lock()


def iter_zip_files_for_tw_date(tw_date: datetime.date) -> Iterable[Tuple[str, str]]:
    """
    Yield (folder_path, filename) for all zip files under a Taiwan date folder.
    Structure:
      BASE_PATH/YYYY/MM/DD/<subfolder-id>/**/*.zip
    """
    dir_path = os.path.join(
        BASE_PATH,
        tw_date.strftime("%Y"),
        tw_date.strftime("%m"),
        tw_date.strftime("%d"),
    )
    if not os.path.isdir(dir_path):
        return

    try:
        # Robust: walk the whole day folder recursively (structure can change).
        for root, _dirs, files in os.walk(dir_path):
            for fn in files:
                if fn.lower().endswith(".zip"):
                    yield root, fn
    except Exception:
        return


def insert_entries(entries: List[Dict[str, Any]]) -> int:
    if not entries:
        return 0
    try:
        conn = connect_db()
        try:
            before = conn.total_changes
            cur = conn.cursor()
            cur.executemany(
                """
                INSERT OR IGNORE INTO raw_entries (
                  utc_ms, ca_ms, ca_date, ca_hour, ca_week, ca_month,
                  filename, folder_path, sn, status, station, part_number,
                  is_bonepile, pb_id
                ) VALUES (
                  :utc_ms, :ca_ms, :ca_date, :ca_hour, :ca_week, :ca_month,
                  :filename, :folder_path, :sn, :status, :station, :part_number,
                  :is_bonepile, :pb_id
                );
                """,
                entries,
            )
            conn.commit()
            # Use total_changes delta because sqlite3 rowcount is unreliable for executemany+INSERT OR IGNORE.
            return int(conn.total_changes - before)
        finally:
            conn.close()
    except sqlite3.OperationalError:
        # Likely schema missing (db was deleted). Re-init once and retry.
        ensure_db_ready(force=True)
        conn = connect_db()
        try:
            before = conn.total_changes
            cur = conn.cursor()
            cur.executemany(
                """
                INSERT OR IGNORE INTO raw_entries (
                  utc_ms, ca_ms, ca_date, ca_hour, ca_week, ca_month,
                  filename, folder_path, sn, status, station, part_number,
                  is_bonepile, pb_id
                ) VALUES (
                  :utc_ms, :ca_ms, :ca_date, :ca_hour, :ca_week, :ca_month,
                  :filename, :folder_path, :sn, :status, :station, :part_number,
                  :is_bonepile, :pb_id
                );
                """,
                entries,
            )
            conn.commit()
            return int(conn.total_changes - before)
        finally:
            conn.close()


def scan_range(start_ca: datetime, end_ca: datetime, state: RawState) -> Dict[str, Any]:
    """
    Scan Oberon files for a CA datetime range and store parsed entries into SQLite.
    Updates state min/max coverage and keys.
    """
    if start_ca.tzinfo is None:
        start_ca = CA_TZ.localize(start_ca)
    if end_ca.tzinfo is None:
        end_ca = CA_TZ.localize(end_ca)
    # Never scan beyond "now" in CA time.
    # Keep second precision for near-real-time updates.
    now_ca = datetime.now(CA_TZ).replace(microsecond=0)
    if end_ca > now_ca:
        end_ca = now_ca
    if start_ca > now_ca:
        return {"ok": False, "error": "start is in the future"}
    if end_ca <= start_ca:
        return {"ok": False, "error": "end must be after start"}

    tw_dates = ca_range_to_tw_dates(start_ca, end_ca)
    new_rows = 0
    seen_min_key = state.min_key
    seen_max_key = state.max_key
    seen_min_path = state.min_path
    seen_max_path = state.max_path

    start_ca_ms = utc_ms(start_ca)
    end_ca_ms = utc_ms(end_ca)

    batch: List[Dict[str, Any]] = []
    visited_zip = 0
    parsed_ok = 0
    ts_ok = 0
    in_range = 0

    def flush():
        nonlocal batch, new_rows
        if not batch:
            return
        inserted = insert_entries(batch)
        new_rows += int(inserted)
        batch = []

    for tw_date in tw_dates:
        for folder_path, fn in iter_zip_files_for_tw_date(tw_date):
            visited_zip += 1
            parsed = parse_test_filename(fn)
            if not parsed:
                continue
            parsed_ok += 1
            sn, status, station, part_number = parsed
            utc_dt = parse_timestamp_from_filename(fn)
            if not utc_dt:
                continue
            ts_ok += 1
            ca_ms, ca_date, ca_hour, ca_week, ca_month = ca_fields_from_utc(utc_dt)
            if ca_ms < start_ca_ms or ca_ms > end_ca_ms:
                continue
            in_range += 1

            is_bp, pb_id = parse_source_token(fn)
            key = (int(utc_ms(utc_dt)), fn)
            # Track min/max key/path for state
            if seen_min_key is None or key < seen_min_key:
                seen_min_key = key
                seen_min_path = folder_path
            if seen_max_key is None or key > seen_max_key:
                seen_max_key = key
                seen_max_path = folder_path

            batch.append(
                {
                    "utc_ms": int(utc_ms(utc_dt)),
                    "ca_ms": int(ca_ms),
                    "ca_date": ca_date,
                    "ca_hour": int(ca_hour),
                    "ca_week": ca_week,
                    "ca_month": ca_month,
                    "filename": fn,
                    "folder_path": folder_path,
                    "sn": sn,
                    "status": status,
                    "station": station,
                    "part_number": part_number,
                    "is_bonepile": is_bp,
                    "pb_id": pb_id,
                }
            )
            if len(batch) >= 2000:
                flush()

    flush()

    # Update coverage based on *actual data present* (prevents "scan coverage" racing ahead of data).
    data_min_ca_ms, data_max_ca_ms = get_db_data_range_ca_ms()
    if data_min_ca_ms is not None:
        state.min_ca_ms = data_min_ca_ms
    if data_max_ca_ms is not None:
        state.max_ca_ms = data_max_ca_ms
    state.min_key = seen_min_key
    state.max_key = seen_max_key
    state.min_path = seen_min_path
    state.max_path = seen_max_path
    state.last_scan_ca_ms = utc_ms(datetime.now(CA_TZ))
    state.save()

    return {
        "ok": True,
        "scanned_tw_days": len(tw_dates),
        "inserted": new_rows,
        "counters": {
            "visited_zip": visited_zip,
            "parsed_ok": parsed_ok,
            "ts_ok": ts_ok,
            "in_range": in_range,
        },
        "coverage": {"min_ca_ms": state.min_ca_ms, "max_ca_ms": state.max_ca_ms},
    }


def ensure_coverage(start_ca: datetime, end_ca: datetime) -> Dict[str, Any]:
    """
    Expand cache coverage to include the requested CA range, scanning only the missing segments.
    """
    # Clamp to "now" so we don't mark future time as covered (second precision).
    now_ca = datetime.now(CA_TZ).replace(microsecond=0)
    if end_ca.tzinfo is None:
        end_ca = CA_TZ.localize(end_ca)
    if start_ca.tzinfo is None:
        start_ca = CA_TZ.localize(start_ca)
    if end_ca > now_ca:
        end_ca = now_ca
    if start_ca > now_ca:
        start_ca = now_ca - timedelta(minutes=1)

    with scan_lock:
        state = RawState.load()

        # Always treat "covered" as actual DB coverage, not attempted coverage.
        data_min_ca_ms, data_max_ca_ms = get_db_data_range_ca_ms()
        # Keep state aligned with DB coverage (self-heal older bad states)
        changed = False
        if data_min_ca_ms is not None and state.min_ca_ms != data_min_ca_ms:
            state.min_ca_ms = data_min_ca_ms
            changed = True
        if data_max_ca_ms is not None and state.max_ca_ms != data_max_ca_ms:
            state.max_ca_ms = data_max_ca_ms
            changed = True
        if changed:
            state.save()

        result: Dict[str, Any] = {"ok": True, "actions": []}
        if state.min_ca_ms is None or state.max_ca_ms is None:
            r = scan_range(start_ca, end_ca, state)
            result["actions"].append({"type": "scan", "range": "initial", "result": r})
            return result

        start_ms = utc_ms(start_ca if start_ca.tzinfo else CA_TZ.localize(start_ca))
        end_ms = utc_ms(end_ca if end_ca.tzinfo else CA_TZ.localize(end_ca))

        if start_ms < state.min_ca_ms:
            r = scan_range(start_ca, datetime.fromtimestamp(state.min_ca_ms / 1000, CA_TZ), state)
            result["actions"].append({"type": "scan", "range": "backfill", "result": r})
        if end_ms > state.max_ca_ms:
            r = scan_range(datetime.fromtimestamp(state.max_ca_ms / 1000, CA_TZ), end_ca, state)
            result["actions"].append({"type": "scan", "range": "forward", "result": r})

        return result


# -----------------------------
# Aggregations for UI
# -----------------------------


def query_entries_in_range(start_ca: datetime, end_ca: datetime) -> List[sqlite3.Row]:
    start_ms = utc_ms(start_ca if start_ca.tzinfo else CA_TZ.localize(start_ca))
    end_ms = utc_ms(end_ca if end_ca.tzinfo else CA_TZ.localize(end_ca))
    try:
        conn = connect_db()
        try:
            cur = conn.execute(
                """
                SELECT *
                FROM raw_entries
                WHERE ca_ms BETWEEN ? AND ?
                ORDER BY sn, utc_ms, filename;
                """,
                (start_ms, end_ms),
            )
            return list(cur.fetchall())
        finally:
            conn.close()
    except sqlite3.OperationalError:
        ensure_db_ready(force=True)
        conn = connect_db()
        try:
            cur = conn.execute(
                """
                SELECT *
                FROM raw_entries
                WHERE ca_ms BETWEEN ? AND ?
                ORDER BY sn, utc_ms, filename;
                """,
                (start_ms, end_ms),
            )
            return list(cur.fetchall())
        finally:
            conn.close()


def cleanup_retention(now_ca: Optional[datetime] = None) -> Dict[str, Any]:
    """
    Delete cached raw entries older than RETENTION_DAYS (based on CA timestamp).
    Keeps cache size bounded and prevents UI from showing ancient data forever.
    """
    global last_retention_cleanup_ms
    try:
        if now_ca is None:
            now_ca = datetime.now(CA_TZ)
        if now_ca.tzinfo is None:
            now_ca = CA_TZ.localize(now_ca)

        cutoff_ca = now_ca - timedelta(days=RETENTION_DAYS)
        cutoff_ms = utc_ms(cutoff_ca)

        conn = connect_db()
        try:
            cur = conn.execute("DELETE FROM raw_entries WHERE ca_ms < ?;", (cutoff_ms,))
            deleted = cur.rowcount if cur.rowcount is not None else 0
            conn.commit()
        finally:
            conn.close()

        # After deleting, clamp scan coverage forward so UI stays consistent.
        data_min_ca_ms, data_max_ca_ms = get_db_data_range_ca_ms()
        state = RawState.load()
        changed = False
        if data_min_ca_ms is not None and (state.min_ca_ms is None or state.min_ca_ms < data_min_ca_ms):
            state.min_ca_ms = data_min_ca_ms
            changed = True
        if data_max_ca_ms is not None and (state.max_ca_ms is None or state.max_ca_ms < data_max_ca_ms):
            state.max_ca_ms = data_max_ca_ms
            changed = True
        if changed:
            state.save()

        with auto_status_lock:
            last_retention_cleanup_ms = utc_ms(now_ca)

        return {"ok": True, "deleted": int(deleted), "cutoff_ca_ms": int(cutoff_ms)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def compute_stats(rows: List[sqlite3.Row], aggregation: str) -> Dict[str, Any]:
    """
    Compute:
      - summary matrix (tested/pass/fail) x (bp/fresh/total)
      - sku rows: part_number -> pass/fail unique SN (SN assigned to latest part_number in range)
      - breakdown rows (daily/weekly/monthly)
    """
    # Group by SN
    sn_tests: Dict[str, List[sqlite3.Row]] = {}
    for r in rows:
        sn = r["sn"]
        sn_tests.setdefault(sn, []).append(r)

    sn_is_bp: Dict[str, int] = {}
    sn_pass: Dict[str, int] = {}
    sn_latest_part: Dict[str, str] = {}
    sn_latest_utc: Dict[str, int] = {}

    for sn, tests in sn_tests.items():
        is_bp = 0
        is_pass = 0
        latest_part = "Unknown"
        latest_utc = -1
        for t in tests:
            if (t["is_bonepile"] or 0) == 1:
                is_bp = 1
            if is_final_pass(t["status"], t["station"], t["part_number"]):
                is_pass = 1
            if int(t["utc_ms"]) > latest_utc:
                latest_utc = int(t["utc_ms"])
                latest_part = t["part_number"] or "Unknown"
        sn_is_bp[sn] = is_bp
        sn_pass[sn] = is_pass
        sn_latest_part[sn] = latest_part
        sn_latest_utc[sn] = latest_utc

    tested_total = len(sn_tests)
    pass_total = sum(1 for v in sn_pass.values() if v)
    fail_total = tested_total - pass_total

    tested_bp = sum(1 for sn, v in sn_is_bp.items() if v == 1)
    tested_fresh = tested_total - tested_bp
    pass_bp = sum(1 for sn in sn_tests.keys() if sn_is_bp.get(sn, 0) == 1 and sn_pass.get(sn, 0) == 1)
    pass_fresh = pass_total - pass_bp
    fail_bp = tested_bp - pass_bp
    fail_fresh = tested_fresh - pass_fresh

    summary = {
        "bp": {"tested": tested_bp, "pass": pass_bp, "fail": fail_bp},
        "fresh": {"tested": tested_fresh, "pass": pass_fresh, "fail": fail_fresh},
        "total": {"tested": tested_total, "pass": pass_total, "fail": fail_total},
    }

    # SKU rows (assign each SN to latest part_number)
    sku_stats: Dict[str, Dict[str, int]] = {}
    for sn in sn_tests.keys():
        sku = sn_latest_part.get(sn, "Unknown") or "Unknown"
        sku_stats.setdefault(sku, {"pass": 0, "fail": 0, "tested": 0})
        sku_stats[sku]["tested"] += 1
        if sn_pass.get(sn, 0) == 1:
            sku_stats[sku]["pass"] += 1
        else:
            sku_stats[sku]["fail"] += 1

    sku_rows = [
        {"sku": sku, "tested": s["tested"], "pass": s["pass"], "fail": s["fail"]}
        for sku, s in sku_stats.items()
    ]
    sku_rows.sort(key=lambda x: (-x["tested"], x["sku"]))

    # Breakdown per period (daily/weekly/monthly)
    # Count unique SN per period and compute pass/bp/fresh based on tests inside that period.
    # (A SN may appear in multiple periods; we count it in each.)
    bucket_key_field = {"daily": "ca_date", "weekly": "ca_week", "monthly": "ca_month"}.get(aggregation, "ca_date")
    bucket_sn_tests: Dict[str, Dict[str, List[sqlite3.Row]]] = {}
    for r in rows:
        bucket = r[bucket_key_field]
        sn = r["sn"]
        bucket_sn_tests.setdefault(bucket, {}).setdefault(sn, []).append(r)

    breakdown_rows: List[Dict[str, Any]] = []
    for bucket, sn_map in bucket_sn_tests.items():
        tested = len(sn_map)
        passed = 0
        bp = 0
        for sn, tests in sn_map.items():
            is_bp_bucket = any((t["is_bonepile"] or 0) == 1 for t in tests)
            if is_bp_bucket:
                bp += 1
            if any(is_final_pass(t["status"], t["station"], t["part_number"]) for t in tests):
                passed += 1
        fresh = tested - bp
        pass_rate = (passed / tested) if tested else 0.0
        breakdown_rows.append(
            {
                "period": bucket,
                "tested": tested,
                "passed": passed,
                "bonepile": bp,
                "fresh": fresh,
                "pass_rate": pass_rate,
            }
        )
    breakdown_rows.sort(key=lambda x: x["period"])

    return {
        "summary": summary,
        "sku_rows": sku_rows,
        "breakdown_rows": breakdown_rows,
    }


def _ts_group_from_part_number(part_number: str) -> str:
    pn = "" if part_number is None else str(part_number).upper()
    m = re.search(r"\bTS(\d+)\b", pn)
    if m:
        return f"TS{int(m.group(1))}"
    return "TS?"


def compute_test_flow(rows: List[sqlite3.Row]) -> Dict[str, Any]:
    """
    Compute station flow table:
    - Stations order fixed: FLA -> FLB -> AST -> FTS -> FCT -> RIN -> NVL
    - Row 1: totals per station (unique SNs) for PASS/FAIL at that station
    - Next rows: grouped by TSx (derived from latest part_number of SN in range),
      then SKU rows sorted ascending. Counts are unique SNs.

    PASS/FAIL per station is based on raw entry status at that station:
    - PASS: SN has at least one row with status='P' at that station (within range)
    - FAIL: SN has at least one row with status='F' at that station (within range)
    A SN can be counted in both for a station if retested (same as raw behavior).
    """
    stations = ["FLA", "FLB", "AST", "FTS", "FCT", "RIN", "NVL"]

    # Group rows by SN
    sn_tests: Dict[str, List[sqlite3.Row]] = {}
    for r in rows:
        sn_tests.setdefault(r["sn"], []).append(r)

    # Determine latest part_number per SN (same rule as SKU table)
    sn_latest_part: Dict[str, str] = {}
    sn_latest_key: Dict[str, Tuple[int, str]] = {}
    for sn, tests in sn_tests.items():
        best_key = (-1, "")
        best_pn = "Unknown"
        for t in tests:
            try:
                ca_ms = int(t["ca_ms"])
            except Exception:
                continue
            fn = t["filename"] or ""
            key = (ca_ms, fn)
            if key > best_key:
                best_key = key
                best_pn = t["part_number"] or "Unknown"
        sn_latest_part[sn] = best_pn
        sn_latest_key[sn] = best_key

    # Totals sets
    total_sets: Dict[str, Dict[str, set]] = {st: {"pass": set(), "fail": set()} for st in stations}

    # Per-SKU sets
    sku_sets: Dict[str, Dict[str, Dict[str, set]]] = {}
    for sn, tests in sn_tests.items():
        sku = sn_latest_part.get(sn, "Unknown") or "Unknown"
        sku_sets.setdefault(sku, {st: {"pass": set(), "fail": set()} for st in stations})
        for t in tests:
            st = str(t["station"] or "").strip().upper()
            if st not in total_sets:
                continue
            status = str(t["status"] or "").strip().upper()
            if status == "P":
                total_sets[st]["pass"].add(sn)
                sku_sets[sku][st]["pass"].add(sn)
            elif status == "F":
                total_sets[st]["fail"].add(sn)
                sku_sets[sku][st]["fail"].add(sn)

    totals = {
        st: {"pass": len(total_sets[st]["pass"]), "fail": len(total_sets[st]["fail"])}
        for st in stations
    }

    # Flat rows with TS column, sort by TS then SKU ascending
    def ts_sort_key(ts: str) -> Tuple[int, int]:
        m = re.match(r"TS(\d+)$", ts)
        if m:
            return (0, int(m.group(1)))
        return (1, 999)

    rows_out: List[Dict[str, Any]] = []
    for sku in sorted(sku_sets.keys()):
        ts = _ts_group_from_part_number(sku)
        rows_out.append(
            {
                "ts": ts,
                "sku": sku,
                "stations": {
                    st: {"pass": len(sku_sets[sku][st]["pass"]), "fail": len(sku_sets[sku][st]["fail"])}
                    for st in stations
                },
            }
        )

    rows_out.sort(key=lambda r: (ts_sort_key(r["ts"]), r["sku"]))

    return {"stations": stations, "totals": totals, "rows": rows_out}


def compute_station_sn_list(
    rows: List[sqlite3.Row],
    station: str,
    outcome: str,
    sku: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    Build SN list filtered by a station outcome:
    - station: e.g. FCT
    - outcome: 'pass' or 'fail' (maps to status P/F)
    - sku: optional exact match to SN latest part_number (in this slice)
    """
    st = str(station or "").strip().upper()
    want_status = "P" if outcome == "pass" else "F"

    # Group rows by SN and compute latest part_number for SKU assignment (in this slice)
    sn_rows: Dict[str, List[sqlite3.Row]] = {}
    for r in rows:
        sn_rows.setdefault(r["sn"], []).append(r)

    sn_latest_part: Dict[str, str] = {}
    for sn, tests in sn_rows.items():
        best_key = (-1, "")
        best_pn = "Unknown"
        for t in tests:
            ca_ms = int(t["ca_ms"])
            fn = t["filename"] or ""
            key = (ca_ms, fn)
            if key > best_key:
                best_key = key
                best_pn = t["part_number"] or "Unknown"
        sn_latest_part[sn] = best_pn

    out: List[Dict[str, Any]] = []
    for sn, tests in sn_rows.items():
        if sku and (sn_latest_part.get(sn, "Unknown") != sku):
            continue

        matched = [t for t in tests if str(t["station"] or "").strip().upper() == st and str(t["status"] or "").strip().upper() == want_status]
        if not matched:
            continue

        # Use latest matched row as "last" context + pass time
        best_key = (-1, "")
        best_row = None
        for t in matched:
            ca_ms = int(t["ca_ms"])
            fn = t["filename"] or ""
            key = (ca_ms, fn)
            if key > best_key:
                best_key = key
                best_row = t

        last_ca_ms = int(best_key[0]) if best_key[0] >= 0 else None
        last_filename = best_row["filename"] if best_row else None
        last_station = best_row["station"] if best_row else None
        last_part_number = best_row["part_number"] if best_row else None
        last_folder_path = best_row["folder_path"] if best_row else None
        last_folder_id = os.path.basename(last_folder_path) if last_folder_path else None

        out.append(
            {
                "sn": sn,
                "result": "PASS" if outcome == "pass" else "FAIL",
                "is_pass": 1 if outcome == "pass" else 0,
                "is_bonepile": 1 if any((t["is_bonepile"] or 0) == 1 for t in tests) else 0,
                "pass_ca_ms": last_ca_ms,
                "last_filename": last_filename,
                "last_station": last_station,
                "last_part_number": last_part_number,
                "last_folder_id": last_folder_id,
                "last_folder_path": last_folder_path,
            }
        )

    out.sort(key=lambda x: (x["pass_ca_ms"] or 0, x["sn"]), reverse=True)
    return out


def compute_station_sn_list_both(
    rows: List[sqlite3.Row],
    station: str,
    sku: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    Return SN list for a station, including PASS and FAIL in one list (unique SN).
    If a SN has both P and F at the station in this slice, result will be "PASS/FAIL".
    """
    st = str(station or "").strip().upper()

    # Group rows by SN and compute latest part_number for SKU assignment (in this slice)
    sn_rows: Dict[str, List[sqlite3.Row]] = {}
    for r in rows:
        sn_rows.setdefault(r["sn"], []).append(r)

    sn_latest_part: Dict[str, str] = {}
    for sn, tests in sn_rows.items():
        best_key = (-1, "")
        best_pn = "Unknown"
        for t in tests:
            ca_ms = int(t["ca_ms"])
            fn = t["filename"] or ""
            key = (ca_ms, fn)
            if key > best_key:
                best_key = key
                best_pn = t["part_number"] or "Unknown"
        sn_latest_part[sn] = best_pn

    out: List[Dict[str, Any]] = []
    for sn, tests in sn_rows.items():
        if sku and (sn_latest_part.get(sn, "Unknown") != sku):
            continue

        station_tests = [
            t
            for t in tests
            if str(t["station"] or "").strip().upper() == st
            and str(t["status"] or "").strip().upper() in ("P", "F")
        ]
        if not station_tests:
            continue

        # Latest per status
        best_p = (-1, "")
        best_f = (-1, "")
        best_p_row = None
        best_f_row = None
        for t in station_tests:
            ca_ms = int(t["ca_ms"])
            fn = t["filename"] or ""
            key = (ca_ms, fn)
            status = str(t["status"] or "").strip().upper()
            if status == "P" and key > best_p:
                best_p = key
                best_p_row = t
            if status == "F" and key > best_f:
                best_f = key
                best_f_row = t

        has_p = best_p_row is not None
        has_f = best_f_row is not None
        if has_p and has_f:
            result = "PASS/FAIL"
        elif has_p:
            result = "PASS"
        else:
            result = "FAIL"

        # Use the latest of (P,F) as the context row
        context_row = best_p_row if best_p >= best_f else best_f_row
        context_key = best_p if best_p >= best_f else best_f
        context_ms = int(context_key[0]) if context_key[0] >= 0 else None

        last_filename = context_row["filename"] if context_row else None
        last_station = context_row["station"] if context_row else None
        last_part_number = context_row["part_number"] if context_row else None
        last_folder_path = context_row["folder_path"] if context_row else None
        last_folder_id = os.path.basename(last_folder_path) if last_folder_path else None

        out.append(
            {
                "sn": sn,
                "result": result,
                "is_pass": 1 if result.startswith("PASS") else 0,
                "is_bonepile": 1 if any((t["is_bonepile"] or 0) == 1 for t in tests) else 0,
                # Use context time in the modal "time" column (kept as pass_ca_ms for compatibility)
                "pass_ca_ms": context_ms,
                "last_filename": last_filename,
                "last_station": last_station,
                "last_part_number": last_part_number,
                "last_folder_id": last_folder_id,
                "last_folder_path": last_folder_path,
            }
        )

    out.sort(key=lambda x: (x["pass_ca_ms"] or 0, x["sn"]), reverse=True)
    return out


def compute_sn_details(rows: List[sqlite3.Row]) -> List[Dict[str, Any]]:
    """
    Build per-SN details for modal drill-down.

    For each SN in the range:
    - is_bonepile: any row has is_bonepile=1
    - is_pass: any row is a final pass (based on part_number rule)
    - pass_ca_ms: latest final pass time (CA ms)
    - fail_ca_ms: latest fail time (any station/status F) (CA ms)
    - last_*: latest seen row (CA ms, filename, station, part_number, folder_path + folder_id)
    """
    sn_map: Dict[str, List[sqlite3.Row]] = {}
    for r in rows:
        sn_map.setdefault(r["sn"], []).append(r)

    out: List[Dict[str, Any]] = []
    for sn, tests in sn_map.items():
        is_bp = False
        pass_ms: Optional[int] = None
        fail_ms: Optional[int] = None

        # Latest seen row (by ca_ms then filename for stability)
        last_row = None
        last_key = (-1, "")

        for t in tests:
            try:
                ca_ms = int(t["ca_ms"])
            except Exception:
                continue

            if (t["is_bonepile"] or 0) == 1:
                is_bp = True

            if t["status"] == "F":
                if fail_ms is None or ca_ms > fail_ms:
                    fail_ms = ca_ms

            if is_final_pass(t["status"], t["station"], t["part_number"]):
                if pass_ms is None or ca_ms > pass_ms:
                    pass_ms = ca_ms

            fn = t["filename"] or ""
            key = (ca_ms, fn)
            if key > last_key:
                last_key = key
                last_row = t

        last_ca_ms = int(last_key[0]) if last_key[0] >= 0 else None
        last_filename = last_row["filename"] if last_row else None
        last_station = last_row["station"] if last_row else None
        last_part_number = last_row["part_number"] if last_row else None
        last_folder_path = last_row["folder_path"] if last_row else None
        last_folder_id = os.path.basename(last_folder_path) if last_folder_path else None

        out.append(
            {
                "sn": sn,
                "result": "PASS" if pass_ms is not None else "FAIL",
                "is_pass": 1 if pass_ms is not None else 0,
                "is_bonepile": 1 if is_bp else 0,
                "pass_ca_ms": pass_ms,
                "fail_ca_ms": fail_ms,
                "last_ca_ms": last_ca_ms,
                "last_filename": last_filename,
                "last_station": last_station,
                "last_part_number": last_part_number,
                "last_folder_id": last_folder_id,
                "last_folder_path": last_folder_path,
            }
        )

    out.sort(key=lambda x: (x["last_ca_ms"] or 0, x["sn"]), reverse=True)
    return out


# -----------------------------
# Job system (in-memory)
# -----------------------------


jobs_lock = threading.Lock()
jobs: Dict[str, Dict[str, Any]] = {}


def new_job_id() -> str:
    return f"job_{int(time.time() * 1000)}_{os.getpid()}"


def set_job(job_id: str, **fields: Any) -> None:
    with jobs_lock:
        jobs.setdefault(job_id, {})
        jobs[job_id].update(fields)


def run_scan_job(job_id: str, start_ca: datetime, end_ca: datetime) -> None:
    try:
        set_job(job_id, status="running", message="Scanning...", started_at=int(time.time()))
        res = ensure_coverage(start_ca, end_ca)
        set_job(job_id, status="done", result=res, finished_at=int(time.time()))
    except Exception as e:
        set_job(job_id, status="error", error=str(e), finished_at=int(time.time()))


def _bonepile_status_payload(state: RawState) -> Dict[str, Any]:
    bf = state.bonepile_file or {}
    return {
        "file": bf,
        "allowed_sheets": BONEPILE_ALLOWED_SHEETS,
        "mapping": state.bonepile_mapping or {},
        "sheets": state.bonepile_sheet_status or {},
    }


def _save_uploaded_bonepile_file(file_storage) -> Dict[str, Any]:
    ensure_dirs()
    # Replace existing file atomically
    tmp_path = BONEPILE_UPLOAD_PATH + ".tmp"
    file_storage.save(tmp_path)
    if os.path.exists(BONEPILE_UPLOAD_PATH):
        try:
            os.remove(BONEPILE_UPLOAD_PATH)
        except Exception:
            pass
    os.replace(tmp_path, BONEPILE_UPLOAD_PATH)
    stat = os.stat(BONEPILE_UPLOAD_PATH)
    now = datetime.now(CA_TZ).replace(microsecond=0)
    return {
        "has_file": True,
        "path": BONEPILE_UPLOAD_PATH,
        "original_name": getattr(file_storage, "filename", None),
        "size_bytes": int(getattr(stat, "st_size", 0)),
        "uploaded_at_ca_ms": utc_ms(now),
    }


def run_bonepile_parse_job(job_id: str, sheets: Optional[List[str]] = None) -> None:
    """
    Parse the uploaded NV/IGS workbook for allowed sheets.
    - Per-sheet: auto-detect header by 'SN' and map columns by header names unless user saved mapping.
    - Writes rows into SQLite bonepile_entries (replaces per-sheet).
    - Updates RawState.bonepile_sheet_status with ok/error for each sheet.
    """
    try:
        set_job(job_id, status="running", message="Parsing workbook...", started_at=int(time.time()))
        ensure_db_ready()
        with scan_lock:
            state = RawState.load()
        if not os.path.exists(BONEPILE_UPLOAD_PATH):
            raise RuntimeError("No uploaded bonepile workbook found")
        wb = _load_bonepile_workbook(BONEPILE_UPLOAD_PATH)
        try:
            all_sheets = list(wb.sheetnames)
            allowed = [s for s in BONEPILE_ALLOWED_SHEETS if s in all_sheets]
            target = allowed if not sheets else [s for s in sheets if s in allowed]

            mapping_cfg = (state.bonepile_mapping or {})
            sheet_status: Dict[str, Any] = state.bonepile_sheet_status or {}

            conn = connect_db()
            try:
                for sheet in target:
                    ws = wb[sheet]
                    
                    # Compute hash of current sheet content
                    current_hash = _hash_sheet_content(ws)
                    prev_status = sheet_status.get(sheet) if isinstance(sheet_status.get(sheet), dict) else {}
                    prev_hash = prev_status.get("content_hash") if isinstance(prev_status.get("content_hash"), str) else None
                    
                    # Skip parsing if hash matches (content unchanged)
                    if prev_hash and prev_hash == current_hash:
                        # Keep previous status but update last_run timestamp
                        prev_status["last_run_ca_ms"] = utc_ms(datetime.now(CA_TZ))
                        prev_status["skipped"] = True
                        prev_status["skip_reason"] = "Content unchanged (hash match)"
                        sheet_status[sheet] = prev_status
                        continue
                    
                    # Determine header row + mapping
                    cfg = (mapping_cfg.get(sheet) or {}) if isinstance(mapping_cfg.get(sheet), dict) else {}
                    header_row = int(cfg.get("header_row") or 0) if cfg.get("header_row") else 0
                    if header_row <= 0:
                        header_row = _find_header_row(ws) or 0
                    if header_row <= 0:
                        sheet_status[sheet] = {
                            "status": "error",
                            "error": "Header row not found (SN)",
                            "last_run_ca_ms": utc_ms(datetime.now(CA_TZ)),
                            "content_hash": current_hash,
                        }
                        continue

                    header_map = _read_header_map(ws, header_row=header_row)

                    # User mapping by header name (preferred) or auto mapping
                    col_map: Dict[str, int] = {}
                    user_cols = cfg.get("columns") if isinstance(cfg.get("columns"), dict) else None
                    if user_cols:
                        for k, v in user_cols.items():
                            if not v:
                                continue
                            # v can be a header string
                            if isinstance(v, str):
                                col_map[k] = int(header_map.get(v.strip().upper(), 0))
                            else:
                                try:
                                    col_map[k] = int(v)
                                except Exception:
                                    col_map[k] = 0
                        # fill missing with auto mapping
                        auto = _auto_mapping_from_headers(header_map)
                        for k, idx in auto.items():
                            col_map.setdefault(k, idx)
                    else:
                        col_map = _auto_mapping_from_headers(header_map)

                    errs = _mapping_errors(col_map, header_map)
                    if errs:
                        sheet_status[sheet] = {
                            "status": "error",
                            "error": "; ".join(errs[:3]),
                            "header_row": header_row,
                            "last_run_ca_ms": utc_ms(datetime.now(CA_TZ)),
                            "content_hash": current_hash,
                        }
                        continue

                    # Replace rows for this sheet
                    conn.execute("DELETE FROM bonepile_entries WHERE sheet = ?;", (sheet,))
                    now_ms = utc_ms(datetime.now(CA_TZ).replace(microsecond=0))
                    inserted = 0
                    empty_sn_streak = 0

                    for excel_row_idx, row in enumerate(
                        ws.iter_rows(min_row=header_row + 1, values_only=True),
                        start=header_row + 1,
                    ):
                        # Stop if we hit a long blank region
                        if row is None:
                            continue
                        sn_val = row[col_map["sn"] - 1] if col_map["sn"] > 0 and col_map["sn"] <= len(row) else None
                        sn = _normalize_sn(sn_val)
                        if not sn:
                            empty_sn_streak += 1
                            if empty_sn_streak >= 200:
                                break
                            continue
                        empty_sn_streak = 0

                        def cell(idx: int) -> str:
                            if idx <= 0 or idx > len(row):
                                return ""
                            v = row[idx - 1]
                            return "" if v is None else str(v).strip()

                        nv_dispo = cell(col_map.get("nv_disposition", 0))
                        igs_action = cell(col_map.get("igs_action", 0))
                        status = cell(col_map.get("status", 0))
                        pic = cell(col_map.get("pic", 0))
                        igs_status = cell(col_map.get("igs_status", 0))
                        nvpn = cell(col_map.get("nvpn", 0))

                        nv_cnt = len(_extract_mmdd_entries(nv_dispo))
                        igs_cnt = len(_extract_mmdd_entries(igs_action))

                        conn.execute(
                            """
                            INSERT OR REPLACE INTO bonepile_entries (
                              sheet, excel_row, sn, nvpn, status, pic, igs_status,
                              nv_disposition, igs_action, nv_dispo_count, igs_action_count, updated_at_ca_ms
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
                            """,
                            (
                                sheet,
                                int(excel_row_idx),
                                sn,
                                nvpn,
                                status,
                                pic,
                                igs_status,
                                nv_dispo,
                                igs_action,
                                int(nv_cnt),
                                int(igs_cnt),
                                int(now_ms),
                            ),
                        )
                        inserted += 1

                    conn.commit()
                    sheet_status[sheet] = {
                        "status": "ok",
                        "rows": int(inserted),
                        "header_row": int(header_row),
                        "last_run_ca_ms": int(now_ms),
                        "content_hash": current_hash,
                    }

                # Save state
                with scan_lock:
                    st = RawState.load()
                    st.bonepile_sheet_status = sheet_status
                    st.save()
            finally:
                conn.close()
        finally:
            try:
                wb.close()
            except Exception:
                pass

        set_job(job_id, status="done", message="Workbook parsed", finished_at=int(time.time()))
    except Exception as e:
        set_job(job_id, status="error", error=str(e), finished_at=int(time.time()))
        with scan_lock:
            st = RawState.load()
            ss = st.bonepile_sheet_status or {}
            ss["_job_error"] = str(e)
            st.bonepile_sheet_status = ss
            st.save()


# -----------------------------
# Routes
# -----------------------------


@app.route("/")
def dashboard():
    return render_template("analytics_dashboard.html")


@app.route("/api/status")
def api_status():
    # Self-heal if cache files were manually deleted.
    ensure_db_ready()
    state = RawState.load()
    data_min_ca_ms, data_max_ca_ms = get_db_data_range_ca_ms()
    with auto_status_lock:
        next_ms = next_auto_scan_ms
        last_cleanup_ms = last_retention_cleanup_ms
    return jsonify(
        {
            "cache": {
                # Data coverage (actual rows present)
                "min_ca_ms": data_min_ca_ms,
                "max_ca_ms": data_max_ca_ms,
                # Scan coverage (what we have attempted to cover)
                "scan_min_ca_ms": state.min_ca_ms,
                "scan_max_ca_ms": state.max_ca_ms,
                "min_key": list(state.min_key) if state.min_key else None,
                "max_key": list(state.max_key) if state.max_key else None,
                "min_path": state.min_path,
                "max_path": state.max_path,
                "last_scan_ca_ms": state.last_scan_ca_ms,
                "scan_interval_seconds": AUTO_SCAN_EVERY_SECONDS,
                "retention_days": RETENTION_DAYS,
                "next_auto_scan_ms": next_ms,
                "last_retention_cleanup_ms": last_cleanup_ms,
            }
            ,
            "bonepile": _bonepile_status_payload(state),
        }
    )


@app.route("/api/events")
def api_events():
    """
    Server-Sent Events stream for lightweight status updates.

    This avoids polling /api/status repeatedly while still updating:
    - data coverage
    - scan coverage
    - last scan time
    - next auto-scan time
    """

    def get_status_payload() -> Dict[str, Any]:
        ensure_db_ready()
        state = RawState.load()
        data_min_ca_ms, data_max_ca_ms = get_db_data_range_ca_ms()
        with auto_status_lock:
            next_ms = next_auto_scan_ms
            last_cleanup_ms = last_retention_cleanup_ms
        return {
            "cache": {
                "min_ca_ms": data_min_ca_ms,
                "max_ca_ms": data_max_ca_ms,
                "scan_min_ca_ms": state.min_ca_ms,
                "scan_max_ca_ms": state.max_ca_ms,
                "last_scan_ca_ms": state.last_scan_ca_ms,
                "scan_interval_seconds": AUTO_SCAN_EVERY_SECONDS,
                "retention_days": RETENTION_DAYS,
                "next_auto_scan_ms": next_ms,
                "last_retention_cleanup_ms": last_cleanup_ms,
            }
            ,
            "bonepile": _bonepile_status_payload(state),
        }

    def gen():
        last_sent = None
        while True:
            try:
                payload = get_status_payload()
                data = json.dumps(payload, separators=(",", ":"))
                if data != last_sent:
                    last_sent = data
                    yield f"event: status\ndata: {data}\n\n"
            except Exception as e:
                # still keep stream alive
                yield f"event: error\ndata: {json.dumps({'error': str(e)})}\n\n"
            time.sleep(2.0)

    return Response(gen(), mimetype="text/event-stream", headers={"Cache-Control": "no-cache"})


@app.route("/api/clear-cache", methods=["POST"])
def api_clear_cache():
    """
    Clear SQLite cache + raw scan state so the system can rescan from scratch.

    Safe reset (cache only):
    - deletes analytics_cache/analytics.db (raw_entries + meta)
    - deletes analytics_cache/raw_state.json
    - clears in-memory jobs
    """
    global db_initialized
    with scan_lock:
        # Clear in-memory jobs
        with jobs_lock:
            jobs.clear()

        # Remove files (best-effort)
        try:
            if os.path.exists(DB_PATH):
                os.remove(DB_PATH)
        except Exception:
            pass
        try:
            if os.path.exists(STATE_PATH):
                os.remove(STATE_PATH)
        except Exception:
            pass
        try:
            if os.path.exists(BONEPILE_UPLOAD_PATH):
                os.remove(BONEPILE_UPLOAD_PATH)
        except Exception:
            pass

        # Force re-init on next access
        db_initialized = False
        ensure_db_ready(force=True)

    return jsonify({"ok": True})


@app.route("/api/job/<job_id>")
def api_job(job_id: str):
    with jobs_lock:
        data = jobs.get(job_id)
    if not data:
        return jsonify({"error": "job not found"}), 404
    return jsonify(data)


@app.route("/api/bonepile/status")
def api_bonepile_status():
    ensure_db_ready()
    state = RawState.load()
    return jsonify(_bonepile_status_payload(state))


@app.route("/api/bonepile/upload", methods=["POST"])
def api_bonepile_upload():
    """
    Upload NV/IGS workbook. The backend stores only the latest file (replaces previous).
    After upload, automatically parse all allowed sheets with auto-detect.
    Sheets with unchanged content (hash match) will be skipped.
    """
    ensure_db_ready()
    if openpyxl is None:
        return jsonify({"error": "openpyxl not installed; cannot accept XLSX"}), 500
    if "file" not in request.files:
        return jsonify({"error": "file is required"}), 400
    f = request.files["file"]
    if not f or not getattr(f, "filename", ""):
        return jsonify({"error": "no file selected"}), 400
    name = str(f.filename)
    if not name.lower().endswith(".xlsx"):
        return jsonify({"error": "only .xlsx is supported for bonepile upload"}), 400

    with scan_lock:
        state = RawState.load()
        meta = _save_uploaded_bonepile_file(f)
        state.bonepile_file = meta
        # Keep existing sheet status (for hash comparison)
        state.bonepile_sheet_status = state.bonepile_sheet_status or {}
        state.save()

    # Auto-parse all allowed sheets (with auto-detect, hash check will skip unchanged sheets)
    job_id = new_job_id()
    set_job(job_id, status="queued", message="Auto-parsing all sheets with auto-detect...")
    t = threading.Thread(target=run_bonepile_parse_job, args=(job_id, None), daemon=True)
    t.start()
    return jsonify({"ok": True, "job_id": job_id, "bonepile_file": meta})


@app.route("/api/bonepile/sheets")
def api_bonepile_sheets():
    """
    Return sheet list, ignore list, and auto-detected header/mapping suggestion for allowed sheets.
    """
    if openpyxl is None:
        return jsonify({"error": "openpyxl not installed; cannot read XLSX"}), 500
    state = RawState.load()
    if not os.path.exists(BONEPILE_UPLOAD_PATH):
        return jsonify({"ok": True, "has_file": False, "allowed": BONEPILE_ALLOWED_SHEETS, "ignored": [], "sheets": {}})
    wb = _load_bonepile_workbook(BONEPILE_UPLOAD_PATH)
    try:
        all_sheets = list(wb.sheetnames)
        ignored = [s for s in all_sheets if s not in BONEPILE_ALLOWED_SHEETS]
        out: Dict[str, Any] = {}
        for sheet in BONEPILE_ALLOWED_SHEETS:
            if sheet not in all_sheets:
                out[sheet] = {"present": False}
                continue
            ws = wb[sheet]
            header_row = _find_header_row(ws) or 0
            header_map = _read_header_map(ws, header_row) if header_row else {}
            auto_map = _auto_mapping_from_headers(header_map) if header_map else {}
            errs = _mapping_errors(auto_map, header_map) if header_map else ["Header row not found (SN)"]
            out[sheet] = {
                "present": True,
                "header_row": int(header_row) if header_row else None,
                "headers": list(header_map.keys())[:80],
                "auto_columns": auto_map,
                "auto_errors": errs,
                "saved_mapping": (state.bonepile_mapping or {}).get(sheet),
                "status": (state.bonepile_sheet_status or {}).get(sheet),
            }
        return jsonify({"ok": True, "has_file": True, "allowed": BONEPILE_ALLOWED_SHEETS, "ignored": ignored, "sheets": out})
    finally:
        try:
            wb.close()
        except Exception:
            pass


@app.route("/api/bonepile/mapping", methods=["POST"])
def api_bonepile_mapping():
    """
    Save mapping for a single sheet:
      { sheet, header_row, columns: {sn, nv_disposition, status, pic, igs_action, igs_status, nvpn?} }
    Values in columns can be header strings (preferred) or 1-based column indices.
    """
    payload = request.json or {}
    sheet = str(payload.get("sheet") or "").strip()
    if sheet not in BONEPILE_ALLOWED_SHEETS:
        return jsonify({"error": "invalid sheet"}), 400
    header_row = int(payload.get("header_row") or 0)
    columns = payload.get("columns") if isinstance(payload.get("columns"), dict) else {}
    if header_row <= 0:
        return jsonify({"error": "header_row must be >= 1"}), 400

    with scan_lock:
        state = RawState.load()
        if state.bonepile_mapping is None:
            state.bonepile_mapping = {}
        state.bonepile_mapping[sheet] = {"header_row": int(header_row), "columns": columns}
        state.save()

    # Trigger re-parse just this sheet in background
    job_id = new_job_id()
    set_job(job_id, status="queued", message=f"Parsing {sheet}...")
    t = threading.Thread(target=run_bonepile_parse_job, args=(job_id, [sheet]), daemon=True)
    t.start()
    return jsonify({"ok": True, "job_id": job_id})


@app.route("/api/bonepile/parse", methods=["POST"])
def api_bonepile_parse():
    """
    Trigger parse job (all sheets or a single sheet).
      { sheet?: "VR-TS1" }
    """
    ensure_db_ready()
    payload = request.json or {}
    sheet = str(payload.get("sheet") or "").strip() if payload.get("sheet") is not None else ""
    sheets: Optional[List[str]] = None
    if sheet:
        if sheet not in BONEPILE_ALLOWED_SHEETS:
            return jsonify({"error": "invalid sheet"}), 400
        sheets = [sheet]
    job_id = new_job_id()
    set_job(job_id, status="queued", message="Bonepile parse queued")
    t = threading.Thread(target=run_bonepile_parse_job, args=(job_id, sheets), daemon=True)
    t.start()
    return jsonify({"ok": True, "job_id": job_id})


@app.route("/api/bonepile/disposition")
def api_bonepile_disposition():
    """
    NV Disposition stats from bonepile_entries.
    Query: aggregation=daily|weekly|monthly (default daily), start_datetime, end_datetime (optional).
    Returns: summary { total, waiting_igs, complete }, by_sku, by_period.
    """
    ensure_db_ready()
    aggregation = request.args.get("aggregation", "daily").strip().lower()
    if aggregation not in ("daily", "weekly", "monthly"):
        aggregation = "daily"
    start_dt = request.args.get("start_datetime")
    end_dt = request.args.get("end_datetime")
    start_ca_ms = None
    end_ca_ms = None
    if start_dt:
        start_ca = _parse_ca_input_datetime(start_dt, is_end=False)
        if start_ca:
            start_ca_ms = utc_ms(start_ca)
    if end_dt:
        end_ca = _parse_ca_input_datetime(end_dt, is_end=True)
        if end_ca:
            end_ca_ms = utc_ms(end_ca)
    try:
        data = compute_disposition_stats(aggregation=aggregation, start_ca_ms=start_ca_ms, end_ca_ms=end_ca_ms)
        return jsonify({"ok": True, **data})
    except sqlite3.OperationalError:
        return jsonify({
            "ok": True,
            "summary": {"total": 0, "waiting_igs": 0, "complete": 0},
            "by_sku": [],
            "by_period": [],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/bonepile/disposition/sn-list", methods=["POST"])
def api_bonepile_disposition_sn_list():
    """
    SN list for disposition drill-down.
    Body: { metric: total|waiting|complete, sku?: string, period?: string, aggregation?: daily|weekly|monthly, start_datetime?, end_datetime? }.
    Returns: rows with sn, last_nv_dispo, last_igs_action, nvpn, status, pic.
    """
    ensure_db_ready()
    payload = request.json or {}
    metric = str(payload.get("metric") or "total").strip().lower()
    if metric not in ("total", "waiting", "complete", "trays_bp", "all_pass_trays"):
        metric = "total"
    sku = (payload.get("sku") or "").strip() or None
    period = (payload.get("period") or "").strip() or None
    aggregation = str(payload.get("aggregation") or "daily").strip().lower()
    if aggregation not in ("daily", "weekly", "monthly"):
        aggregation = "daily"
    start_dt = payload.get("start_datetime")
    end_dt = payload.get("end_datetime")
    start_ca_ms = None
    end_ca_ms = None
    if start_dt:
        start_ca = _parse_ca_input_datetime(start_dt, is_end=False)
        if start_ca:
            start_ca_ms = utc_ms(start_ca)
    if end_dt:
        end_ca = _parse_ca_input_datetime(end_dt, is_end=True)
        if end_ca:
            end_ca_ms = utc_ms(end_ca)
    try:
        rows = compute_disposition_sn_list(metric=metric, sku=sku, period=period, aggregation=aggregation, start_ca_ms=start_ca_ms, end_ca_ms=end_ca_ms)
        return jsonify({"ok": True, "count": len(rows), "rows": rows})
    except sqlite3.OperationalError:
        return jsonify({"ok": True, "count": 0, "rows": []})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/scan", methods=["POST"])
def api_scan():
    payload = request.json or {}
    start_dt = payload.get("start_datetime")
    end_dt = payload.get("end_datetime")
    if not start_dt or not end_dt:
        return jsonify({"error": "start_datetime and end_datetime required"}), 400
    start_ca = _parse_ca_input_datetime(start_dt, is_end=False)
    end_ca = _parse_ca_input_datetime(end_dt, is_end=True)
    if not start_ca or not end_ca:
        return jsonify({"error": "datetime format must be YYYY-MM-DD HH:MM (optional :SS)"}), 400
    # Clamp end to now so manual scan doesn't "reserve" future coverage.
    now_ca = datetime.now(CA_TZ).replace(microsecond=0)
    if end_ca > now_ca:
        end_ca = now_ca
    if start_ca > now_ca:
        return jsonify({"error": "start is in the future"}), 400

    job_id = new_job_id()
    set_job(job_id, status="queued", message="Queued")
    t = threading.Thread(target=run_scan_job, args=(job_id, start_ca, end_ca), daemon=True)
    t.start()
    return jsonify({"job_id": job_id, "status": "queued"})


@app.route("/api/query", methods=["POST"])
def api_query():
    payload = request.json or {}
    start_dt = payload.get("start_datetime")
    end_dt = payload.get("end_datetime")
    aggregation = (payload.get("aggregation") or "daily").strip().lower()
    if aggregation not in ("daily", "weekly", "monthly"):
        aggregation = "daily"

    if not start_dt or not end_dt:
        return jsonify({"error": "start_datetime and end_datetime required"}), 400
    start_ca = _parse_ca_input_datetime(start_dt, is_end=False)
    end_ca = _parse_ca_input_datetime(end_dt, is_end=True)
    if not start_ca or not end_ca:
        return jsonify({"error": "datetime format must be YYYY-MM-DD HH:MM (optional :SS)"}), 400
    # Clamp end to now so cache/coverage logic stays truthful.
    now_ca = datetime.now(CA_TZ).replace(microsecond=0)
    if end_ca > now_ca:
        end_ca = now_ca
    if start_ca > now_ca:
        return jsonify({"error": "start is in the future"}), 400
    if end_ca <= start_ca:
        return jsonify({"error": "end must be after start"}), 400

    # IMPORTANT DESIGN CHANGE:
    # Query must NEVER auto-trigger scans (that causes loops / "scan nonstop").
    # Scanning is done by:
    # - Manual scan button (/api/scan)
    # - Background auto-scan loop
    # Query returns whatever is currently in SQLite + a coverage flag.
    start_ms = utc_ms(start_ca)
    end_ms = utc_ms(end_ca)
    data_min_ca_ms, data_max_ca_ms = get_db_data_range_ca_ms()
    is_fully_covered = (
        data_min_ca_ms is not None
        and data_max_ca_ms is not None
        and start_ms >= int(data_min_ca_ms)
        and end_ms <= int(data_max_ca_ms)
    )

    rows = query_entries_in_range(start_ca, end_ca)
    computed = compute_stats(rows, aggregation=aggregation)
    test_flow = compute_test_flow(rows)
    return jsonify(
        {
            "needs_scan": False,
            "aggregation": aggregation,
            "summary": computed["summary"],
            "sku_rows": computed["sku_rows"][:200],
            "breakdown_rows": computed["breakdown_rows"],
            "counts": {"raw_rows": len(rows), "unique_sns": len(set([r["sn"] for r in rows]))},
            "coverage": {"min_ca_ms": data_min_ca_ms, "max_ca_ms": data_max_ca_ms},
            "is_fully_covered": bool(is_fully_covered),
            "test_flow": test_flow,
        }
    )


def _csv_response(text: str, filename: str) -> Response:
    resp = Response(text, mimetype="text/csv; charset=utf-8")
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp.headers["Cache-Control"] = "no-store"
    return resp


def _xlsx_response(data: bytes, filename: str) -> Response:
    resp = Response(data, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp.headers["Cache-Control"] = "no-store"
    return resp


def _copy_cell_style(src_cell, tgt_cell):
    """Copy font, fill, alignment, border, number_format from source cell to target cell."""
    if src_cell.has_style:
        tgt_cell.font = src_cell.font.copy()
        tgt_cell.fill = src_cell.fill.copy()
        tgt_cell.alignment = src_cell.alignment.copy()
        tgt_cell.border = src_cell.border.copy()
        tgt_cell.number_format = src_cell.number_format


def _build_export_xlsx(
    export_kind: str,
    computed: Dict[str, Any],
    start_s: str,
    end_s: str,
    start_ca: datetime,
    end_ca: datetime,
) -> Tuple[bytes, str]:
    """
    Build XLSX using templates; preserve template formatting.
    export_kind: 'summary' | 'sku'
    start_ca, end_ca: datetime objects for header text
    Returns (xlsx_bytes, filename).
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed; cannot export XLSX")

    # Format datetime for header: "YYYY-MM-DD HH:MM"
    start_str = start_ca.strftime("%Y-%m-%d %H:%M")
    end_str = end_ca.strftime("%Y-%m-%d %H:%M")
    header_text = f"Testing from {start_str} to {end_str}"

    if export_kind == "summary":
        path = TRAY_SUMMARY_TEMPLATE_PATH
        if not os.path.isfile(path):
            raise FileNotFoundError(f"Tray Summary template not found: {path}")
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # Template may have merged cells A1:D1, unmerge first if exists
        merged_ranges = list(ws.merged_cells.ranges)
        for mr in merged_ranges:
            mr_str = str(mr)
            # Unmerge if it overlaps with row 1 and includes column A
            if 'A1' in mr_str or mr_str.startswith('A1:'):
                ws.unmerge_cells(mr_str)
        # Get fill color from A1 before clearing it
        from openpyxl.styles import PatternFill
        a1_cell = ws.cell(row=1, column=1)
        a1_fill = a1_cell.fill.copy() if a1_cell.has_style and a1_cell.fill else None
        # Clear fill from A1
        if a1_cell.has_style:
            a1_cell.fill = PatternFill()
        # Fill from column B-D only, leave A1 empty
        header_cell = ws.cell(row=1, column=2, value=header_text)
        # Merge cells B1:D1 for header (A1 stays empty)
        ws.merge_cells('B1:D1')
        # Style header: bold, center alignment
        from openpyxl.styles import Font, Alignment
        header_cell.font = Font(bold=True, size=12)
        header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # Apply fill color from A1 to header (B1:D1)
        if a1_fill:
            header_cell.fill = a1_fill
        # Auto-adjust column width based on header text length
        # Estimate: each character ~1.2 units, divide by 3 for B+C+D columns, add padding
        text_length = len(header_text)
        estimated_width = max(text_length * 1.2 / 3, 15)
        for col in ['B', 'C', 'D']:
            current_width = ws.column_dimensions[col].width if col in ws.column_dimensions and ws.column_dimensions[col].width else 0
            ws.column_dimensions[col].width = max(current_width, estimated_width)
        
        s = computed["summary"]
        # Template: row 2 = headers (BP, FRESH, TOTAL), row 3 = TOTAL/tested, row 4 = PASS, row 5 = FAIL
        # Map: row 3 -> "tested", row 4 -> "pass", row 5 -> "fail"
        ws.cell(row=3, column=1, value="TOTAL")  # Keep label as "TOTAL" in template
        ws.cell(row=3, column=2, value=s["bp"].get("tested", 0))
        ws.cell(row=3, column=3, value=s["fresh"].get("tested", 0))
        ws.cell(row=3, column=4, value=s["total"].get("tested", 0))
        ws.cell(row=4, column=1, value="PASS")
        ws.cell(row=4, column=2, value=s["bp"].get("pass", 0))
        ws.cell(row=4, column=3, value=s["fresh"].get("pass", 0))
        ws.cell(row=4, column=4, value=s["total"].get("pass", 0))
        ws.cell(row=5, column=1, value="FAIL")
        ws.cell(row=5, column=2, value=s["bp"].get("fail", 0))
        ws.cell(row=5, column=3, value=s["fresh"].get("fail", 0))
        ws.cell(row=5, column=4, value=s["total"].get("fail", 0))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read(), f"summary_{start_s}_to_{end_s}.xlsx"

    if export_kind == "sku":
        path = SKU_SUMMARY_TEMPLATE_PATH
        if not os.path.isfile(path):
            raise FileNotFoundError(f"SKU Summary template not found: {path}")
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # Insert header row at the top
        ws.insert_rows(1)
        # Get fill color from A2 (original A1, now shifted down) before clearing it
        from openpyxl.styles import PatternFill
        a2_cell = ws.cell(row=2, column=1)
        a2_fill = a2_cell.fill.copy() if a2_cell.has_style and a2_cell.fill else None
        # Clear fill from A1 and A2
        a1_cell = ws.cell(row=1, column=1)
        if a1_cell.has_style:
            a1_cell.fill = PatternFill()
        if a2_cell.has_style:
            a2_cell.fill = PatternFill()
        # Fill from column B-D only, leave A1 empty
        header_cell = ws.cell(row=1, column=2, value=header_text)
        # Merge cells B1:D1 for header (A1 stays empty)
        ws.merge_cells('B1:D1')
        # Style header: bold, center alignment
        from openpyxl.styles import Font, Alignment
        header_cell.font = Font(bold=True, size=12)
        header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # Apply fill color from A2 (original A1) to header (B1:D1)
        if a2_fill:
            header_cell.fill = a2_fill
        # Auto-adjust column width based on header text length
        # Estimate: each character ~1.2 units, divide by 3 for B+C+D columns, add padding
        text_length = len(header_text)
        estimated_width = max(text_length * 1.2 / 3, 15)
        for col in ['B', 'C', 'D']:
            current_width = ws.column_dimensions[col].width if col in ws.column_dimensions and ws.column_dimensions[col].width else 0
            ws.column_dimensions[col].width = max(current_width, estimated_width)
        
        sku_rows = computed.get("sku_rows") or []
        # Template: row 2 = header (SKU, TESTED, PASS, FAIL), data rows start at row 3
        # (shifted down by 1 row due to header insertion)
        # Ensure header is set
        ws.cell(row=2, column=1, value="SKU")
        ws.cell(row=2, column=2, value="TESTED")
        ws.cell(row=2, column=3, value="PASS")
        ws.cell(row=2, column=4, value="FAIL")
        # Template has data rows 3-6 (4 rows). For additional SKUs beyond row 6, copy style from row 3
        first_data_row = 3
        template_last_data_row = 5
        for i, r in enumerate(sku_rows):
            row_num = first_data_row + i
            ws.cell(row=row_num, column=1, value=r.get("sku") or "")
            ws.cell(row=row_num, column=2, value=r.get("tested") or 0)
            ws.cell(row=row_num, column=3, value=r.get("pass") or 0)
            ws.cell(row=row_num, column=4, value=r.get("fail") or 0)
            # Copy style from template data row (row 2) to new rows beyond template
            if row_num > template_last_data_row:
                for col in range(1, 5):
                    src_cell = ws.cell(row=first_data_row, column=col)
                    tgt_cell = ws.cell(row=row_num, column=col)
                    _copy_cell_style(src_cell, tgt_cell)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read(), f"sku_{start_s}_to_{end_s}.xlsx"

    raise ValueError(f"Unsupported export_kind for XLSX: {export_kind}")


def _excel_text_cell(value: Any) -> str:
    """
    Make CSV cells safer for Excel auto-parsing.

    Excel often interprets strings like "183/66" as dates. To prevent that,
    emit an Excel text formula: ="183/66".
    """
    if value is None:
        return ""
    s = str(value)
    # Prevent scientific notation for long numeric IDs (e.g. SN).
    if re.fullmatch(r"\d{10,}", s):
        return f'="{s}"'
    # Common Excel auto-date patterns.
    if re.fullmatch(r"\d{1,4}/\d{1,4}", s) or re.fullmatch(r"\d{1,4}/\d{1,4}/\d{1,4}", s):
        return f'="{s}"'
    return s


def _fmt_ca_ms(ms: Optional[int]) -> str:
    if not ms:
        return ""
    try:
        dt = datetime.fromtimestamp(int(ms) / 1000, CA_TZ)
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ""


def _normalize_sn(v: Any) -> Optional[str]:
    """
    Normalize SN:
    - Must start with '18' and be 13 digits.
    - Excel may store it as float/scientific; coerce safely.
    """
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    # Handle scientific notation or floats like 1.830125000128E+12
    try:
        if re.fullmatch(r"\d+(\.\d+)?E\+\d+", s, flags=re.IGNORECASE):
            s = str(int(float(s)))
    except Exception:
        pass
    # Strip trailing .0
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    s = re.sub(r"[^\d]", "", s)
    if len(s) == 13 and s.startswith("18"):
        return s
    return None


def _extract_mmdd_entries(text: Any) -> List[str]:
    """
    Extract "entries" from a cell that may contain multiple mm/dd markers.
    Returns list of raw segments (strings) starting at each mm/dd marker.
    """
    if text is None:
        return []
    raw = str(text)
    if not raw.strip():
        return []
    # Find all mm/dd occurrences
    matches = list(re.finditer(r"\b\d{1,2}/\d{1,2}\b", raw))
    if not matches:
        return []
    out: List[str] = []
    for i, m in enumerate(matches):
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(raw)
        seg = raw[start:end].strip()
        if seg:
            out.append(seg)
    return out


def _load_bonepile_workbook(path: str):
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed; cannot read XLSX")
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def _hash_sheet_content(ws, max_rows: int = 10000) -> str:
    """
    Compute SHA256 hash of sheet content (first max_rows rows, all columns).
    Used to detect if sheet content changed since last parse.
    """
    h = hashlib.sha256()
    row_count = 0
    for row in ws.iter_rows(max_row=max_rows, values_only=True):
        if row_count >= max_rows:
            break
        # Convert row to string representation (normalize None to empty)
        row_str = "|".join(str(v if v is not None else "") for v in row)
        h.update(row_str.encode("utf-8"))
        h.update(b"\n")
        row_count += 1
    # Include row count in hash
    h.update(str(row_count).encode("utf-8"))
    return h.hexdigest()


def _find_header_row(ws, max_rows: int = 300) -> Optional[int]:
    """
    Return 1-based row index of header row containing 'SN' cell.
    """
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
        if not row:
            continue
        for v in row:
            if v is None:
                continue
            if str(v).strip().upper() == "SN":
                return i
    return None


def _read_header_map(ws, header_row: int, max_cols: int = 80) -> Dict[str, int]:
    """
    Build case-insensitive header -> 1-based column index map.
    """
    header_map: Dict[str, int] = {}
    for j, cell in enumerate(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)), start=1):
        if j > max_cols:
            break
        if cell is None:
            continue
        name = str(cell).strip()
        if not name:
            continue
        header_map[name.strip().upper()] = j
    return header_map


def _auto_mapping_from_headers(header_map: Dict[str, int]) -> Dict[str, int]:
    """
    Auto-map required fields by header names.
    """
    def pick(*names: str) -> Optional[int]:
        for n in names:
            idx = header_map.get(n.upper())
            if idx:
                return idx
        return None

    m: Dict[str, int] = {}
    m["sn"] = pick("SN") or 0
    m["nv_disposition"] = pick("NV DISPOSITION", "NV DISPO", "NV DISPOSITION ") or 0
    m["status"] = pick("STATUS") or 0
    m["pic"] = pick("PIC") or 0
    m["igs_action"] = pick("IGS ACTION") or 0
    m["igs_status"] = pick("IGS STATUS") or 0
    # Optional part number/SKU column (varies by file)
    m["nvpn"] = pick("NVPN", "PART NUMBER", "PART NUMBERS", "SKU") or 0
    return m


def _mapping_errors(mapping: Dict[str, int], header_map: Dict[str, int]) -> List[str]:
    errors: List[str] = []
    for k in BONEPILE_REQUIRED_FIELDS:
        if int(mapping.get(k) or 0) <= 0:
            errors.append(f"Missing column for '{k}'")
    if errors:
        # Provide context: first few headers
        sample = ", ".join(list(header_map.keys())[:25])
        errors.append(f"Available headers: {sample}")
    return errors


def _last_mmdd_entry(text: Any) -> str:
    """Return the last mm/dd entry segment from cell text, or full text if no mm/dd."""
    entries = _extract_mmdd_entries(text)
    if entries:
        return entries[-1].strip()
    return (str(text) if text is not None else "").strip()


def _is_pass_status(status_norm: str) -> bool:
    """
    Check if normalized status string indicates a pass status.
    Returns True if status contains: "PASS", "ALL PASS", "PASS ALL", or "PASSED" (case-insensitive).
    """
    if not status_norm:
        return False
    # Check for various pass patterns
    return (
        "PASS" in status_norm or
        "ALL PASS" in status_norm or
        "PASS ALL" in status_norm or
        "PASSED" in status_norm
    )


def _last_mmdd_only(text: Any) -> Optional[Tuple[int, int]]:
    """Return (month, day) from the last mm/dd in cell text, or None."""
    raw = (str(text) if text is not None else "").strip()
    if not raw:
        return None
    matches = list(re.finditer(r"\b(\d{1,2})/(\d{1,2})\b", raw))
    if not matches:
        return None
    m = matches[-1]
    try:
        month = int(m.group(1))
        day = int(m.group(2))
        if 1 <= month <= 12 and 1 <= day <= 31:
            return (month, day)
    except (ValueError, IndexError):
        pass
    return None


def _disposition_period_from_row(
    row: Dict[str, Any], aggregation: str, fallback_ca_ms: Optional[int] = None
) -> str:
    """
    Return period key from row: use last mm/dd in nv_disposition or igs_action (year from updated_at).
    Fallback to _disposition_period_from_ca_ms(updated_at) if no mm/dd.
    """
    ca_ms = row.get("updated_at_ca_ms") or fallback_ca_ms
    year = None
    if ca_ms is not None:
        try:
            dt = datetime.fromtimestamp(ca_ms / 1000.0, tz=CA_TZ)
            year = dt.year
        except Exception:
            pass
    mmdd_nv = _last_mmdd_only(row.get("nv_disposition"))
    mmdd_igs = _last_mmdd_only(row.get("igs_action"))
    mmdd = mmdd_igs or mmdd_nv
    if mmdd is not None and year is not None:
        try:
            d = date(year, mmdd[0], mmdd[1])
            if aggregation == "monthly":
                return d.strftime("%Y-%m")
            if aggregation == "weekly":
                days_since_sunday = (d.weekday() + 1) % 7
                week_start = d - timedelta(days=days_since_sunday)
                week_end = week_start + timedelta(days=6)
                return f"{week_start.strftime('%Y-%m-%d')}~{week_end.strftime('%Y-%m-%d')}"
            return d.strftime("%Y-%m-%d")
        except (ValueError, TypeError):
            pass
    return _disposition_period_from_ca_ms(ca_ms, aggregation)


def _disposition_period_from_ca_ms(ca_ms: Optional[int], aggregation: str) -> str:
    """Return period key (ca_date, ca_week, or ca_month) from updated_at_ca_ms."""
    if ca_ms is None:
        return ""
    try:
        dt = datetime.fromtimestamp(ca_ms / 1000.0, tz=CA_TZ)
    except Exception:
        return ""
    if aggregation == "weekly":
        days_since_sunday = (dt.weekday() + 1) % 7
        week_start = (dt - timedelta(days=days_since_sunday)).date()
        week_end = week_start + timedelta(days=6)
        return f"{week_start.strftime('%Y-%m-%d')}~{week_end.strftime('%Y-%m-%d')}"
    if aggregation == "monthly":
        return dt.strftime("%Y-%m")
    return dt.strftime("%Y-%m-%d")


def compute_disposition_stats(aggregation: str = "daily", start_ca_ms: Optional[int] = None, end_ca_ms: Optional[int] = None) -> Dict[str, Any]:
    """
    Compute NV Disposition stats from bonepile_entries.
    
    Logic:
    1. Total Dispositions: Đếm SNs có NV disposition date (mm/dd cuối cùng) trong date range
    2. Waiting IGS: Status=FAIL, PIC=IGS, mm/dd từ NV disposition cuối cùng trong date range
    3. By Period:
       - Total: Phân loại SNs theo mm/dd từ NV disposition cuối cùng = ngày của header period
       - Waiting IGS: Phân loại SNs theo mm/dd từ IGS action cuối cùng = ngày của header period
    
    Returns: summary { total, waiting_igs, complete }, by_sku [ { sku, total, waiting_igs, complete } ], by_period [ { period, total, waiting_igs, complete } ].
    """
    conn = connect_db()
    try:
        # Get ALL rows (no date filter on SQL level)
        query = "SELECT sn, nvpn, status, pic, nv_disposition, igs_action, updated_at_ca_ms FROM bonepile_entries;"
        rows = conn.execute(query).fetchall()
    finally:
        conn.close()

    def _norm(s: Any) -> str:
        return (str(s) if s is not None else "").strip().upper()

    def _row_dict(r) -> Dict[str, Any]:
        return {k: r[k] for k in r.keys()} if hasattr(r, "keys") else dict(r)

    if start_ca_ms is not None and end_ca_ms is not None:
        start_d = datetime.fromtimestamp(start_ca_ms / 1000.0, tz=CA_TZ).date()
        end_d = datetime.fromtimestamp(end_ca_ms / 1000.0, tz=CA_TZ).date()
        year = start_d.year  # Use year from date range
    else:
        start_d = None
        end_d = None
        year = datetime.now(CA_TZ).year

    def _date_to_period(d: date) -> str:
        if aggregation == "monthly":
            return d.strftime("%Y-%m")
        elif aggregation == "weekly":
            days_since_sunday = (d.weekday() + 1) % 7
            week_start = d - timedelta(days=days_since_sunday)
            week_end = week_start + timedelta(days=6)
            return f"{week_start.strftime('%Y-%m-%d')}~{week_end.strftime('%Y-%m-%d')}"
        else:
            return d.strftime("%Y-%m-%d")

    # Per SN, keep latest row
    sn_latest: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        rd = _row_dict(r)
        sn = (rd.get("sn") or "").strip()
        if not sn:
            continue
        ca_ms = rd.get("updated_at_ca_ms")
        existing = sn_latest.get(sn)
        if existing is None or (ca_ms or 0) > (existing.get("updated_at_ca_ms") or 0):
            sn_latest[sn] = {
                "row": r,
                "rd": rd,
                "updated_at_ca_ms": ca_ms,
            }
    
    # Step 1: Count Total Dispositions - SNs có NV disposition date trong date range
    total_sns: Dict[str, Dict[str, Any]] = {}
    for sn, data in sn_latest.items():
        rd = data["rd"]
        nv_dispo_text = rd.get("nv_disposition")
        if not nv_dispo_text:
            continue
        
        # Get mm/dd cuối cùng từ NV disposition
        mmdd_nv = _last_mmdd_only(nv_dispo_text)
        if mmdd_nv is None:
            continue
        
        # Build date from mm/dd
        try:
            nv_date = date(year, mmdd_nv[0], mmdd_nv[1])
            # If date seems too old, try next year
            if start_d and nv_date < start_d - timedelta(days=60):
                nv_date = date(year + 1, mmdd_nv[0], mmdd_nv[1])
        except (ValueError, TypeError):
            continue
        
        # Check if date is in range
        if start_d is not None and end_d is not None:
            if not (start_d <= nv_date <= end_d):
                continue
        
        sku = (rd.get("nvpn") or "").strip() or "Unknown"
        period = _date_to_period(nv_date)
        
        total_sns[sn] = {
            "sku": sku,
            "nv_date": nv_date,
            "period": period,
        }

    # Step 2: Count Waiting IGS - SNs với Status=FAIL, PIC=IGS, mm/dd cuối cùng trong date range
    waiting_sns: Dict[str, Dict[str, Any]] = {}
    for sn, data in sn_latest.items():
        r = data["row"]
        rd = data["rd"]
        
        # Check Waiting IGS criteria: Status=FAIL, PIC=IGS
        status_ok = _norm(rd.get("status")) == "FAIL"
        pic_ok = _norm(rd.get("pic")) == "IGS"
        if not (status_ok and pic_ok):
            continue
        
        # Get mm/dd cuối cùng từ NV disposition
        mmdd_nv = _last_mmdd_only(rd.get("nv_disposition"))
        if mmdd_nv is None:
            continue
        
        # Build date from mm/dd
        try:
            nv_date = date(year, mmdd_nv[0], mmdd_nv[1])
            # If date seems too old, try next year
            if start_d and nv_date < start_d - timedelta(days=60):
                nv_date = date(year + 1, mmdd_nv[0], mmdd_nv[1])
        except (ValueError, TypeError):
            continue
        
        # Check if date is in range
        if start_d is not None and end_d is not None:
            if not (start_d <= nv_date <= end_d):
                continue
        
        # Get mm/dd từ IGS action cuối cùng (for period classification)
        mmdd_igs = _last_mmdd_only(rd.get("igs_action"))
        igs_date = None
        if mmdd_igs is not None:
            try:
                igs_date = date(year, mmdd_igs[0], mmdd_igs[1])
                if start_d and igs_date < start_d - timedelta(days=60):
                    igs_date = date(year + 1, mmdd_igs[0], mmdd_igs[1])
            except (ValueError, TypeError):
                pass
        
        sku = (rd.get("nvpn") or "").strip() or "Unknown"
        period_nv = _date_to_period(nv_date)
        period_igs = _date_to_period(igs_date) if igs_date else None
        
        waiting_sns[sn] = {
            "sku": sku,
            "nv_date": nv_date,
            "igs_date": igs_date,
            "period_nv": period_nv,
            "period_igs": period_igs,
        }

    # Step 3: Build by_period and by_sku
    by_period: Dict[str, Dict[str, Any]] = {}
    by_sku: Dict[str, Dict[str, Any]] = {}
    
    # Total dispositions by period (from SNs)
    for sn, info in total_sns.items():
        period = info["period"]
        by_period.setdefault(period, {"period": period, "total": 0, "waiting_igs": 0, "complete": 0})
        by_period[period]["total"] += 1
        
        sku = info["sku"]
        by_sku.setdefault(sku, {"sku": sku, "total": 0, "waiting_igs": 0, "complete": 0})
        by_sku[sku]["total"] += 1
    
    # Waiting IGS by period (from SNs)
    for sn, info in waiting_sns.items():
        waiting_period = info["period_igs"] if info["period_igs"] else info["period_nv"]
        by_period.setdefault(waiting_period, {"period": waiting_period, "total": 0, "waiting_igs": 0, "complete": 0})
        by_period[waiting_period]["waiting_igs"] += 1
        
        sku = info["sku"]
        by_sku.setdefault(sku, {"sku": sku, "total": 0, "waiting_igs": 0, "complete": 0})
        by_sku[sku]["waiting_igs"] += 1

    # Calculate complete
    for sku, d in by_sku.items():
        d["complete"] = d["total"] - d.get("waiting_igs", 0)
    for period, d in by_period.items():
        d["complete"] = d["total"] - d.get("waiting_igs", 0)

    # Summary
    summary_total = len(total_sns)
    summary_waiting = len(waiting_sns)
    summary_complete = summary_total - summary_waiting
    
    summary = {
        "total": summary_total,
        "waiting_igs": summary_waiting,
        "complete": summary_complete
    }

    # Count unique trays (SNs) in BP and trays with ALL PASS status
    # IMPORTANT: Count from ALL rows in bonepile_entries (not filtered by date range)
    # This gives total trays in the excel file, not just in the filter range
    conn_all = connect_db()
    try:
        all_rows = conn_all.execute(
            "SELECT sn, nvpn, status, updated_at_ca_ms FROM bonepile_entries;"
        ).fetchall()
    finally:
        conn_all.close()
    
    # For each SN, keep the latest row (by updated_at_ca_ms) to determine status
    sn_latest_row: Dict[str, Dict[str, Any]] = {}
    for r in all_rows:
        sn = (r["sn"] or "").strip()
        if not sn:
            continue
        ca_ms = r["updated_at_ca_ms"] or 0
        existing = sn_latest_row.get(sn)
        if existing is None or ca_ms > (existing.get("updated_at_ca_ms") or 0):
            sn_latest_row[sn] = {
                "sn": sn,
                "nvpn": (r["nvpn"] or "").strip() or "Unknown",
                "status": (r["status"] or "").strip(),
                "updated_at_ca_ms": ca_ms,
            }
    
    unique_trays_bp = len(sn_latest_row)
    all_pass_trays = 0
    tray_by_sku: Dict[str, Dict[str, int]] = {}  # {sku: {total_trays: X, all_pass_trays: Y}}
    
    for sn, d in sn_latest_row.items():
        status_norm = _norm(d["status"])
        if _is_pass_status(status_norm):
            all_pass_trays += 1
        
        sku = d["nvpn"]
        tray_by_sku.setdefault(sku, {"sku": sku, "total_trays": 0, "all_pass_trays": 0})
        tray_by_sku[sku]["total_trays"] += 1
        if _is_pass_status(status_norm):
            tray_by_sku[sku]["all_pass_trays"] += 1
    
    summary["unique_trays_bp"] = unique_trays_bp
    summary["all_pass_trays"] = all_pass_trays
    tray_by_sku_list = sorted(tray_by_sku.values(), key=lambda x: (x["sku"]))

    by_sku_list = sorted(by_sku.values(), key=lambda x: (x["sku"]))
    by_period_list = sorted(by_period.values(), key=lambda x: (x["period"]))

    # Filter by_period to only include periods within user's date range
    if start_ca_ms is not None and end_ca_ms is not None:
        try:
            start_d = datetime.fromtimestamp(start_ca_ms / 1000.0, tz=CA_TZ).date()
            end_d = datetime.fromtimestamp(end_ca_ms / 1000.0, tz=CA_TZ).date()
            filtered = []
            for p in by_period_list:
                period_str = p.get("period") or ""
                if aggregation == "daily" and re.match(r"^\d{4}-\d{2}-\d{2}$", period_str):
                    pd = datetime.strptime(period_str, "%Y-%m-%d").date()
                    if start_d <= pd <= end_d:
                        filtered.append(p)
                elif aggregation == "monthly" and re.match(r"^\d{4}-\d{2}$", period_str):
                    pd = datetime.strptime(period_str + "-01", "%Y-%m-%d").date()
                    if start_d <= pd <= end_d:
                        filtered.append(p)
                elif aggregation == "weekly" and "~" in period_str:
                    part = period_str.split("~")[0]
                    if re.match(r"^\d{4}-\d{2}-\d{2}$", part):
                        pd = datetime.strptime(part, "%Y-%m-%d").date()
                        if pd <= end_d and (pd + timedelta(days=6)) >= start_d:
                            filtered.append(p)
                else:
                    filtered.append(p)
            by_period_list = filtered
        except Exception:
            pass

    return {"summary": summary, "by_sku": by_sku_list, "by_period": by_period_list, "tray_by_sku": tray_by_sku_list}


def compute_disposition_sn_list(
    metric: str, sku: Optional[str] = None, period: Optional[str] = None, aggregation: str = "daily", start_ca_ms: Optional[int] = None, end_ca_ms: Optional[int] = None
) -> List[Dict[str, Any]]:
    """
    Return list of { sn, last_nv_dispo, last_igs_action, nvpn, status, pic } for drill-down.
    Uses same logic as compute_disposition_stats: find KPI SNs first, then filter by period/sku.
    metric: total | waiting | complete | trays_bp | all_pass_trays
    """
    def _norm(s: Any) -> str:
        return (str(s) if s is not None else "").strip().upper()

    def _row_dict(r) -> Dict[str, Any]:
        return {k: r[k] for k in r.keys()} if hasattr(r, "keys") else dict(r)

    # Special handling for trays_bp and all_pass_trays: get ALL rows (not filtered by date range)
    if metric in ("trays_bp", "all_pass_trays"):
        conn = connect_db()
        try:
            rows = conn.execute(
                "SELECT sheet, excel_row, sn, nvpn, status, pic, nv_disposition, igs_action, updated_at_ca_ms FROM bonepile_entries;"
            ).fetchall()
        finally:
            conn.close()
        
        # Per SN keep latest row
        sn_rows: Dict[str, Dict[str, Any]] = {}
        for r in rows:
            rd = _row_dict(r)
            sn = (rd.get("sn") or "").strip()
            if not sn:
                continue
            ca_ms = rd.get("updated_at_ca_ms")
            existing = sn_rows.get(sn)
            if existing is None or (ca_ms or 0) > (existing.get("updated_at_ca_ms") or 0):
                sn_rows[sn] = {
                    "sn": sn,
                    "nvpn": (rd.get("nvpn") or "").strip() or "Unknown",
                    "status": (rd.get("status") or "").strip(),
                    "pic": (rd.get("pic") or "").strip(),
                    "nv_disposition": rd.get("nv_disposition"),
                    "igs_action": rd.get("igs_action"),
                    "updated_at_ca_ms": ca_ms,
                }
        
        # Filter by metric
        out: List[Dict[str, Any]] = []
        for sn, d in sn_rows.items():
            status_norm = _norm(d.get("status"))
            row_sku = d.get("nvpn")
            if sku and sku != "__TOTAL__" and row_sku != sku:
                continue
            if metric == "all_pass_trays" and not _is_pass_status(status_norm):
                continue
            # trays_bp: include all
            out.append({
                "sn": sn,
                "last_nv_dispo": _last_mmdd_entry(d.get("nv_disposition")),
                "last_igs_action": _last_mmdd_entry(d.get("igs_action")),
                "nvpn": row_sku,
                "status": d.get("status"),
                "pic": d.get("pic"),
            })
        out.sort(key=lambda x: (x["sn"]))
        return out

    # For waiting/complete/total: use same logic as compute_disposition_stats
    conn = connect_db()
    try:
        rows = conn.execute(
            "SELECT sheet, excel_row, sn, nvpn, status, pic, nv_disposition, igs_action, updated_at_ca_ms FROM bonepile_entries;"
        ).fetchall()
    finally:
        conn.close()

    # Step 1: Get latest row per SN
    sn_data: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        rd = _row_dict(r)
        sn = (rd.get("sn") or "").strip()
        if not sn:
            continue
        ca_ms = rd.get("updated_at_ca_ms")
        existing = sn_data.get(sn)
        if existing is None or (ca_ms or 0) > (existing.get("updated_at_ca_ms") or 0):
            sn_data[sn] = {
                "row": r,
                "rd": rd,
                "updated_at_ca_ms": ca_ms,
            }

    # Step 2: Find KPI SNs (Status=FAIL, PIC=IGS, mm/dd from NV disposition in date range)
    if start_ca_ms is not None and end_ca_ms is not None:
        start_d = datetime.fromtimestamp(start_ca_ms / 1000.0, tz=CA_TZ).date()
        end_d = datetime.fromtimestamp(end_ca_ms / 1000.0, tz=CA_TZ).date()
    else:
        start_d = None
        end_d = None

    kpi_sns: Dict[str, Dict[str, Any]] = {}
    for sn, data in sn_data.items():
        r = data["row"]
        rd = data["rd"]
        
        status_ok = _norm(rd.get("status")) == "FAIL"
        pic_ok = _norm(rd.get("pic")) == "IGS"
        
        # For "waiting" metric, must match KPI criteria
        if metric == "waiting" and not (status_ok and pic_ok):
            continue
        # For "complete" metric, must NOT match waiting criteria
        if metric == "complete" and (status_ok and pic_ok):
            continue
        
        mmdd_nv = _last_mmdd_only(rd.get("nv_disposition"))
        if mmdd_nv is None:
            continue
        
        # Determine year for mm/dd: use year from date range if available, otherwise from updated_at
        ca_ms = data["updated_at_ca_ms"]
        year = None
        if start_d is not None:
            # Use year from start date range (most reliable)
            year = start_d.year
        elif ca_ms is not None:
            try:
                dt = datetime.fromtimestamp(ca_ms / 1000.0, tz=CA_TZ)
                year = dt.year
            except Exception:
                pass
        
        if year is None:
            continue
        
        try:
            nv_date = date(year, mmdd_nv[0], mmdd_nv[1])
        except (ValueError, TypeError):
            continue
        
        # If date is in the past compared to updated_at, try next year
        if ca_ms is not None:
            try:
                dt = datetime.fromtimestamp(ca_ms / 1000.0, tz=CA_TZ)
                updated_date = dt.date()
                if nv_date < updated_date - timedelta(days=30):
                    # If mm/dd is more than 30 days before updated_at, likely next year
                    try:
                        nv_date_next_year = date(year + 1, mmdd_nv[0], mmdd_nv[1])
                        if start_d is None or (start_d <= nv_date_next_year <= end_d):
                            nv_date = nv_date_next_year
                            year = year + 1
                    except (ValueError, TypeError):
                        pass
            except Exception:
                pass
        
        # Check date range
        if start_d is not None and end_d is not None:
            if not (start_d <= nv_date <= end_d):
                continue
        
        # Calculate periods
        def _date_to_period(d: date) -> str:
            if aggregation == "monthly":
                return d.strftime("%Y-%m")
            elif aggregation == "weekly":
                days_since_sunday = (d.weekday() + 1) % 7
                week_start = d - timedelta(days=days_since_sunday)
                week_end = week_start + timedelta(days=6)
                return f"{week_start.strftime('%Y-%m-%d')}~{week_end.strftime('%Y-%m-%d')}"
            else:
                return d.strftime("%Y-%m-%d")
        
        period_nv = _date_to_period(nv_date)
        mmdd_igs = _last_mmdd_only(rd.get("igs_action"))
        period_igs = None
        if mmdd_igs is not None:
            try:
                igs_date = date(year, mmdd_igs[0], mmdd_igs[1])
                # If IGS date is in the past compared to NV date, try next year
                if igs_date < nv_date - timedelta(days=30):
                    try:
                        igs_date = date(year + 1, mmdd_igs[0], mmdd_igs[1])
                    except (ValueError, TypeError):
                        pass
                period_igs = _date_to_period(igs_date)
            except (ValueError, TypeError):
                pass
        
        # Filter by period (for waiting: use IGS period, for total: use NV period)
        if period and period != "__TOTAL__":
            if metric == "waiting":
                if period_igs and period_igs != period:
                    continue
                elif not period_igs and period_nv != period:
                    continue
            else:
                if period_nv != period:
                    continue
        
        # Filter by SKU
        row_sku = (r["nvpn"] or "").strip() or "Unknown"
        if sku and sku != "__TOTAL__" and row_sku != sku:
            continue
        
        kpi_sns[sn] = {
            "row": r,
            "rd": rd,
            "nvpn": row_sku,
        }

    # Step 3: Build output
    out: List[Dict[str, Any]] = []
    for sn, info in kpi_sns.items():
        r = info["row"]
        rd = info["rd"]
        out.append({
            "sn": sn,
            "last_nv_dispo": _last_mmdd_entry(rd.get("nv_disposition")),
            "last_igs_action": _last_mmdd_entry(rd.get("igs_action")),
            "nvpn": info["nvpn"],
            "status": (rd.get("status") or "").strip(),
            "pic": (rd.get("pic") or "").strip(),
        })
    out.sort(key=lambda x: (x["sn"]))
    return out


def _parse_range_from_payload(payload: Dict[str, Any]) -> Tuple[Optional[Response], Optional[datetime], Optional[datetime]]:
    start_dt = payload.get("start_datetime")
    end_dt = payload.get("end_datetime")
    if not start_dt or not end_dt:
        return jsonify({"error": "start_datetime and end_datetime required"}), None, None
    start_ca = _parse_ca_input_datetime(start_dt, is_end=False)
    end_ca = _parse_ca_input_datetime(end_dt, is_end=True)
    if not start_ca or not end_ca:
        return jsonify({"error": "datetime format must be YYYY-MM-DD HH:MM (optional :SS)"}), None, None

    now_ca = datetime.now(CA_TZ).replace(microsecond=0)
    if end_ca > now_ca:
        end_ca = now_ca
    if start_ca > now_ca:
        return jsonify({"error": "start is in the future"}), None, None
    if end_ca <= start_ca:
        return jsonify({"error": "end must be after start"}), None, None
    return None, start_ca, end_ca


@app.route("/api/export", methods=["POST"])
def api_export():
    """
    Export CSV for a specific dashboard container or the whole dashboard.

    Payload:
      - start_datetime: "YYYY-MM-DD HH:MM"
      - end_datetime: "YYYY-MM-DD HH:MM"
      - aggregation: daily|weekly|monthly (used for breakdown)
      - export: one of: summary|sku|breakdown|test_flow|dashboard
    """
    payload = request.json or {}
    export_kind = (payload.get("export") or "dashboard").strip().lower()
    aggregation = (payload.get("aggregation") or "daily").strip().lower()
    if aggregation not in ("daily", "weekly", "monthly"):
        aggregation = "daily"

    err, start_ca, end_ca = _parse_range_from_payload(payload)
    if err is not None:
        return err
    assert start_ca is not None and end_ca is not None

    rows = query_entries_in_range(start_ca, end_ca)
    computed = compute_stats(rows, aggregation=aggregation)
    test_flow = compute_test_flow(rows)
    details = compute_sn_details(rows)

    start_s = start_ca.strftime("%Y%m%d_%H%M")
    end_s = end_ca.strftime("%Y%m%d_%H%M")
    export_format = (payload.get("format") or "csv").strip().lower()

    # Handle XLSX export for summary and sku
    if export_format == "xlsx" and export_kind in ("summary", "sku"):
        try:
            data_xlsx, filename = _build_export_xlsx(export_kind, computed, start_s, end_s, start_ca, end_ca)
            return _xlsx_response(data_xlsx, filename)
        except FileNotFoundError as e:
            return jsonify({"error": str(e)}), 404
        except Exception as e:
            return jsonify({"error": f"XLSX export failed: {str(e)}"}), 500

    # Build CSV
    out = io.StringIO()
    w = csv.writer(out, lineterminator="\n")

    def write_summary():
        w.writerow(["metric", "bp", "fresh", "total"])
        s = computed["summary"]
        for m in ("tested", "pass", "fail"):
            w.writerow([m, s["bp"][m], s["fresh"][m], s["total"][m]])

    def write_sku():
        w.writerow(["sku", "tested", "pass", "fail"])
        for r in computed["sku_rows"]:
            w.writerow([r.get("sku"), r.get("tested"), r.get("pass"), r.get("fail")])

    def write_breakdown():
        w.writerow(["period", "tested", "passed", "bonepile", "fresh", "pass_rate"])
        for r in computed["breakdown_rows"]:
            w.writerow(
                [
                    r.get("period"),
                    r.get("tested"),
                    r.get("passed"),
                    r.get("bonepile"),
                    r.get("fresh"),
                    f"{float(r.get('pass_rate') or 0.0):.4f}",
                ]
            )

    def write_test_flow():
        stations = test_flow.get("stations") or []
        w.writerow(["ts", "sku"] + list(stations))
        totals = test_flow.get("totals") or {}
        w.writerow(
            ["-", "TOTAL"]
            + [
                _excel_text_cell(f"{(totals.get(st) or {}).get('pass', 0)}/{(totals.get(st) or {}).get('fail', 0)}")
                for st in stations
            ]
        )
        for r in (test_flow.get("rows") or []):
            row_vals = [r.get("ts"), r.get("sku")]
            st_map = r.get("stations") or {}
            for st in stations:
                cell = st_map.get(st) or {}
                row_vals.append(_excel_text_cell(f"{cell.get('pass', 0)}/{cell.get('fail', 0)}"))
            w.writerow(row_vals)

    if export_kind == "summary":
        write_summary()
        return _csv_response(out.getvalue(), f"summary_{start_s}_to_{end_s}.csv")
    if export_kind == "sku":
        write_sku()
        return _csv_response(out.getvalue(), f"sku_{start_s}_to_{end_s}.csv")
    if export_kind == "breakdown":
        write_breakdown()
        return _csv_response(out.getvalue(), f"breakdown_{aggregation}_{start_s}_to_{end_s}.csv")
    if export_kind in ("test_flow", "testflow"):
        write_test_flow()
        return _csv_response(out.getvalue(), f"test_flow_{start_s}_to_{end_s}.csv")
    if export_kind == "disposition_summary":
        start_ca_ms = utc_ms(start_ca)
        end_ca_ms = utc_ms(end_ca)
        dispo_data = compute_disposition_stats(aggregation="daily", start_ca_ms=start_ca_ms, end_ca_ms=end_ca_ms)
        w.writerow(["metric", "value"])
        summary = dispo_data.get("summary", {})
        w.writerow(["Total Dispositions", summary.get("total", 0)])
        w.writerow(["Waiting IGS", summary.get("waiting_igs", 0)])
        w.writerow(["Complete", summary.get("complete", 0)])
        return _csv_response(out.getvalue(), f"disposition_summary_{start_s}_to_{end_s}.csv")

    # dashboard (single table, per-SN; matches user's requested "one table only")
    # Columns are designed for Excel-friendly review.
    sn_rows: Dict[str, List[sqlite3.Row]] = {}
    for r in rows:
        sn_rows.setdefault(r["sn"], []).append(r)

    # Map details by SN for extra columns
    by_sn = {d.get("sn"): d for d in (details or []) if d.get("sn")}

    station_order = ["FLA", "FLB", "AST", "FTS", "FCT", "RIN", "NVL"]

    w.writerow(
        [
            "SN",
            "Bonepile",
            "Pass/Fail",
            "Part Numbers",
            "Stations",
            "Test Count",
            "Pass Count",
            "Fail Count",
            "First Seen (CA)",
            "Last Seen (CA)",
            "Last Station",
            "Last Folder ID",
            "Last Filename",
            "Last Final Pass Time (CA)",
            "Last Fail Time (CA)",
        ]
    )

    for sn in sorted(sn_rows.keys()):
        tests = sn_rows.get(sn) or []
        # Status counts
        p_cnt = 0
        f_cnt = 0
        is_bp = False
        stations_seen: set = set()
        part_seen: set = set()
        min_ca = None
        max_ca = None
        for t in tests:
            st = str(t["station"] or "").strip().upper()
            if st:
                stations_seen.add(st)
            pn = t["part_number"] or ""
            if pn:
                part_seen.add(pn)
            status = str(t["status"] or "").strip().upper()
            if status == "P":
                p_cnt += 1
            elif status == "F":
                f_cnt += 1
            if (t["is_bonepile"] or 0) == 1:
                is_bp = True
            try:
                ca_ms = int(t["ca_ms"])
                if min_ca is None or ca_ms < min_ca:
                    min_ca = ca_ms
                if max_ca is None or ca_ms > max_ca:
                    max_ca = ca_ms
            except Exception:
                pass

        # Order stations nicely
        ordered = [s for s in station_order if s in stations_seen]
        extras = sorted([s for s in stations_seen if s not in station_order])
        stations_txt = ", ".join(ordered + extras)

        d = by_sn.get(sn) or {}
        last_part = d.get("last_part_number") or ""
        # Prefer latest part number for the main column, but keep multiple if needed
        if last_part and len(part_seen) > 1:
            parts_txt = "; ".join([last_part] + sorted([p for p in part_seen if p != last_part]))
        elif last_part:
            parts_txt = last_part
        else:
            parts_txt = "; ".join(sorted(part_seen))

        # PASS/FAIL per dashboard: align with our UI "is_pass" definition (final pass)
        passfail = "PASS" if int(d.get("is_pass") or 0) == 1 else "FAIL"

        w.writerow(
            [
                _excel_text_cell(sn),
                "Yes" if is_bp else "No",
                passfail,
                parts_txt,
                stations_txt,
                len(tests),
                p_cnt,
                f_cnt,
                _fmt_ca_ms(min_ca),
                _fmt_ca_ms(max_ca),
                d.get("last_station") or "",
                d.get("last_folder_id") or "",
                d.get("last_filename") or "",
                _fmt_ca_ms(d.get("pass_ca_ms")),
                _fmt_ca_ms(d.get("fail_ca_ms")),
            ]
        )

    return _csv_response(out.getvalue(), f"dashboard_{start_s}_to_{end_s}.csv")


@app.route("/api/sn-list", methods=["POST"])
def api_sn_list():
    """
    Drill-down list of SNs for a metric bucket.

    Payload:
      - start_datetime: "YYYY-MM-DD HH:MM"
      - end_datetime: "YYYY-MM-DD HH:MM"
      - segment: "bp" | "fresh" | "total"
      - metric: "tested" | "pass" | "fail"
    """
    payload = request.json or {}
    start_dt = payload.get("start_datetime")
    end_dt = payload.get("end_datetime")
    segment = (payload.get("segment") or "total").strip().lower()
    metric = (payload.get("metric") or "tested").strip().lower()
    sku = payload.get("sku")  # optional: filter by last part_number
    period = payload.get("period")  # optional: filter by bucket value (daily/weekly/monthly)
    aggregation = (payload.get("aggregation") or "").strip().lower()
    station = payload.get("station")  # optional: station drilldown
    station_outcome = (payload.get("station_outcome") or "").strip().lower()  # pass/fail/both

    if segment not in ("bp", "fresh", "total"):
        segment = "total"
    if metric not in ("tested", "pass", "fail"):
        metric = "tested"

    if not start_dt or not end_dt:
        return jsonify({"error": "start_datetime and end_datetime required"}), 400

    start_ca = _parse_ca_input_datetime(start_dt, is_end=False)
    end_ca = _parse_ca_input_datetime(end_dt, is_end=True)
    if not start_ca or not end_ca:
        return jsonify({"error": "datetime format must be YYYY-MM-DD HH:MM (optional :SS)"}), 400

    now_ca = datetime.now(CA_TZ).replace(microsecond=0)
    if end_ca > now_ca:
        end_ca = now_ca
    if end_ca <= start_ca:
        return jsonify({"error": "end must be after start"}), 400

    rows = query_entries_in_range(start_ca, end_ca)

    # Optional: filter rows by requested bucket (time breakdown drilldown)
    if period and aggregation in ("daily", "weekly", "monthly"):
        key_field = {"daily": "ca_date", "weekly": "ca_week", "monthly": "ca_month"}[aggregation]
        try:
            rows = [r for r in rows if (r[key_field] == period)]
        except Exception:
            pass

    if station and station_outcome in ("pass", "fail"):
        details = compute_station_sn_list(rows, station=station, outcome=station_outcome, sku=sku)
    elif station and station_outcome == "both":
        details = compute_station_sn_list_both(rows, station=station, sku=sku)
    else:
        details = compute_sn_details(rows)

    # Optional: SKU filter (match the SN's latest part number in the selected slice)
    if sku:
        sku_norm = str(sku).strip()
        details = [d for d in details if (d.get("last_part_number") or "") == sku_norm]

    # Segment filter
    if segment == "bp":
        details = [d for d in details if d.get("is_bonepile") == 1]
    elif segment == "fresh":
        details = [d for d in details if d.get("is_bonepile") == 0]

    # Metric filter
    if metric == "pass":
        details = [d for d in details if d.get("is_pass") == 1]
    elif metric == "fail":
        details = [d for d in details if d.get("is_pass") == 0]

    return jsonify(
        {
            "segment": segment,
            "metric": metric,
            "sku": sku,
            "period": period,
            "aggregation": aggregation,
            "count": len(details),
            "rows": details[:5000],
        }
    )


# -----------------------------
# Background auto-scan
# -----------------------------


def auto_scan_loop():
    global next_auto_scan_ms
    # Run retention cleanup at most every 12 hours
    last_cleanup_time = 0.0
    while True:
        loop_started = time.time()
        try:
            now_ca = datetime.now(CA_TZ).replace(microsecond=0)
            # Refresh window: last N minutes (CA). We delete cache in this range then rescan so data is always fresh.
            start_ca = now_ca - timedelta(minutes=REFRESH_WINDOW_MINUTES)
            cutoff_ms = utc_ms(start_ca)

            with scan_lock:
                state = RawState.load()
                ensure_db_ready()
                # Delete raw_entries in the refresh window so rescan accepts all data in that window (fresh).
                try:
                    conn = connect_db()
                    try:
                        conn.execute("DELETE FROM raw_entries WHERE ca_ms >= ?", (cutoff_ms,))
                        conn.commit()
                    finally:
                        conn.close()
                    # Rescan last REFRESH_WINDOW_MINUTES and insert; scan_range updates state from DB after insert.
                    scan_range(start_ca, now_ca, state)
                except Exception:
                    pass

            # Retention cleanup (best-effort)
            if (loop_started - last_cleanup_time) >= (12 * 60 * 60):
                cleanup_retention(now_ca=now_ca)
                last_cleanup_time = loop_started
        except Exception:
            pass

        # Always pause full interval between runs (prevents "scan nonstop" when scan work > interval).
        with auto_status_lock:
            next_auto_scan_ms = int((time.time() + AUTO_SCAN_EVERY_SECONDS) * 1000)
        time.sleep(float(AUTO_SCAN_EVERY_SECONDS))


def main():
    ensure_db_ready(force=True)

    t = threading.Thread(target=auto_scan_loop, daemon=True)
    t.start()

    app.run(host="0.0.0.0", port=5555, debug=False)


if __name__ == "__main__":
    main()

